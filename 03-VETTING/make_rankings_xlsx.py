#!/usr/bin/env python3
"""
Build a polished, candidate-relative .xlsx from a job-vetting rankings CSV.

Usage:
    python make_rankings_xlsx.py <input-rankings.csv> [output.xlsx] \
        [--config jail.config.json] [--quarantined N]

Candidate-relative coloring (V2 Unit 5): comp and location are colored against the
candidate's own preferences in jail.config.json (comp target/floor; per-arrangement
location ratings), NOT hardcoded thresholds. If the config is missing or partial, those
columns fall back to NEUTRAL (grey) with an "Unknown / no prefs" label — never a wrong
guess. New columns make the fit legible: "Comp Fit", "Location Fit", "Lane Fit".

- final_score is colored to match the Status bands (>=80 / 70-79 / 60-69 / <60).
- The four sub-scores keep the fine-grained 0-100 bucket ramp.
- "Job Post Title + Link" cells are real clickable hyperlinks.
- Backward-compatible with legacy CSV field names (passion_score / quality_of_life_score).
- --quarantined N adds a note that N thin/failed prep posts were not ranked.
"""
import argparse
import csv
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    HAVE_OPENPYXL = True
except ImportError:  # the pure fit/config logic below is importable + testable without openpyxl
    HAVE_OPENPYXL = False

FONT = "Arial"

# ---- Output column layout: (header, source-CSV-header or None) ----
# None = computed here (Comp/Location Fit) or a blank manual column.
COLUMNS = [
    ("Applied",                 None),
    ("Status?",                 "Status?"),
    ("Category",                "Category"),
    ("Company",                 "Company"),
    ("Job Post Title + Link",   "Job Post Title + Link"),
    ("Location?",               "Location?"),
    ("Location Fit",            None),   # computed from Location? + config
    ("Comp Range",              "Comp Range"),
    ("Comp Fit",                None),   # computed from Comp Range + config
    ("Have Intro?",             None),
    ("Decline/Down Date?",      None),
    ("Notes",                   None),
    ("lane",                    "lane"),
    ("Lane Fit",                "Lane Fit"),   # from the scorer (vet-jobs writes it)
    ("desire_score",            "desire_score"),
    ("market_perception_score", "market_perception_score"),
    ("company_style_score",     "company_style_score"),
    ("practicality_score",      "practicality_score"),
    ("final_score",             "final_score"),
    ("mission_fit_notes",       "mission_fit_notes"),
    ("scope_fit_notes",         "scope_fit_notes"),
    ("top_reasons",             "top_reasons"),
    ("top_concerns",            "top_concerns"),
    ("job_file",                "job_file"),
    ("ClaudeStatus",            None),
]

COL_ALIASES = {
    "desire_score":       "passion_score",
    "practicality_score": "quality_of_life_score",
}

SCORE_COLS = ["desire_score", "market_perception_score", "company_style_score", "practicality_score", "final_score"]
SUB_SCORE_COLS = ["desire_score", "market_perception_score", "company_style_score", "practicality_score"]

# Sub-score 0-100 ramp (unchanged).
SCORE_BUCKETS = [
    ("85+", "5CE05C"), ("80-84", "88E888"), ("75-79", "D9EAD3"), ("70-74", "FCF3CE"),
    ("65-69", "FCE5D5"), ("60-64", "FAD7C0"), ("50-59", "F4D4D4"), ("<50", "ECBFBF"),
]
# final_score bands aligned to Status thresholds (statusFor in vet-jobs: 80 / 70 / 60).
FINAL_STRONG = "88E888"   # >= 80  Apply ASAP
FINAL_MAYBE  = "D9EAD3"   # 70-79  Apply If Time
FINAL_WEAK   = "FCF3CE"   # 60-69  Backup Lane
FINAL_SKIP   = "ECBFBF"   # < 60   Skip

# Candidate-preference -> color. The config stores the PREFERENCE word; styling lives here.
RATING_COLORS = {
    "preferred": "A9D08E",  # green
    "ok":        "FFE699",  # yellow
    "stretch":   "F4B183",  # orange
    "no":        "F4A6A6",  # red
    None:        "D9D9D9",  # grey / unknown
}
GREEN, YELLOW, RED, GREY = "A9D08E", "FFE699", "F4A6A6", "D9D9D9"

CAT_PALETTE = [
    "D9E1F2", "E2EFDA", "FCE4D6", "FFF2CC", "EAD1DC",
    "DDEBF7", "E2F0D9", "FBE5D6", "F4CCCC", "D9D2E9",
]

WRAP_COLS = {
    "Category", "Location?", "Location Fit", "Job Post Title + Link", "lane", "Lane Fit",
    "Status?", "Notes", "mission_fit_notes", "scope_fit_notes", "top_reasons", "top_concerns",
}
CENTER_COLS = {
    "Applied", "Have Intro?", "Decline/Down Date?", "Comp Range", "Comp Fit",
    "desire_score", "market_perception_score", "company_style_score", "practicality_score", "final_score",
}
WIDTHS = {
    "Applied": 9, "Status?": 24, "Category": 26, "Company": 20, "Job Post Title + Link": 52,
    "Location?": 24, "Location Fit": 18, "Comp Range": 12, "Comp Fit": 16, "Have Intro?": 12,
    "Decline/Down Date?": 16, "Notes": 28, "lane": 22, "Lane Fit": 22,
    "desire_score": 9, "market_perception_score": 11, "company_style_score": 11,
    "practicality_score": 11, "final_score": 10, "mission_fit_notes": 44, "scope_fit_notes": 44,
    "top_reasons": 48, "top_concerns": 48, "job_file": 28, "ClaudeStatus": 18,
}


# --------------------------------------------------------------------------- #
# Candidate-relative fit logic (pure — no openpyxl needed)
# --------------------------------------------------------------------------- #
def load_config(path) -> dict:
    if not path:
        return {}
    try:
        data = json.loads(Path(path).read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def comp_fit(text, cfg) -> tuple:
    """(label, color). Uses the TOP of the posted range vs the candidate's floor/target."""
    t = (text or "").strip()
    nums = [int(n) for n in re.findall(r"\d+", t)]
    if not t or "?" in t or not nums:
        return ("Unknown", GREY)
    comp = cfg.get("comp") or {}
    floor, target = comp.get("floor_base"), comp.get("target_base")
    if floor is None and target is None:
        return ("No comp prefs", GREY)
    high = max(nums)
    if floor is not None and high < floor:
        return ("Below floor", RED)
    if target is not None and high >= target:
        return ("Meets/above target", GREEN)
    if target is not None:
        return ("Near target", YELLOW)   # not below floor (or no floor), but short of target
    return ("Above floor", GREEN)         # only a floor is known and the range clears it


def location_fit(text, cfg) -> tuple:
    """(label, color). Maps the normalized Location? string + the candidate's home metro to
    one arrangement, then colors it by the candidate's rating for that arrangement."""
    loc = (text or "").strip().lower()
    locp = cfg.get("location") or {}
    arr = locp.get("arrangements") or {}
    aliases = [a.lower() for a in (locp.get("home_metro_aliases") or []) if a]
    home = (locp.get("home_metro") or "").strip().lower()
    if home:
        aliases.append(home)

    def rc(kind):
        return RATING_COLORS.get(arr.get(kind))  # rating word -> color; missing -> grey

    if not loc or "unknown" in loc or "unclear" in loc:
        return ("Unclear", GREY)
    if "remote" in loc:
        if "state" in loc:
            return ("Remote (state-restricted)", rc("remote"))
        return ("Remote", rc("remote"))
    if any(k in loc for k in ("irl", "onsite", "on-site", "hybrid", "in-office", "in office")):
        m = re.search(r"(\d+)\s*day", loc)
        days = int(m.group(1)) if m else None
        onsite = ("onsite" in loc or "on-site" in loc) or (days is not None and days >= 5)
        mode = "onsite" if onsite else "hybrid"
        if aliases and any(a in loc for a in aliases):
            return (f"Home {mode}", rc(f"home_{mode}"))
        if aliases:
            return (f"Other {mode}", rc(f"other_{mode}"))
        return (f"{mode.capitalize()} (home metro not set)", GREY)
    return ("Unclear", GREY)


def config_complete_enough(cfg) -> bool:
    comp = cfg.get("comp") or {}
    loc = (cfg.get("location") or {}).get("arrangements") or {}
    has_comp = comp.get("floor_base") is not None or comp.get("target_base") is not None
    has_loc = any(v is not None for v in loc.values())
    return has_comp or has_loc


# --------------------------------------------------------------------------- #
def parse_title_link(val):
    if not val:
        return val, None
    idx = val.find(" | http")
    if idx == -1:
        idx = val.find(" |http")
    if idx != -1:
        return val[:idx].strip(), val[idx:].lstrip(" |").strip()
    return val.strip(), None


def read_records(path):
    with open(path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers) if h.strip()}

    def get(row, name):
        i = idx.get(name)
        if i is None and name in COL_ALIASES:
            i = idx.get(COL_ALIASES[name])
        return row[i] if (i is not None and i < len(row)) else ""

    records = []
    for row in rows[1:]:
        if not any(c.strip() for c in row):
            continue
        rec = {key: get(row, key) for _, key in COLUMNS if key is not None}
        for s in SCORE_COLS:
            v = str(rec.get(s, "")).strip()
            rec[s] = int(v) if v.lstrip("-").isdigit() else (v or None)
        records.append(rec)
    records.sort(key=lambda r: (r.get("final_score") or 0), reverse=True)
    return records


def solid(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _lane_fit_color(text):
    t = (text or "").lower()
    if "high" in t:
        return "E2EFDA"   # faint green
    if "low" in t:
        return "F2F2F2"   # faint grey (low confidence)
    return None           # medium / unknown -> no fill


def build(input_csv, output_xlsx, config_path=None, quarantined=0):
    if not HAVE_OPENPYXL:
        raise RuntimeError("openpyxl is required to build the .xlsx (install requirements.txt in the venv).")
    cfg = load_config(config_path)
    records = read_records(input_csv)
    for rec in records:
        rec["_comp_fit"] = comp_fit(rec.get("Comp Range", ""), cfg)
        rec["_loc_fit"] = location_fit(rec.get("Location?", ""), cfg)

    cat_color = {}
    for rec in records:
        cat = (rec.get("Category") or "").strip()
        if cat and cat not in cat_color:
            cat_color[cat] = CAT_PALETTE[len(cat_color) % len(CAT_PALETTE)]

    wb = Workbook()
    ws = wb.active
    ws.title = "Job Rankings"
    ncols = len(COLUMNS)
    last_row = len(records) + 1
    letter = {hdr: get_column_letter(i + 1) for i, (hdr, _) in enumerate(COLUMNS)}

    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    hdr_fill = solid("305496")
    hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    for c, (hdr, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(1, c, hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30

    # Header notes: quarantine count + missing-config guidance.
    notes = []
    if quarantined and int(quarantined) > 0:
        notes.append(f"Prep quarantined {quarantined} thin/failed post(s); they were NOT ranked. "
                     f"See '0 - Prep Report/'.")
    if not config_complete_enough(cfg):
        notes.append("Candidate preferences (jail.config.json comp/location) are missing or empty, "
                     "so Comp Fit / Location Fit are neutral. Run /intake or fill jail.config.json "
                     "for candidate-relative coloring.")
    if notes:
        ws["A1"].comment = Comment("\n".join(notes), "JAIL")

    base_font = Font(name=FONT, size=10, color="000000")
    link_font = Font(name=FONT, size=10, color="0563C1", underline="single")
    for r, rec in enumerate(records, start=2):
        for c, (hdr, key) in enumerate(COLUMNS, start=1):
            if hdr == "Comp Fit":
                val = rec["_comp_fit"][0]
            elif hdr == "Location Fit":
                val = rec["_loc_fit"][0]
            else:
                val = "" if key is None else rec.get(key, "")
            if val is None:
                val = ""
            cell = ws.cell(r, c)
            cell.border = BORDER
            if hdr == "Job Post Title + Link" and val:
                title, url = parse_title_link(str(val))
                cell.value = title
                if url:
                    cell.hyperlink = url
                    cell.font = link_font
                else:
                    cell.font = base_font
            else:
                cell.value = val
                cell.font = base_font
            if hdr in CENTER_COLS:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif hdr in WRAP_COLS:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top")

        comp_c = rec["_comp_fit"][1]
        loc_c = rec["_loc_fit"][1]
        ws[f"{letter['Comp Range']}{r}"].fill = solid(comp_c)
        ws[f"{letter['Comp Fit']}{r}"].fill = solid(comp_c)
        ws[f"{letter['Location?']}{r}"].fill = solid(loc_c)
        ws[f"{letter['Location Fit']}{r}"].fill = solid(loc_c)
        lf = _lane_fit_color(rec.get("Lane Fit", ""))
        if lf:
            ws[f"{letter['Lane Fit']}{r}"].fill = solid(lf)
        cat = (rec.get("Category") or "").strip()
        if cat in cat_color:
            ws[f"{letter['Category']}{r}"].fill = solid(cat_color[cat])

    for hdr, _ in COLUMNS:
        ws.column_dimensions[letter[hdr]].width = WIDTHS.get(hdr, 16)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{last_row}"

    if last_row < 2:
        wb.save(output_xlsx)
        return 0

    # Sub-score 0-100 ramp.
    for s in SUB_SCORE_COLS:
        col = letter[s]
        rng = f"{col}2:{col}{last_row}"
        rules = [
            CellIsRule(operator="greaterThanOrEqual", formula=["85"], fill=solid(SCORE_BUCKETS[0][1]), stopIfTrue=True),
            CellIsRule(operator="between", formula=["80", "84"], fill=solid(SCORE_BUCKETS[1][1]), stopIfTrue=True),
            CellIsRule(operator="between", formula=["75", "79"], fill=solid(SCORE_BUCKETS[2][1]), stopIfTrue=True),
            CellIsRule(operator="between", formula=["70", "74"], fill=solid(SCORE_BUCKETS[3][1]), stopIfTrue=True),
            CellIsRule(operator="between", formula=["65", "69"], fill=solid(SCORE_BUCKETS[4][1]), stopIfTrue=True),
            CellIsRule(operator="between", formula=["60", "64"], fill=solid(SCORE_BUCKETS[5][1]), stopIfTrue=True),
            CellIsRule(operator="between", formula=["50", "59"], fill=solid(SCORE_BUCKETS[6][1]), stopIfTrue=True),
            CellIsRule(operator="lessThan", formula=["50"], fill=solid(SCORE_BUCKETS[7][1]), stopIfTrue=True),
        ]
        for rule in rules:
            ws.conditional_formatting.add(rng, rule)

    # final_score colored to the Status bands.
    fcol = letter["final_score"]
    frng = f"{fcol}2:{fcol}{last_row}"
    for rule in [
        CellIsRule(operator="greaterThanOrEqual", formula=["80"], fill=solid(FINAL_STRONG), stopIfTrue=True),
        CellIsRule(operator="between", formula=["70", "79"], fill=solid(FINAL_MAYBE), stopIfTrue=True),
        CellIsRule(operator="between", formula=["60", "69"], fill=solid(FINAL_WEAK), stopIfTrue=True),
        CellIsRule(operator="lessThan", formula=["60"], fill=solid(FINAL_SKIP), stopIfTrue=True),
    ]:
        ws.conditional_formatting.add(frng, rule)

    wb.save(output_xlsx)
    return len(records)


def main(argv):
    parser = argparse.ArgumentParser(description="Build the candidate-relative rankings .xlsx.")
    parser.add_argument("input_csv")
    parser.add_argument("output_xlsx", nargs="?", default=None)
    parser.add_argument("--config", default=None, help="path to jail.config.json (candidate preferences)")
    parser.add_argument("--quarantined", type=int, default=0, help="count of prep-quarantined posts (note only)")
    args = parser.parse_args(argv[1:])
    out = args.output_xlsx or re.sub(r"\.csv$", ".xlsx", args.input_csv, flags=re.I)
    if out == args.input_csv:
        out = args.input_csv + ".xlsx"
    n = build(args.input_csv, out, config_path=args.config, quarantined=args.quarantined)
    print(f"Wrote {out} ({n} jobs).")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
