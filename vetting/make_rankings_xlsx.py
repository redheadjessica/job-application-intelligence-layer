#!/usr/bin/env python3
"""
Build a polished, conditionally-formatted .xlsx from a job-vetting rankings CSV.

Usage:
    python make_rankings_xlsx.py <input-rankings.csv> [output.xlsx]

If the output path is omitted it is derived from the input
(e.g. 06-02-26-rankings.csv -> 06-02-26-rankings.xlsx).

Design notes (kept in sync with CLAUDE.md / the run-batch pipeline):
- Columns are re-ordered into a review-friendly layout and the sheet is sorted by
  final_score descending (highest priority first).
- Score columns and Location use LIVE Excel conditional-formatting rules
  (solid fills, black text) so they re-color if you edit a value.
- Comp Range and Category use a STATIC fill computed here. (Comp parsing a "min-max" string
  into a number inside an Excel formula is brittle; Category colors are assigned from a
  data-driven palette per distinct value.) They are correct at generation time; they will not
  re-color if you hand-edit those cells later.
- "Job Post Title + Link" cells are written as real clickable hyperlinks: the job title is
  the display text, the URL is the hyperlink target.
- Backward-compatible: if the CSV uses old field names (passion_score, company_style_score,
  quality_of_life_score) they are aliased to the current names (desire_score, practicality_score).
"""
import csv
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

FONT = "Arial"

# ---- Output column layout: (header, source-CSV-header or None for a blank manual column) ----
# None = blank column for manual entry; string = read from that CSV column name.
COLUMNS = [
    ("Applied",                 None),
    ("Status?",                 "Status?"),
    ("Category",                "Category"),
    ("Company",                 "Company"),
    ("Job Post Title + Link",   "Job Post Title + Link"),
    ("Location?",               "Location?"),
    ("Comp Range",              "Comp Range"),
    ("Have Intro?",             None),
    ("Decline/Down Date?",      None),
    ("Notes",                   None),
    ("lane",                    "lane"),
    ("desire_score",            "desire_score"),
    ("market_perception_score", "market_perception_score"),
    ("company_style_score",     "company_style_score"),
    ("practicality_score",      "practicality_score"),
    ("final_score",             "final_score"),
    ("mission_fit_notes",       "mission_fit_notes"),
    ("scope_fit_notes",         "scope_fit_notes"),
    ("top_reasons",             "top_reasons"),
    ("top_concerns",            "top_concerns"),
    ("job_file",                "job_file"),
    ("ClaudeStatus",            None),
]

# Aliases: if a CSV column is missing, try the legacy name instead.
# Handles the transition from the old 4-score system to the current 3-score system.
COL_ALIASES = {
    "desire_score":       "passion_score",          # old CSV used passion_score
    "practicality_score": "quality_of_life_score",  # old CSV used quality_of_life_score
    # company_style_score kept the same name — no alias needed
}

SCORE_COLS = ["desire_score", "market_perception_score", "company_style_score", "practicality_score", "final_score"]

# Score buckets: (label, fill hex). Highest first; black text everywhere.
# Bright green top, soft pastel ramp, soft pink (not red) at bottom.
SCORE_BUCKETS = [
    ("85+",   "5CE05C"),  # bright green
    ("80-84", "88E888"),  # light bright green
    ("75-79", "D9EAD3"),  # pale sage green
    ("70-74", "FCF3CE"),  # pale yellow
    ("65-69", "FCE5D5"),  # pale peach
    ("60-64", "FAD7C0"),  # light peach / tan
    ("50-59", "F4D4D4"),  # light pink
    ("<50",   "ECBFBF"),  # soft deeper pink
]

# Comp tiers (by midpoint of the range, in $thousands).
COMP_EXCELLENT = "A9D08E"   # >= 230  top of target
COMP_SOLID     = "C6E0B4"   # 200-229 solid / acceptable
COMP_BELOW     = "F8CBAD"   # 170-199 below target
COMP_LOW       = "F4CCCC"   # < 170   low
COMP_MISSING   = "D9D9D9"   # ?? / unknown

LOC_GREEN  = "A9D08E"   # remote
LOC_YELLOW = "FFE699"   # remote but restricted to certain states
LOC_PINK   = "F4CCCC"   # another city / on-site elsewhere

# Category fills: a generic, data-driven palette. Distinct "Category" values are
# assigned colors from this list in order of appearance (see build()).
CAT_PALETTE = [
    "D9E1F2",  # light blue
    "E2EFDA",  # light green
    "FCE4D6",  # light peach
    "FFF2CC",  # light yellow
    "EAD1DC",  # light pink
    "DDEBF7",  # paler blue
    "E2F0D9",  # paler green
    "FBE5D6",  # paler peach
    "F4CCCC",  # soft rose
    "D9D2E9",  # light lavender
]

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_COLS = {
    "Category", "Location?", "Job Post Title + Link", "lane", "Status?", "Notes",
    "mission_fit_notes", "scope_fit_notes", "top_reasons", "top_concerns",
}
CENTER_COLS = {
    "Applied", "Have Intro?", "Decline/Down Date?", "Comp Range",
    "desire_score", "market_perception_score", "company_style_score", "practicality_score", "final_score",
}
WIDTHS = {
    "Applied": 9,
    "Status?": 24,
    "Category": 26,
    "Company": 20,
    "Job Post Title + Link": 52,
    "Location?": 26,
    "Comp Range": 12,
    "Have Intro?": 12,
    "Decline/Down Date?": 16,
    "Notes": 30,
    "lane": 26,
    "desire_score": 9,
    "market_perception_score": 11,
    "company_style_score": 11,
    "practicality_score": 11,
    "final_score": 10,
    "mission_fit_notes": 46,
    "scope_fit_notes": 46,
    "top_reasons": 50,
    "top_concerns": 50,
    "job_file": 30,
    "ClaudeStatus": 20,
}


def solid(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def comp_fill(text):
    t = (text or "").strip()
    if not t or "?" in t:
        return COMP_MISSING
    nums = [int(n) for n in re.findall(r"\d+", t)]
    if not nums:
        return COMP_MISSING
    mid = sum(nums) / len(nums)
    if mid >= 230:
        return COMP_EXCELLENT
    if mid >= 200:
        return COMP_SOLID
    if mid >= 170:
        return COMP_BELOW
    return COMP_LOW


def parse_title_link(val):
    """Split 'Role Title | https://...' into (title, url). Returns (val, None) if no URL."""
    if not val:
        return val, None
    idx = val.find(" | http")
    if idx == -1:
        idx = val.find(" |http")
    if idx != -1:
        title = val[:idx].strip()
        url = val[idx:].lstrip(" |").strip()
        return title, url
    return val.strip(), None


def read_records(path):
    with open(path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers) if h.strip()}

    def get(row, name):
        """Read a value by column name, falling back to legacy alias if the primary is absent."""
        i = idx.get(name)
        if i is None and name in COL_ALIASES:
            i = idx.get(COL_ALIASES[name])
        return row[i] if (i is not None and i < len(row)) else ""

    records = []
    for row in rows[1:]:
        if not any(c.strip() for c in row):
            continue
        rec = {}
        for _, key in COLUMNS:
            if key is None:
                continue
            rec[key] = get(row, key)
        for s in SCORE_COLS:
            v = str(rec.get(s, "")).strip()
            rec[s] = int(v) if v.lstrip("-").isdigit() else (v or None)
        records.append(rec)

    # Sort by final_score descending (highest priority first).
    records.sort(key=lambda r: (r.get("final_score") or 0), reverse=True)
    return records


def build(input_csv, output_xlsx):
    records = read_records(input_csv)

    # Data-driven category palette: assign palette colors to the distinct
    # "Category" values in order of appearance (cycling if there are more
    # categories than palette entries). Applied as a static per-row fill below.
    cat_color = {}
    for rec in records:
        cat = (rec.get("Category") or "").strip()
        if cat and cat not in cat_color:
            cat_color[cat] = CAT_PALETTE[len(cat_color) % len(CAT_PALETTE)]

    wb = Workbook()
    ws = wb.active
    ws.title = "Job Rankings"

    ncols = len(COLUMNS)
    last_row = len(records) + 1
    letter = {hdr: get_column_letter(i + 1) for i, (hdr, _) in enumerate(COLUMNS)}

    # ---- Header row ----
    hdr_fill = solid("305496")
    hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    for c, (hdr, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(1, c, hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30

    # ---- Data rows ----
    base_font = Font(name=FONT, size=10, color="000000")
    link_font = Font(name=FONT, size=10, color="0563C1", underline="single")
    for r, rec in enumerate(records, start=2):
        for c, (hdr, key) in enumerate(COLUMNS, start=1):
            val = "" if key is None else rec.get(key, "")
            if val is None:
                val = ""

            cell = ws.cell(r, c)
            cell.border = BORDER

            # Real hyperlink for the title+link column
            if hdr == "Job Post Title + Link" and val:
                title, url = parse_title_link(str(val))
                cell.value = title
                if url:
                    cell.hyperlink = url
                    cell.font = link_font
                else:
                    cell.font = base_font
            else:
                cell.value = val
                cell.font = base_font

            if hdr in CENTER_COLS:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif hdr in WRAP_COLS:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top")

        # Static comp fill
        ws[f"{letter['Comp Range']}{r}"].fill = solid(comp_fill(rec.get("Comp Range", "")))

        # Static category fill (data-driven palette)
        cat = (rec.get("Category") or "").strip()
        if cat in cat_color:
            ws[f"{letter['Category']}{r}"].fill = solid(cat_color[cat])

    # ---- Column widths ----
    for hdr, _ in COLUMNS:
        ws.column_dimensions[letter[hdr]].width = WIDTHS.get(hdr, 16)

    # ---- Header niceties: freeze, filter ----
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{last_row}"

    if last_row < 2:
        wb.save(output_xlsx)
        return 0

    # ---- Conditional formatting: score buckets (live, solid fill, black text) ----
    for s in SCORE_COLS:
        col = letter[s]
        rng = f"{col}2:{col}{last_row}"
        rules = [
            CellIsRule(operator="greaterThanOrEqual", formula=["85"], fill=solid(SCORE_BUCKETS[0][1]), stopIfTrue=True),
            CellIsRule(operator="between", formula=["80", "84"], fill=solid(SCORE_BUCKETS[1][1]), stopIfTrue=True),
            CellIsRule(operator="between", formula=["75", "79"], fill=solid(SCORE_BUCKETS[2][1]), stopIfTrue=True),
            CellIsRule(operator="between", formula=["70", "74"], fill=solid(SCORE_BUCKETS[3][1]), stopIfTrue=True),
            CellIsRule(operator="between", formula=["65", "69"], fill=solid(SCORE_BUCKETS[4][1]), stopIfTrue=True),
            CellIsRule(operator="between", formula=["60", "64"], fill=solid(SCORE_BUCKETS[5][1]), stopIfTrue=True),
            CellIsRule(operator="between", formula=["50", "59"], fill=solid(SCORE_BUCKETS[6][1]), stopIfTrue=True),
            CellIsRule(operator="lessThan", formula=["50"], fill=solid(SCORE_BUCKETS[7][1]), stopIfTrue=True),
        ]
        for rule in rules:
            ws.conditional_formatting.add(rng, rule)

    # ---- Conditional formatting: Location ----
    loc = letter["Location?"]
    loc_rng = f"{loc}2:{loc}{last_row}"
    a = f"{loc}2"
    loc_rules = [
        # State-restricted remote -> caution (yellow), checked before plain "remote" -> green.
        FormulaRule(formula=[f'AND(ISNUMBER(SEARCH("remote",{a})),ISNUMBER(SEARCH("states",{a})))'], fill=solid(LOC_YELLOW), stopIfTrue=True),
        FormulaRule(formula=[f'ISNUMBER(SEARCH("remote",{a}))'], fill=solid(LOC_GREEN), stopIfTrue=True),
        FormulaRule(formula=[(
            f'OR(ISNUMBER(SEARCH("san francisco",{a})),ISNUMBER(SEARCH("sf",{a})),'
            f'ISNUMBER(SEARCH("new york",{a})),ISNUMBER(SEARCH("nyc",{a})),'
            f'ISNUMBER(SEARCH("boston",{a})),ISNUMBER(SEARCH("chicago",{a})),'
            f'ISNUMBER(SEARCH("austin",{a})),ISNUMBER(SEARCH("los angeles",{a})),'
            f'ISNUMBER(SEARCH("seattle",{a})),ISNUMBER(SEARCH("denver",{a})),'
            f'ISNUMBER(SEARCH("atlanta",{a})))'
        )], fill=solid(LOC_PINK), stopIfTrue=True),
    ]
    for rule in loc_rules:
        ws.conditional_formatting.add(loc_rng, rule)

    wb.save(output_xlsx)
    return len(records)


def main(argv):
    if len(argv) < 2:
        print(__doc__)
        return 2
    inp = argv[1]
    out = argv[2] if len(argv) > 2 else re.sub(r"\.csv$", ".xlsx", inp, flags=re.I)
    if out == inp:
        out = inp + ".xlsx"
    n = build(inp, out)
    print(f"Wrote {out} ({n} jobs).")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
