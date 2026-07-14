#!/usr/bin/env python3
"""
Apply the candidate's "On Ice" overlay to a rankings spreadsheet (and its sibling CSV).

Philosophy: never have two simultaneous applications at the same company (bad signal).

Rules:
  1. Any role whose COMPANY is in on-ice-companies.txt (the candidate applied & is awaiting a response)
     -> Status = "Apply Eventually: On Ice (Applied to Another Position at this Company)".
  2. For any OTHER company that has MULTIPLE roles in this ranking (and the candidate has NOT applied there),
     keep only the single TOP-scoring role's status; set all the lower-scoring roles -> On Ice.
  3. Single-role, not-applied companies keep whatever status the vet assigned.

This must be re-run after every vet, because the vet re-computes Status from score and would
otherwise clobber the On-Ice tracking.

Usage:
  python apply_on_ice.py --rankings "__READY_TO_REVIEW__PRIVATE_GITIGNORED/<batch>/1 - Rankings/<batch>-rankings.xlsx"
                         [--companies PRIVATE__YOUR_FILES_GITIGNORED/03-VETTING__YOUR_PRIVATE_INFO/on-ice-companies.txt]
"""
import argparse
import csv
import os
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl required\n"); sys.exit(1)

ON_ICE = "Apply Eventually: On Ice (Applied to Another Position at this Company)"


def norm(s):
    return "".join(ch for ch in (s or "").lower() if ch.isalnum())


def load_companies(path):
    out = []
    for line in open(path, encoding="utf-8"):
        line = line.strip()
        if line and not line.startswith("#"):
            out.append(line)
    return out


def is_applied(company, applied):
    c = norm(company)
    if not c:
        return False  # blank/summary rows never match
    # an applied-company name must appear inside the job's company (not the reverse — avoids a
    # short applied name spuriously matching, and empty strings matching everything)
    return any(na and na in c for na in (norm(a) for a in applied))


def main(argv):
    ap = argparse.ArgumentParser()
    ap.add_argument("--rankings", required=True)
    ap.add_argument("--companies", default="PRIVATE__YOUR_FILES_GITIGNORED/03-VETTING__YOUR_PRIVATE_INFO/on-ice-companies.txt")
    a = ap.parse_args(argv[1:])
    applied = load_companies(a.companies)

    wb = openpyxl.load_workbook(a.rankings)
    ws = wb.active
    H = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    # Header names match the current tracker layout (vet-jobs.js HEADERS). "Final Score" is looked up
    # by prefix/contains rather than an exact literal, since the score-column LABEL is dynamic
    # (score-dimensions.json / the candidate's scoring card) and the engine-owned default is
    # "FINAL Weighted Score" today but may not always be — this keeps the on-ice pass from breaking
    # again the next time that label changes.
    final_col_name = next((h for h in H if h and "final" in h.lower()), None)
    if final_col_name is None:
        sys.stderr.write("Could not find a Final-Score-like column header in the rankings sheet\n"); sys.exit(1)
    sc, co, jf, fc = H["Status? [You Change]"], H["Company"], H["Job File"], H[final_col_name]

    def fs(r):
        v = ws.cell(r, fc).value
        return int(v) if v not in (None, "") and str(v).lstrip("-").isdigit() else 0

    # Job rows end at the first blank Company cell — make_rankings_xlsx.py appends a blank gap row
    # plus a merged section-color legend block below the data, which must NOT be iterated (writing
    # to a merged, non-anchor cell raises AttributeError; it also isn't job data).
    last_data_row = 1
    for r in range(2, ws.max_row + 1):
        if not (ws.cell(r, co).value or "").strip():
            break
        last_data_row = r
    rows = [(r, ws.cell(r, co).value, fs(r)) for r in range(2, last_data_row + 1)]
    changed = {}  # job_file -> new status

    # Rule 1: applied companies -> all roles On Ice
    for r, c, f in rows:
        if is_applied(c, applied):
            ws.cell(r, sc, ON_ICE)
            changed[ws.cell(r, jf).value] = ON_ICE

    # Rule 2: not-applied companies with >1 role -> keep top scorer, rest On Ice
    groups = defaultdict(list)
    for r, c, f in rows:
        if not is_applied(c, applied):
            groups[norm(c)].append((f, r))
    for k, lst in groups.items():
        if len(lst) > 1:
            lst.sort(key=lambda x: -x[0])  # highest final_score first
            for f, r in lst[1:]:
                ws.cell(r, sc, ON_ICE)
                changed[ws.cell(r, jf).value] = ON_ICE

    wb.save(a.rankings)

    # Patch the sibling CSV so the paste-into-Google-Sheet source matches.
    csvp = a.rankings[:-5] + ".csv" if a.rankings.endswith(".xlsx") else a.rankings + ".csv"
    if os.path.exists(csvp):
        recs = list(csv.DictReader(open(csvp, newline="", encoding="utf-8")))
        if recs:
            fields = list(recs[0].keys())
            stf = next((k for k in fields if k and k.strip().lower().startswith("status?")), None)
            jff = next((k for k in fields if k == "Job File"), None)
            if stf and jff:
                for row in recs:
                    if row.get(jff) in changed:
                        row[stf] = changed[row[jff]]
                w = csv.DictWriter(open(csvp, "w", newline="", encoding="utf-8"), fieldnames=fields)
                w.writeheader(); w.writerows(recs)

    print(f"On-Ice applied to {len(changed)} role(s):")
    for k in sorted(changed):
        print(f"  {k}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
