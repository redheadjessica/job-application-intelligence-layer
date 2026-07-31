#!/usr/bin/env python3
"""
Build a polished job-search TRACKER .xlsx from a job-vetting rankings CSV.

The CSV (written by vet-jobs) is CLEAN DATA: a header + one row per job, in final-score order, in
the tracker layout (human-editable columns first — headers ending "? [You ...]" — then AI
scoring/detail). No divider rows: the data stays sortable (no merged cells) and a user can paste
rows into their own tracker without dragging duplicate dividers.

This script STYLES that data and adds tracker affordances:
- header frozen (freeze A2), every cell wrapped + centered + top-aligned,
- a Status dropdown (inline list, like the reference workbook — shows the arrow; applied to the job
  rows only),
- auto-filter on the job rows so "Sort A->Z" is one click (the legend below is outside its range),
- lifecycle/stoplight fills for Status / Lane / Comp Fit, plus the fixed 4-hex Working Location
  palette applied to the Working Location cells,
- Final Score / sub-score color ramps,
- a SEPARATE section-color legend block below the jobs (merged A:J bars) — a palette you can copy a
  bar from if you sort and want visual breaks; it is NOT mixed into the sortable data,
- an Instructions tab.

Note on "chips": the rounded colored dropdown pills are a Google Sheets-native feature configured in
Sheets; they cannot be embedded in an .xlsx. This file gives flat cell fills + a working dropdown.

Output-contract normalization + the Working Location color mapper live in norm_contracts.py (single
source, same directory); this script re-normalizes contract-governed cells on read so regenerating an
OLD CSV also repairs its text and colors. Candidate-relative fit LABEL text still originates in
vet-jobs.js.
Usage:
    python make_rankings_xlsx.py <input-rankings.csv> [output.xlsx] \
        [--config jail.config.json] [--quarantined N]
"""
import argparse
import csv
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import norm_contracts  # noqa: E402 — shared output-contract normalizers + color mappers

try:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    HAVE_OPENPYXL = True
except ImportError:  # the pure label->color/config logic below stays importable + testable
    HAVE_OPENPYXL = False

FONT = "Arial"

# ---- The tracker column layout (must match vet-jobs.js HEADERS). The CSV header row drives the
# actual order; these names locate the columns that need special styling. ----
H_STATUS = "Status? [You Change]"
H_LANE = "Lane"
H_COMPANY = "Company"
H_TITLE = "Job Post Title + Link"
H_WORKLOC = "Working Location"
H_COMPRANGE = "Comp Range"
H_POSTED = norm_contracts.H_POSTED    # "ATS First Posted Date" (static date or `Unknown`)
H_UPDATED = norm_contracts.H_UPDATED  # "ATS Last Updated Date" (static date or BLANK)
H_JOBFILE = "Job File"
H_LANEFIT = "Lane Fit"
H_COMPFIT = "Comp Fit"
H_DATACOMPLETE = "Data Completeness"  # per-job comp/location capture quality (green/amber/red)
H_PRACTNOTES = norm_contracts.H_PRACTNOTES  # "Comp + Lifestyle Fit Notes" — prose companion to the
# practicality score; sits immediately after it. Static name even though that score's label is
# candidate-relabelable, exactly like "Mission Fit Notes".
LEGEND_MERGE_TO = 10  # section bars merge A:J (columns 1..10 — the human-facing block)

# ---- Score-column labels/weights/definitions are DYNAMIC (see load_meta / DEFAULT_METADATA below).
# H_FINAL / SUB_SCORE_COLS / SCORE_COLS are no longer module-level constants — they are resolved once
# per run inside build() from the per-run "<batch>-rankings.meta.json" written by vet-jobs.js (or this
# DEFAULT_METADATA fallback, kept in sync with the engine-owned ENGINE__PUBLIC_GIT_TRACKED/03-VETTING/
# score-dimensions.json, if no meta file is found — e.g. when this script is run standalone). ----
DEFAULT_METADATA = {
    "order": ["final", "market", "desire", "style", "practicality"],
    "final": {"label": "FINAL Weighted Score"},
    "desire": {"label": "Your Desire Score", "weight_pct": 35,
               "definition": "Estimates how much you'd likely want the role if hired and if logistics were workable. May consider mission, product, users, problems, scope, career direction, and personal interests. Should not primarily measure compensation, location, or whether the employer is likely to hire you."},
    "market": {"label": "How They May See Your Profile", "weight_pct": 30,
               "definition": "Estimates how competitive and legible you may appear to this employer before tailoring, based on the canonical summary profile available during vetting and the job posting. It does not use the newly tailored resume. A preference for the company, mission, or lane is not evidence that the employer will see you as qualified."},
    "style": {"label": "Culture Fit Score", "weight_pct": 20,
              "definition": "Estimates how well the company's apparent working style, values, product culture, and environment may suit you, based on the evidence actually available. Job postings provide incomplete culture evidence. When little reliable information is available, this score should remain closer to neutral and should be treated as lower-confidence."},
    "practicality": {"label": "Comp + Lifestyle Fit Score", "weight_pct": 15,
                     "definition": "Estimates how well compensation, location, work arrangement, travel, schedule, and other practical considerations fit your stated preferences. A lower score reduces the opportunity's priority but is not automatically a veto."},
}


def load_meta(csv_path) -> dict:
    """Load the per-run '<batch>-rankings.meta.json' written by vet-jobs.js (sibling of the CSV).
    Falls back to DEFAULT_METADATA per-key if the file is missing, unreadable, or incomplete —
    this keeps the script usable standalone (tests, ad-hoc regeneration) without a meta file."""
    p = Path(csv_path)
    meta_path = p.with_name(p.stem + ".meta.json")
    data = {}
    try:
        data = json.loads(meta_path.read_text(encoding="utf-8"))
    except Exception:
        data = {}
    merged = {}
    for key in ("final", "desire", "market", "style", "practicality"):
        merged[key] = {**DEFAULT_METADATA.get(key, {}), **(data.get(key) or {})}
    merged["order"] = data.get("order") or DEFAULT_METADATA["order"]
    return merged

# ---- Status vocabulary: the 12 dropdown values (vet-jobs statusFor outputs are a subset). ----
# Defined ONCE in norm_contracts, alongside the repair function — a second copy here would be a
# place for the two to drift, and a status that drifts loses its dropdown match and its color.
STATUS_VALUES = norm_contracts.STATUS_VALUES

# Section legend labels -> (fill, font) for the merged bars. Vivid palette matching the reference
# workbook; Interviewed is brown to read distinct from the Closed/Down red.
SECTION_COLORS = {
    "**Currently In Talks**":        ("0000FF", "FFFFFF"),
    "Applied: Awaiting Response":    ("00FFFF", "000000"),
    "Apply ASAP":                    ("00FF00", "000000"),
    "Apply Eventually":              ("FFFF00", "000000"),
    "Closed / Not Moving Forward":   ("FF0000", "FFFFFF"),
    "Interviewed, Rejected":         ("7B3F00", "FFFFFF"),
}

# Per-status cell fills (lifecycle stoplight, sampled from the reference workbook's chips).
STATUS_COLORS = {
    "**Currently In Talks**":                                                ("2F5597", "FFFFFF"),
    "Applied: Awaiting Response":                                            ("9DC3E6", "000000"),
    "Apply Again??":                                                         ("D9D9D9", "000000"),
    "Apply ASAP: High Prio":                                                 ("375623", "FFFFFF"),
    "Apply Eventually: Apply If Time":                                       ("A9D08E", "000000"),
    "Apply Eventually: Backup Lane":                                         ("FFE699", "000000"),
    "Apply Eventually: On Ice (Applied to Another Position at this Company)": ("F4B183", "000000"),
    "Apply Eventually: Or Skip It":                                          ("F8CBAD", "000000"),
    "Declined (Applied, Rejected)":                                          ("A52A2A", "FFFFFF"),
    "Down (Applied, No Response)":                                           ("A52A2A", "FFFFFF"),
    "Down: Closed Before Applying":                                          ("A52A2A", "FFFFFF"),
    "Interviewed: Rejected":                                                 ("7B3F00", "FFFFFF"),
}

# HARD-STOP row marker (Jessica, 7/16/26): vet-jobs.js sets this exact status string, never a
# score, when its scoring agent could not verify a job file actually contains real posting
# content (see the 488KB-Microsoft-boilerplate incident this fixes). This must be impossible to
# miss — the whole row gets this loud fill, not just the Status cell, and it overrides every
# other per-cell color below (lane, comp, location) since none of that data is trustworthy either.
NEEDS_REFETCH_STATUS = "⚠️ NEEDS RE-FETCH — content not verified"
NEEDS_REFETCH_FILL = "FF00FF"  # magenta — deliberately jarring, shouldn't blend with anything else

# Sub-score 0-100 ramp.
SCORE_BUCKETS = [
    ("85+", "5CE05C"), ("80-84", "88E888"), ("75-79", "D9EAD3"), ("70-74", "FCF3CE"),
    ("65-69", "FCE5D5"), ("60-64", "FAD7C0"), ("50-59", "F4D4D4"), ("<50", "ECBFBF"),
]
# Final Score bands aligned to the Status thresholds (statusFor in vet-jobs: 80 / 70 / 60).
FINAL_STRONG, FINAL_MAYBE, FINAL_WEAK, FINAL_SKIP = "88E888", "D9EAD3", "FCF3CE", "ECBFBF"

GREEN, YELLOW, RED, GREY = "A9D08E", "FFE699", "F4A6A6", "D9D9D9"
RATING_COLORS = {"preferred": "A9D08E", "ok": "FFE699", "stretch": "F4B183", "no": "F4A6A6", None: "D9D9D9"}
COMP_LABEL_COLORS = {
    "Below floor": RED, "Meets/above target": GREEN, "Above floor": GREEN,
    "Near target": YELLOW, "Unknown": GREY, "No comp prefs": GREY,
}

# Structural (non-score) column widths — these header strings don't change with score relabeling.
WIDTHS = {
    "Applied Date? [You Fill In]": 16, H_STATUS: 30, H_LANE: 22, H_COMPANY: 18, H_TITLE: 46,
    H_WORKLOC: 22, H_COMPRANGE: 12, H_POSTED: 14, H_UPDATED: 14, "Have Intro? [You Add]": 14, "Your Notes? [You Add]": 26,
    "Decline/Down Date? [You Add]": 16, H_PRACTNOTES: 46, "Your Desire Score Notes": 40,
    "Profile Score Notes": 40,
    "Top Reasons Notes": 46, "Top Concerns Notes": 46, "Job File": 28,
    "Tailored? (Base Resume)": 26,
    H_LANEFIT: 22, H_COMPFIT: 16, H_DATACOMPLETE: 24, "Cover Letter Drafted?": 12,
}
# Score-column widths, keyed by dimension (not by the dynamic label string) — applied at runtime
# once the effective labels are resolved, since the label text itself may change per-run.
SCORE_WIDTHS_BY_KEY = {"final": 13, "market": 30, "desire": 11, "style": 12, "practicality": 16}


# --------------------------------------------------------------------------- #
# Config + label->color logic (pure — no openpyxl needed)
# --------------------------------------------------------------------------- #
def load_config(path) -> dict:
    if not path:
        return {}
    try:
        data = json.loads(Path(path).read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def config_complete_enough(cfg) -> bool:
    comp = cfg.get("comp") or {}
    loc = (cfg.get("location") or {}).get("arrangements") or {}
    has_comp = comp.get("floor_base") is not None or comp.get("target_base") is not None
    has_loc = any(v is not None for v in loc.values())
    return has_comp or has_loc


def resolve_submitted_applications_link(cfg):
    """Explicit, config-owned destination ONLY — no directory scanning, no year inference, no
    "latest folder" guessing. Returns (path_or_None, warning_or_None); caller renders a working link
    when a path comes back, or a clear warning (never a silently-guessed alternate) when it doesn't."""
    configured = ((cfg.get("archive") or {}).get("current_year_path") or "").strip()
    if not configured:
        return None, "No submitted-applications path is configured — set \"archive.current_year_path\" in jail.config.json to enable this link."
    if not Path(configured).is_dir():
        return None, f"Configured submitted-applications path does not exist: {configured} — check \"archive.current_year_path\" in jail.config.json."
    return configured, None


def comp_color(label) -> str:
    return COMP_LABEL_COLORS.get((label or "").strip(), GREY)


# ---- Data Completeness: per-job comp/location capture quality ----------------- #
# vet-jobs.js is the single source of the LABEL text (preferring prep's field_status, else its own
# row fallback) and writes it into the CSV column. This script maps that label TEXT -> a color, and
# — for older CSVs written before the column existed — reproduces vet-jobs.js's ROW FALLBACK here so
# regenerating an old batch still gets a populated, colored column + top flag. Keep the vocabulary
# ("complete" / "not verified" / "unknown" / "not posted") in sync with vet-jobs.js.
COMPLETE_GREEN, COMPLETE_AMBER, COMPLETE_RED = GREEN, YELLOW, RED


# ONE implementation, in norm_contracts — the CSV migration pass back-fills the column with the
# same helper this build uses on read, so a legacy CSV and a current one land identically.
fallback_completeness = norm_contracts.fallback_completeness


def completeness_category(value):
    """complete / benign (only 'not posted') / attention (any 'not verified' / 'unknown'). Mirrors
    vet-jobs.js completenessCategory; drives both the cell color and the top-flag membership."""
    v = (value or "").strip().lower()
    if not v:
        return None
    # A comp-source CONFLICT is attention-worthy even when the capture is otherwise
    # complete ("✓ complete · ⚠ comp conflicting") — checked before "complete".
    if "conflicting" in v:
        return "attention"
    if "complete" in v:
        return "complete"
    if "not verified" in v or "unknown" in v:
        return "attention"
    if "not posted" in v:
        return "benign"
    return "attention"


def completeness_color(value):
    cat = completeness_category(value)
    return {"complete": COMPLETE_GREEN, "benign": COMPLETE_AMBER, "attention": COMPLETE_RED}.get(cat)


def completeness_summary(records):
    """Build the loud top-flag line from the ranked records: '⚠️ Incomplete captures (N): Company —
    Role (flag); ...' listing only rows we could not verify (skips pure 'not posted'), or an
    all-clear line when none. Returns (text, has_attention). Shared by the XLSX banner + Instructions
    tab; the Markdown rankings carry an equivalent line built in vet-jobs.js."""
    attention = [rec for rec in (records or [])
                 if completeness_category(rec.get(H_DATACOMPLETE)) == "attention"
                 and (rec.get(H_STATUS) or "").strip() != NEEDS_REFETCH_STATUS]
    if not attention:
        return "✓ Data completeness: all captures complete.", False
    parts = []
    for rec in attention:
        role = parse_title_link(rec.get(H_TITLE) or "")[0]
        flag = (rec.get(H_DATACOMPLETE) or "").lstrip("⚠ ").strip()
        parts.append(f"{(rec.get(H_COMPANY) or '').strip()} — {role} ({flag})")
    return f"⚠️ Incomplete captures ({len(attention)}): " + "; ".join(parts), True


def loc_color(workloc, cfg) -> str:
    """Color Working Location from the canonical Working Location text — a thin
    call into norm_contracts.working_location_color(), the ONE deterministic mapper (authoritative
    spec, 2026-07-29). Returns exactly one of the 4 spec hexes (42FF35 / FDFF43 / FA9C31 / F82C1F,
    black text): remote -> green; acceptable home-metro office at exactly 1-3 days -> yellow;
    >3 days / open-ended minimums ("3+", "at least 3") / unknown cadence / Unknown -> orange;
    required in-person outside the home geography -> red. NO grey. This SUPERSEDES the July-14 rule
    (which put exactly-3-days at orange and Unknown at grey) and the older per-arrangement-rating
    lookup; jail.config.json's location.arrangements no longer drives this color."""
    return norm_contracts.working_location_color(workloc, cfg)


# Lane color is keyed to the HEALTH DOMAIN itself (Jessica's rule, 7/16/26), not lane-priority
# matching: green is reserved for health jobs only, so it reads as a domain signal at a glance
# rather than colliding with "this matches one of my priority lanes" (which used to also paint
# non-health p1 lanes green). Mental Health gets a distinctly brighter green than other health
# subcategories; nothing outside "Health - ..." is ever colored green.
MENTAL_HEALTH_GREEN = "00B050"
OTHER_HEALTH_GREEN = "C6E0B4"


def lane_color(lane_str, cfg=None):
    """Color the (job-centric) Lane cell purely from the Lane text's own bucket — "Health - ..." only.
    Bright green for "Health - Mental Health" (normalized exactly by vet-jobs.js), a lighter green for
    every other Health subcategory, and no fill at all for every non-Health bucket."""
    s = (lane_str or "").strip()
    if not s.lower().startswith("health"):
        return None
    if s.strip().lower() == "health - mental health":
        return MENTAL_HEALTH_GREEN
    return OTHER_HEALTH_GREEN


def lane_fit_color(lane_fit_str, cfg):
    """Color the (candidate-relative) Lane Fit cell by how well it maps to the candidate's priority
    lanes, reading the Lane Fit string ("Primary (confidence) ..."). p1 lane -> green, p2+ -> amber,
    Outside -> red/grey. Independent of lane_color() above, which is the job-centric Lane cell and is
    now keyed to the health domain instead — the two columns answer different questions and can
    legitimately show different colors on the same row."""
    s = (lane_fit_str or "").strip()
    if not s:
        return None
    m = re.match(r"^(.*?)\s*\((high|medium|low)\)", s)
    primary = (m.group(1) if m else s).strip()
    conf = m.group(2) if m else "medium"
    if primary.lower() == "outside lanes":
        return {"high": RED, "low": GREY}.get(conf, "F4CCCC")
    lanes = {(l.get("name") or "").strip().lower(): l.get("priority") for l in (cfg.get("lanes") or [])}
    pr = lanes.get(primary.lower())
    if pr == 1:
        return GREEN
    if pr in (2, 3):
        return YELLOW
    return "C6E0B4"


def inline_list_formula(values):
    """Build an inline data-validation list like the reference workbook. A single quoted literal is
    capped ~255 chars in Excel, so long lists are split at comma boundaries into <=240-char chunks
    concatenated with & — e.g. "a,b,"&"c,d". Reproduces the exact comma-joined list when evaluated."""
    joined = ",".join(values)
    if len(joined) <= 240:
        return '"' + joined + '"'
    chunks, cur = [], ""
    for i, v in enumerate(values):
        piece = v + ("," if i < len(values) - 1 else "")
        if cur and len(cur) + len(piece) > 240:
            chunks.append(cur)
            cur = piece
        else:
            cur += piece
    if cur:
        chunks.append(cur)
    return "&".join('"' + c + '"' for c in chunks)


# --------------------------------------------------------------------------- #
def parse_title_link(val):
    if not val:
        return val, None
    idx = val.find(" | http")
    if idx == -1:
        idx = val.find(" |http")
    if idx != -1:
        return val[:idx].strip(), val[idx:].lstrip(" |").strip()
    return val.strip(), None


def read_records(path, score_cols):
    """Return (headers, records). Records are data dicts keyed by header, order preserved
    (vet-jobs already sorted by final score). The CSV carries no divider rows.
    score_cols: the 5 effective score-column header strings (int-parsed for conditional formatting)."""
    with open(path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    headers = [h.strip() for h in rows[0]]
    records = []
    for row in rows[1:]:
        if not any(c.strip() for c in row):
            continue
        rec = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
        for s in score_cols:
            v = str(rec.get(s, "")).strip()
            rec[s] = int(v) if v.lstrip("-").isdigit() else (v or None)
        records.append(rec)
    return headers, records


def solid(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


# The human-managed columns. THE RULE (2026-07-30): a column is yours to fill in ONLY when its
# header carries an explicit marker — `[You Fill In]`, `[You Change]`, or `[You Add]`. A bare `?`
# in a header does NOT mean human-managed: `Tailored? (Base Resume)` and `Cover Letter Drafted?`
# are written by the pipeline (update_rankings_row.py) as those steps complete.
HUMAN_MARKERS = ("[You Fill In]", "[You Change]", "[You Add]")


def human_managed_columns(headers):
    return [h for h in headers if any(m in h for m in HUMAN_MARKERS)]


def pipeline_filled_columns(headers):
    """Columns whose header ends in `?` but which the PIPELINE fills — the exception the
    Instructions tab has to state explicitly, or the `?` reads as an invitation."""
    return [h for h in headers
            if h in ("Tailored? (Base Resume)", "Cover Letter Drafted?")]


def build_column_guide(headers):
    """The Instructions tab's column enumeration, generated FROM the real header row so it can
    never drift from the workbook (a test asserts the two agree)."""
    yours = human_managed_columns(headers)
    pipeline = pipeline_filled_columns(headers)
    return [
        ("Which columns are yours", True),
        ("Only columns whose header carries an explicit marker are yours to manage: "
         + ", ".join(yours) + ".", False),
        ("A bare \"?\" in a header does NOT mean the column is yours. "
         + ", ".join(pipeline)
         + " are filled in by the pipeline as the tailoring and cover-letter steps complete — "
           "leave them alone and they will populate themselves.", False),
        ("Everything else is AI-scored or machine-derived and is re-derived on every "
         "regeneration, so edits there are overwritten.", False),
        ("", False),
    ]


INSTRUCTIONS = [
    ("This spreadsheet is two things at once", True),
    ("A ranking report (AI-scored) AND your living job-search tracker.", False),
    ("", False),
    ("Where to start", True),
    ("Rows are ordered best-fit first by Final Score.", False),
    ("", False),
    ("Status", True),
    ("\"Status? [You Change]\" starts from the AI's Final Score:  >=80 Apply ASAP  ·  70-79 Apply If Time  ·  60-69 Backup Lane  ·  <60 Or Skip It.  Pick a new value from the dropdown as each application progresses.", False),
    ("\"On Ice\" means you already applied to a different role at that company — don't double-apply.", False),
    ("", False),
    ("Grouping & sorting", True),
    ("To group by stage, sort by the Status column (use the header filter button, or Data > Sort). The job rows have no merged cells, so sorting works cleanly.", False),
    ("Below the jobs is a legend of section colors. If you like visual breaks between groups, copy a bar in above a group after sorting — optional.", False),
    ("Pasting jobs into your own tracker? Copy only the job rows, NOT the legend bars, so you don't end up with duplicate dividers.", False),
    ("", False),
    ("ATS dates", True),
    ("\"ATS First Posted Date\" is the date the EMPLOYER's job board first published the posting — not the date this batch was fetched. When no verified ATS date exists it reads \"Unknown\" rather than sitting blank, and it is never guessed from a URL, job id, or search-result age.", False),
    ("\"ATS Last Updated Date\" is the job board's own last-updated timestamp, when it publishes one. It follows the SAME rule as the column beside it: when the ATS provides no update date it reads \"Unknown\", never blank — some boards never expose the field at all, and some postings have simply never been updated. Neither date is ever substituted with JAIL's fetch date, a file date, a deadline, or anything inferred from the posting's wording.", False),
    ("Both are fixed dates on purpose: an age or \"days open\" number baked into a saved sheet goes stale and starts misleading you. Neither column is ever left blank, and neither date feeds scoring, ranking, or prioritization.", False),
    ("", False),
    ("Colors & the dropdown", True),
    ("Status, Lane, Lane Fit, Comp Fit, Comp Range and Data Completeness cells are color-coded. Working Location uses one fixed 4-color palette: green = remote genuinely available; yellow = acceptable home-metro office at exactly 1-3 days; orange = unknown location/cadence, >3 days, or open-ended minimums (\"3+ days\"); red = required in-person outside your home geography. In Google Sheets you can layer the native rounded \"chip\" dropdown on top if you prefer that look (that is a Sheets feature, not part of the file).", False),
    ("", False),
    ("AI detail columns", True),
    ("The score block and the notes columns to its right are AI-generated.  \"Lane\" = what the job actually is;  \"Lane Fit\" = how that maps to your target lanes.  \"Comp + Lifestyle Fit Notes\" is the rationale/audit trail behind that score.", False),
    ("", False),
    ("CSV companion", True),
    ("A matching .csv has the same columns + job rows (no colors, dropdown, legend, or tabs).", False),
]


def build_score_section(meta):
    """The dynamic 'How to read these rankings' section — labels, weights, and definitions come from
    the per-run meta (score-dimensions.json defaults, or the candidate's scoring card where weights
    override it), NOT hardcoded here, so this stays correct if a user changes their scoring system."""
    order = meta.get("order") or DEFAULT_METADATA["order"]
    rows = [("How to read these rankings", True),
            ("Jobs are scored across four independent dimensions. A high or low score in one area should not automatically pull the other scores toward it.", False),
            ("", False)]
    for key in order:
        dim = meta.get(key) or DEFAULT_METADATA.get(key, {})
        label = dim.get("label", key)
        weight = dim.get("weight_pct")
        header = f"{label}" + (f"  (weight: {weight}%)" if key != "final" and weight is not None else "")
        rows.append((header, True))
        if dim.get("definition"):
            rows.append((dim["definition"], False))
        rows.append(("", False))
    return rows


def build_instructions(ws, meta=None, archive_link=None, archive_warning=None,
                       completeness_line=None, completeness_attention=False, headers=None):
    meta = meta or DEFAULT_METADATA
    ws.column_dimensions["A"].width = 110
    t = ws.cell(1, 1, "How to use this tracker")
    t.font = Font(name=FONT, bold=True, size=14, color="305496")
    r = 3
    # Loud data-completeness flag at the very top — mirrors the banner on the Job Rankings sheet.
    if completeness_line:
        c = ws.cell(r, 1, completeness_line)
        c.font = Font(name=FONT, bold=True, size=11, color=("C00000" if completeness_attention else "375623"))
        c.fill = solid("FCE4D6" if completeness_attention else "E2EFDA")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        r += 2
    guide = build_column_guide(headers) if headers else []
    for text, is_header in INSTRUCTIONS + guide + build_score_section(meta):
        c = ws.cell(r, 1, text)
        c.font = Font(name=FONT, bold=is_header, size=11, color=("305496" if is_header else "000000"))
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        r += 1
    r += 1
    hdr = ws.cell(r, 1, "Submitted Applications")
    hdr.font = Font(name=FONT, bold=True, size=11, color="305496")
    r += 1
    if archive_link:
        link_cell = ws.cell(r, 1, archive_link)
        link_cell.hyperlink = f"file://{archive_link}"
        link_cell.font = Font(name=FONT, size=11, color="0563C1", underline="single")
        link_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    else:
        note = ws.cell(r, 1, archive_warning or "Submitted-applications link not configured.")
        note.font = Font(name=FONT, italic=True, size=10, color=("C00000" if archive_warning else "808080"))
        note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def build(input_csv, output_xlsx, config_path=None, quarantined=0):
    if not HAVE_OPENPYXL:
        raise RuntimeError("openpyxl is required to build the .xlsx (install requirements.txt in the venv).")
    cfg = load_config(config_path)
    meta = load_meta(input_csv)
    order = meta.get("order") or DEFAULT_METADATA["order"]
    # Effective score-column header strings for THIS run, in the requested display order — resolved
    # from meta (the scoring card's actual weights + score-dimensions.json's labels), not hardcoded.
    label_of = {k: meta[k]["label"] for k in order}
    H_FINAL = label_of["final"]
    SUB_SCORE_COLS = [label_of[k] for k in order if k != "final"]
    SCORE_COLS = [H_FINAL] + SUB_SCORE_COLS
    for key in order:
        WIDTHS[label_of[key]] = SCORE_WIDTHS_BY_KEY.get(key, 14)

    # ---- ONE canonical row collection (B3, divergence prevention) --------------- #
    # Every repair the workbook needs — migration, WL/lane/comp/status normalization,
    # posted-date + completeness back-fill — is applied to the CSV ITSELF first, via
    # the shared normalize pass. The workbook is then built from the repaired CSV, so
    # the two artifacts can never diverge: the class where the XLSX showed a repaired
    # value (a back-filled posted date) that the CSV lacked is structurally gone.
    # vet-jobs.js runs the same pass right after scoring, making this a no-op there;
    # it does the work for a standalone regeneration of an old or hand-made CSV.
    norm_contracts.normalize_rankings_csv(
        str(input_csv), cfg,
        score_labels={k: label_of.get(k) for k in norm_contracts.SCORE_SLOTS})

    headers, records = read_records(input_csv, SCORE_COLS)
    # Residual on-read normalization (belt-and-suspenders; a no-op after the pass above).
    for rec in records:
        if H_WORKLOC in rec:
            rec[H_WORKLOC] = norm_contracts.normalize_working_location(rec.get(H_WORKLOC), cfg)
        if H_COMPRANGE in rec:
            rec[H_COMPRANGE] = norm_contracts.normalize_comp_range(rec.get(H_COMPRANGE))
            # Re-derive Comp Fit from the normalized range (midpoint rule) so old CSVs
            # written under the superseded high-endpoint-only rule regenerate with
            # honest colors. Comp colors keep the COMP_LABEL_COLORS palette.
            if H_COMPFIT in rec:
                rec[H_COMPFIT] = norm_contracts.comp_fit_label(rec.get(H_COMPRANGE), cfg)
        if H_LANE in rec:
            # Lane taxonomy repair ("Work Tools - X" -> "Work - X"); Lane Fit is
            # candidate data and is deliberately NOT touched.
            rec[H_LANE] = norm_contracts.normalize_lane(rec.get(H_LANE))
        # Status spelling drift ("Apply if Time" -> "Apply If Time"). A near-miss silently costs the
        # cell its dropdown match, its lifecycle fill and its filter grouping, so repair it here too
        # — but only the spelling. The NEEDS_REFETCH sentinel is deliberately not a tracker value and
        # is left alone by normalize_status (it warns, never rewrites, an unrecognized value).
        if H_STATUS in rec and (rec.get(H_STATUS) or "").strip() != NEEDS_REFETCH_STATUS:
            rec[H_STATUS] = norm_contracts.normalize_status(rec.get(H_STATUS))
    # ---- Migrate to the 28-column contract, then back-fill the derived columns ----
    # The CSV migration pass (norm_contracts.normalize_rankings_csv, which vet-jobs.js runs right
    # after scoring) normally leaves nothing to do here — for a migrated CSV this is a no-op. It
    # stays for a HAND-MADE or never-normalized CSV, which must still regenerate a correct
    # workbook: same renames, same drop, same inserts, same order, ONE shared definition.
    table = [list(headers)] + [[rec.get(h, "") for h in headers] for rec in records]
    if norm_contracts.migrate_rankings_headers(
            table, score_labels={k: label_of.get(k) for k in norm_contracts.SCORE_SLOTS}):
        headers = [str(h).strip() for h in table[0]]
        records = [dict(zip(headers, row)) for row in table[1:]]
    base_dir = Path(input_csv).resolve().parent
    for rec in records:
        # Never blank: a verified employer date from the capture, else the `Unknown` placeholder.
        if (rec.get(H_POSTED) or "").strip() in ("", norm_contracts.UNKNOWN_POSTED_DATE):
            rec[H_POSTED] = (norm_contracts.posted_date_from_capture(
                rec.get(H_JOBFILE), base_dir=base_dir) or norm_contracts.UNKNOWN_POSTED_DATE)
        # Every row colored: back-fill a blank completeness cell with the shared helper.
        if not (rec.get(H_DATACOMPLETE) or "").strip():
            rec[H_DATACOMPLETE] = fallback_completeness(rec.get(H_COMPRANGE), rec.get(H_WORKLOC))
        rec.setdefault(H_PRACTNOTES, "")
    ncols = len(headers)
    letter = {h: get_column_letter(i + 1) for i, h in enumerate(headers)}
    legend_letter = get_column_letter(LEGEND_MERGE_TO)

    wb = Workbook()
    ws = wb.active
    ws.title = "Job Rankings"
    archive_link, archive_warning = resolve_submitted_applications_link(cfg)
    compl_line, compl_attention = completeness_summary(records)
    build_instructions(wb.create_sheet("Instructions"), meta=meta, archive_link=archive_link,
                       archive_warning=archive_warning, completeness_line=compl_line,
                       completeness_attention=compl_attention, headers=headers)   # second tab

    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)
    LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
    # Free-text columns read left-to-right like prose — left-align them (Jessica, 7/16/26). Everything
    # else (dates, scores, short fit labels) stays centered.
    LEFT_ALIGN_HEADERS = {
        H_LANE, H_COMPANY, H_TITLE, H_WORKLOC, H_POSTED, H_UPDATED,
        "Have Intro? [You Add]", "Your Notes? [You Add]", "Decline/Down Date? [You Add]",
        H_PRACTNOTES, "Your Desire Score Notes", "Profile Score Notes", "Top Reasons Notes",
        "Top Concerns Notes",
        "Job File", "Tailored? (Base Resume)", H_LANEFIT, H_COMPFIT,
    }
    base_font = Font(name=FONT, size=10, color="000000")
    link_font = Font(name=FONT, size=10, color="0563C1", underline="single")

    # Header row.
    hdr_fill = solid("305496")
    hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = LEFT_ALIGN if h in LEFT_ALIGN_HEADERS else ALIGN
        cell.border = BORDER
    ws.row_dimensions[1].height = 30

    notes = []
    if quarantined and int(quarantined) > 0:
        notes.append(f"Prep quarantined {quarantined} thin/failed post(s); they were NOT ranked. See '0 - Prep Report/'.")
    if not config_complete_enough(cfg):
        notes.append("Candidate preferences (jail.config.json comp/location) are missing or empty, so Comp Fit is "
                     "neutral and Working Location cannot recognize your home metro (out-of-metro "
                     "offices show orange instead of red). Run /intake or fill jail.config.json.")
    if notes:
        ws[f"{letter[H_STATUS]}1"].comment = Comment("\n".join(notes), "JAIL")

    # Data rows (job rows only — no dividers).
    for i, rec in enumerate(records):
        r = i + 2
        for c, h in enumerate(headers, start=1):
            val = rec.get(h, "")
            if val is None:
                val = ""
            cell = ws.cell(r, c)
            cell.border = BORDER
            cell.alignment = LEFT_ALIGN if h in LEFT_ALIGN_HEADERS else ALIGN
            if h == H_TITLE and val:
                title, url = parse_title_link(str(val))
                cell.value = title
                if url:
                    cell.hyperlink = url
                    cell.font = link_font
                else:
                    cell.font = base_font
            else:
                cell.value = val
                cell.font = base_font
        st = (rec.get(H_STATUS) or "").strip()
        if st in STATUS_COLORS:
            fh, foh = STATUS_COLORS[st]
            ws[f"{letter[H_STATUS]}{r}"].fill = solid(fh)
            ws[f"{letter[H_STATUS]}{r}"].font = Font(name=FONT, size=10, bold=True, color=foh)
        lc = lane_color(rec.get(H_LANE, ""))
        if lc:
            ws[f"{letter[H_LANE]}{r}"].fill = solid(lc)
        lfc = lane_fit_color(rec.get(H_LANEFIT, ""), cfg)
        if lfc:
            ws[f"{letter[H_LANEFIT]}{r}"].fill = solid(lfc)
        cc = comp_color(rec.get(H_COMPFIT, ""))
        ws[f"{letter[H_COMPRANGE]}{r}"].fill = solid(cc)
        ws[f"{letter[H_COMPFIT]}{r}"].fill = solid(cc)
        loc = loc_color(rec.get(H_WORKLOC, ""), cfg)
        ws[f"{letter[H_WORKLOC]}{r}"].fill = solid(loc)
        dcc = completeness_color(rec.get(H_DATACOMPLETE, ""))
        if dcc:
            ws[f"{letter[H_DATACOMPLETE]}{r}"].fill = solid(dcc)

        # HARD-STOP override: wins over every other fill above — none of a NEEDS-RE-FETCH row's
        # data (lane, comp, location, scores) is trustworthy, so the whole row gets flagged, not
        # just the Status cell.
        if st == NEEDS_REFETCH_STATUS:
            for c in range(1, ncols + 1):
                cell = ws.cell(r, c)
                cell.fill = solid(NEEDS_REFETCH_FILL)
                cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")

    n = len(records)
    last_data_row = n + 1  # header is row 1
    for h in headers:
        ws.column_dimensions[letter[h]].width = WIDTHS.get(h, 16)
    ws.freeze_panes = "A2"  # freeze the header row only

    if n >= 1:
        # Header filter buttons over the job rows only (legend below is outside this range).
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{last_data_row}"
        # Status dropdown — inline list (shows the arrow), applied to the job rows only.
        dv = DataValidation(type="list", formula1=inline_list_formula(STATUS_VALUES),
                            showDropDown=False, allowBlank=True)
        ws.add_data_validation(dv)
        dv.add(f"{letter[H_STATUS]}2:{letter[H_STATUS]}{last_data_row}")

        # Score color ramps.
        for s in SUB_SCORE_COLS:
            col = letter[s]
            rng = f"{col}2:{col}{last_data_row}"
            for rule in [
                CellIsRule(operator="greaterThanOrEqual", formula=["85"], fill=solid(SCORE_BUCKETS[0][1]), stopIfTrue=True),
                CellIsRule(operator="between", formula=["80", "84"], fill=solid(SCORE_BUCKETS[1][1]), stopIfTrue=True),
                CellIsRule(operator="between", formula=["75", "79"], fill=solid(SCORE_BUCKETS[2][1]), stopIfTrue=True),
                CellIsRule(operator="between", formula=["70", "74"], fill=solid(SCORE_BUCKETS[3][1]), stopIfTrue=True),
                CellIsRule(operator="between", formula=["65", "69"], fill=solid(SCORE_BUCKETS[4][1]), stopIfTrue=True),
                CellIsRule(operator="between", formula=["60", "64"], fill=solid(SCORE_BUCKETS[5][1]), stopIfTrue=True),
                CellIsRule(operator="between", formula=["50", "59"], fill=solid(SCORE_BUCKETS[6][1]), stopIfTrue=True),
                CellIsRule(operator="lessThan", formula=["50"], fill=solid(SCORE_BUCKETS[7][1]), stopIfTrue=True),
            ]:
                ws.conditional_formatting.add(rng, rule)
        fcol = letter[H_FINAL]
        frng = f"{fcol}2:{fcol}{last_data_row}"
        for rule in [
            CellIsRule(operator="greaterThanOrEqual", formula=["80"], fill=solid(FINAL_STRONG), stopIfTrue=True),
            CellIsRule(operator="between", formula=["70", "79"], fill=solid(FINAL_MAYBE), stopIfTrue=True),
            CellIsRule(operator="between", formula=["60", "69"], fill=solid(FINAL_WEAK), stopIfTrue=True),
            CellIsRule(operator="lessThan", formula=["60"], fill=solid(FINAL_SKIP), stopIfTrue=True),
        ]:
            ws.conditional_formatting.add(frng, rule)

    # Loud DATA-COMPLETENESS flag — a merged banner just below the jobs, above the section legend,
    # so the reader sees at a glance which rows' scores were computed against comp/location that could
    # NOT be verified (skips pure "not posted" — a benign employer omission). Same summary the Markdown
    # rankings carry at the top; also mirrored on the Instructions tab.
    summ, has_attention = completeness_summary(records)
    cr = last_data_row + 2
    fill_hex, font_hex = ("FCE4D6", "C00000") if has_attention else ("E2EFDA", "375623")
    banner = ws.cell(cr, 1, summ)
    banner.fill = solid(fill_hex)
    banner.font = Font(name=FONT, bold=True, size=11, color=font_hex)
    banner.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(f"A{cr}:{legend_letter}{cr}")
    ws.row_dimensions[cr].height = 30

    # Section-color legend — a SEPARATE block below the jobs (blank-row gap), merged A:J. Not mixed
    # into the sortable data; a palette to copy a bar from after sorting if you want visual breaks.
    lr = cr + 2
    lbl = ws.cell(lr, 1, "Section colors  (optional — after sorting by Status, copy a bar in above a group; do NOT paste these into your own tracker's data)")
    lbl.font = Font(name=FONT, size=9, italic=True, color="808080")
    lbl.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"A{lr}:{legend_letter}{lr}")
    for j, (label, (fill_hex, font_hex)) in enumerate(SECTION_COLORS.items()):
        br = lr + 1 + j
        ws.merge_cells(f"A{br}:{legend_letter}{br}")
        cell = ws.cell(br, 1, label)
        cell.fill = solid(fill_hex)
        cell.font = Font(name=FONT, bold=True, size=11, color=font_hex)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[br].height = 20

    wb.save(output_xlsx)
    return n


def main(argv):
    parser = argparse.ArgumentParser(description="Build the job-search tracker .xlsx from a rankings CSV.")
    parser.add_argument("input_csv")
    parser.add_argument("output_xlsx", nargs="?", default=None)
    parser.add_argument("--config", default=None, help="path to jail.config.json (candidate preferences)")
    parser.add_argument("--quarantined", type=int, default=0, help="count of prep-quarantined posts (note only)")
    args = parser.parse_args(argv[1:])
    out = args.output_xlsx or re.sub(r"\.csv$", ".xlsx", args.input_csv, flags=re.I)
    if out == args.input_csv:
        out = args.input_csv + ".xlsx"
    n = build(args.input_csv, out, config_path=args.config, quarantined=args.quarantined)
    print(f"Wrote {out} ({n} jobs).")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
