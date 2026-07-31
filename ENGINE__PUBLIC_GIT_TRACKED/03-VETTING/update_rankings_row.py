#!/usr/bin/env python3
"""Write a tailor/cover-letter result BACK into a batch's rankings CSV + XLSX.

Why this exists: vet-jobs.js writes "Base Resume Used" as a blank at vet time with the comment
"filled later by the tailor step" — but nothing ever filled it. The tailor agent already returns
`recommended_base`; it was simply discarded. The column was therefore empty in every batch ever
produced, and had to be reconstructed by hand from the per-job `application_resume_output*.md`
files. Same story for "Cover Letter?" — there was no way to see which jobs had a letter written.

This closes that handoff. Both the tailor step and the cover-letter step call it per job.

Matching: by canonical URL first, then job-file name. URL is primary because the same posting is
often re-fetched under different filenames across batches (e.g. Everyday Health exists as
`...__pm.txt` and `...__everyday-health.txt`; Google as `google__...` and `product-manager-...`)
— all of which resolve to one canonical URL via prep_common.normalize_url.

Both CSV and XLSX are edited IN PLACE (never regenerated) so the user's own manual edits,
formatting, and column renames survive.

    python update_rankings_row.py --batch "<batch dir>" --job-file "airtable__product-manager.txt" \
        [--url "https://..."] [--base "Anthropic — PM, Consumer (6/25/26)"] [--cover-letter] \
        [--base-score 70 --improved-score 85 --why "..."]

The resume-comparison block (--base-score/--improved-score/--why) is written the same in-place way,
and is VALIDATED BEFORE anything is written: a block that violates the contract (not a multiple of
five, out of range, improved below base, a delta that does not match, a missing reason) aborts the
whole run rather than writing a half-correct row into the tracker. `--delta` is derived, not
accepted, so the stored value can never disagree with its two operands.
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import norm_contracts  # noqa: E402 — the contract owns the column names + their invariants

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "02-PREP"))
try:
    from prep_common import normalize_url
except Exception:  # prep deps unavailable — fall back to exact string compare
    def normalize_url(u):
        return (u or "").strip().lower().rstrip("/")

# Column headers are matched by PREFIX, not equality: users rename them locally (e.g. a trailing
# "- <custom field>" suffix). Prefix-matching keeps that working.
#
# BOTH header generations are accepted on READ, so this keeps working against an older rankings
# file that still says "Base Resume Used" / "Cover Letter?" while writing the CURRENT contract
# names. Without the legacy fallback the tailor and cover-letter workflows would silently write
# nothing into an unmigrated file — the exact failure this script exists to fix.
H_BASE_PREFIXES = ("tailored? (base resume)", "tailored?", "base resume used")
H_COVER_PREFIXES = ("cover letter drafted?", "cover letter")
# The contract names this script writes when it has to CREATE the column.
H_BASE_CURRENT = "Tailored? (Base Resume)"
H_COVER_CURRENT = "Cover Letter Drafted?"
# Semantics (contract, 2026-07-30): both columns stay BLANK until the step completes —
# `Tailored? (Base Resume)` then holds the exact base-resume name, `Cover Letter Drafted?`
# the literal `Yes` (not "Y": the header asks a question, so the answer reads as one).
COVER_LETTER_YES = "Yes"
H_JOBFILE = "job file"
H_TITLE_PREFIX = "job post title"

# The resume-comparison block. Unlike the two columns above, these have no legacy spelling — they
# are new in the 32-column contract — so each is matched on its own exact name, lowercased.
H_RESUME_COLS = norm_contracts.RESUME_COMPARISON_COLUMNS
H_BASE_SCORE = norm_contracts.H_BASE_SCORE
H_IMPROVED_SCORE = norm_contracts.H_IMPROVED_SCORE
H_DELTA = norm_contracts.H_DELTA
H_WHY_IMPROVES = norm_contracts.H_WHY_IMPROVES

MONTHS = {m: i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"], 1)}

# A base name terminates at its parenthesized date; everything after is prose the agent added
# ("...chassis, merged with", "...adapted significantly", "...copied and renamed to X.pages").
_DATE_END = re.compile(r'^(.*?\((?:\d{1,2}/\d{1,2}/\d{2}|\d{1,2}/\d{2}|[A-Z][a-z]{2,8}\s+\d{4})\))')
_PROSE_CUT = re.compile(
    r'(?:,\s*(?:adapted|merged|chassis|retitled|copied|from the)|\.\s|\s+chassis\b|\s+—\s*see\b)', re.I)
# A leading label, in case a caller passes a whole markdown line rather than the bare value.
_LABEL = re.compile(r'^\s*(?:primary|chosen|recommended|selected)\s+base\s*(?:\(merge\))?\s*:\s*|'
                    r'^\s*base\s+(?:chosen|used|actually used)\s*:\s*', re.I)


def terse_base(s: str) -> str:
    """Normalize a verbose agent-returned base into the tracker's terse house style.

    'Anthropic — PM, Consumer (6/25/26), copied and renamed to Jessica-Barnett-Resume - ....pages'
        -> 'Anthropic — PM, Consumer (6/25/26)'
    'Dropbox Principal PM — Teams & Collab (Jan 2026)'  -> 'Dropbox Principal PM — Teams & Collab (1/26)'
    """
    s = re.sub(r'\*\*', '', (s or "").strip())
    s = re.sub(r'^\s*[-*•]\s+', '', s)   # leading markdown bullet
    s = _LABEL.sub('', s).strip()
    # Collapse a date-paren that carries extra words INSIDE it: "(7/1/26 finalized submission)"
    # -> "(7/1/26)". Do this BEFORE _DATE_END so the paren is a clean date to anchor on.
    s = re.sub(r'\((\d{1,2}/\d{1,2}/\d{2}|\d{1,2}/\d{2})[^)]*\)', r'(\1)', s)
    m = _DATE_END.match(s)
    if m:
        s = m.group(1)
    else:
        # No parenthesized date to anchor on. If there's a BARE date ("..., 6/11/26 (canonical...)"),
        # truncate right after it — everything past the date is trailing prose. Otherwise fall back
        # to cutting at the first prose connector.
        bare = re.search(r'\d{1,2}/\d{1,2}/\d{2}|\d{1,2}/\d{2}', s)
        s = s[:bare.end()] if bare else _PROSE_CUT.split(s, maxsplit=1)[0]
    s = s.strip().rstrip('.,').strip()

    def _d(mm):
        mon = mm.group(1).lower()[:3]
        return f"({MONTHS[mon]}/{mm.group(2)[-2:]})" if mon in MONTHS else mm.group(0)

    s = re.sub(r'\(([A-Z][a-z]{2,8})\s+(\d{4})\)', _d, s)
    return s.replace("Professional Services", "Prof. Services")


def _col(headers, prefixes):
    """Index of the first header matching ANY accepted prefix (current contract name first,
    then the legacy spelling), or None. `prefixes` may be a single string."""
    if isinstance(prefixes, str):
        prefixes = (prefixes,)
    for prefix in prefixes:
        for i, h in enumerate(headers):
            if (h or "").strip().lower().startswith(prefix):
                return i
    return None


def _url_from_title(cell_value: str):
    m = re.search(r'https?://\S+', str(cell_value or ""))
    return m.group(0) if m else None


def _is_match(row_jobfile, row_title, job_file, url):
    if url:
        row_url = _url_from_title(row_title)
        if row_url and normalize_url(row_url) == normalize_url(url):
            return True
    return bool(job_file) and (row_jobfile or "").strip() == job_file.strip()


def build_plan(base, cover_letter, resume):
    """The writes this run intends, as (accepted_prefixes, name_to_create_if_absent, value).

    One list drives the CSV, the XLSX and the Markdown, so the three artifacts cannot diverge by
    someone remembering to extend two of them. `resume` is the validated 4-tuple
    (base_score, improved_score, delta, why) or None."""
    plan = []
    if base:
        plan.append((H_BASE_PREFIXES, H_BASE_CURRENT, base))
    if cover_letter:
        # Created under the LEGACY name deliberately: this only fires on a pre-contract file, and
        # matching that file's own generation keeps a later migration a pure rename.
        plan.append((H_COVER_PREFIXES, "Cover Letter?", COVER_LETTER_YES))
    if resume:
        for name, value in zip(H_RESUME_COLS, resume):
            plan.append(((name.lower(),), name, str(value)))
    return plan


def update_csv(path: Path, job_file, url, base, cover_letter, resume=None) -> bool:
    rows = list(csv.reader(path.open(newline="", encoding="utf-8")))
    if not rows:
        return False
    headers = rows[0]
    plan = build_plan(base, cover_letter, resume)
    ci_jf, ci_title = _col(headers, H_JOBFILE), _col(headers, H_TITLE_PREFIX)

    # Resolve every target column, appending any the file predates. Appending (rather than
    # inserting at the contract position) is what keeps this edit non-destructive: no existing
    # column's data shifts, and the normalize pass reorders to the contract on the next full run.
    targets = []
    for prefixes, create_name, value in plan:
        ci = _col(headers, prefixes)
        if ci is None:
            headers.append(create_name)
            ci = len(headers) - 1
        targets.append((ci, value))
    for r in rows[1:]:
        if len(r) < len(headers):
            r.extend([""] * (len(headers) - len(r)))

    hit = False
    for r in rows[1:]:
        jf = r[ci_jf] if ci_jf is not None and ci_jf < len(r) else ""
        ti = r[ci_title] if ci_title is not None and ci_title < len(r) else ""
        if not _is_match(jf, ti, job_file, url):
            continue
        hit = True
        for ci, value in targets:
            r[ci] = value
    if hit:
        with path.open("w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(rows)
    return hit


def _style_new_columns(ws, created):
    """Make a column added to an EXISTING workbook look native.

    This path exists because the tracker is hand-finished and must not be regenerated to gain a
    column — regenerating would rebuild the workbook and discard the candidate's own formatting.
    So the new column copies the header style already in the sheet, takes its width and its
    conditional-format ramp from the generator, and inherits the body cells' alignment/border
    from its left-hand neighbour. Without this the four columns land as bare white cells against
    a fully styled sheet, which reads as breakage."""
    try:
        from copy import copy
        from openpyxl.utils import get_column_letter
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        import make_rankings_xlsx as mrx
    except Exception:
        return
    last_row = ws.max_row
    template_hdr = ws.cell(1, 1)
    for ci, name in created:
        col = ci + 1
        hdr = ws.cell(1, col)
        for attr in ("fill", "font", "alignment", "border"):
            try:
                setattr(hdr, attr, copy(getattr(template_hdr, attr)))
            except Exception:
                pass
        try:
            ws.column_dimensions[get_column_letter(col)].width = mrx.WIDTHS.get(name, 16)
        except Exception:
            pass
        # Body cells inherit from the neighbour so wrap/centre/borders stay consistent.
        neighbour = ws.cell(2, col - 1) if col > 1 else None
        if neighbour is not None:
            for r in range(2, last_row + 1):
                cell = ws.cell(r, col)
                for attr in ("alignment", "border", "font"):
                    try:
                        setattr(cell, attr, copy(getattr(neighbour, attr)))
                    except Exception:
                        pass
        if name in norm_contracts.RESUME_SCORE_COLUMNS and last_row >= 2:
            try:
                mrx.apply_resume_comparison_formatting(
                    ws, get_column_letter(col), name, last_row)
            except Exception:
                pass
    # Widen the auto-filter to cover the new columns (rows unchanged, so the legend block below
    # the jobs stays outside it). Otherwise the added columns have no filter button and the
    # sheet's one-click "Sort A->Z" silently stops short of them.
    try:
        ref = ws.auto_filter.ref
        if ref and ":" in ref:
            start, end = ref.split(":")
            row_end = "".join(ch for ch in end if ch.isdigit())
            ws.auto_filter.ref = f"{start}:{get_column_letter(ws.max_column)}{row_end}"
    except Exception:
        pass


def update_xlsx(path: Path, job_file, url, base, cover_letter, resume=None) -> bool:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False
    wb = load_workbook(path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    plan = build_plan(base, cover_letter, resume)
    ci_jf, ci_title = _col(headers, H_JOBFILE), _col(headers, H_TITLE_PREFIX)

    targets, created = [], []
    for prefixes, create_name, value in plan:
        ci = _col(headers, prefixes)
        if ci is None:
            ci = len(headers)
            headers.append(create_name)
            ws.cell(1, ci + 1, create_name)
            created.append((ci, create_name))
        targets.append((ci, value))
    if created:
        _style_new_columns(ws, created)

    hit = False
    for r in range(2, ws.max_row + 1):
        jf = ws.cell(r, ci_jf + 1).value if ci_jf is not None else ""
        tcell = ws.cell(r, ci_title + 1) if ci_title is not None else None
        # Prefer the real hyperlink target over the display text.
        ti = ""
        if tcell is not None:
            ti = (tcell.hyperlink.target if tcell.hyperlink and tcell.hyperlink.target
                  else str(tcell.value or ""))
        if not _is_match(jf, ti, job_file, url):
            continue
        hit = True
        for ci, value in targets:
            # Scores go in as numbers so the conditional-format ramp and any sort behave; the
            # prose column stays text.
            cell_value = int(value) if str(value).lstrip("-").isdigit() else value
            ws.cell(r, ci + 1, cell_value)
    if hit:
        wb.save(path)
    return hit


# --------------------------------------------------------------------------- #
# Markdown — the third artifact, and the one that silently goes stale.
#
# The rankings .md is generated independently in vet-jobs.js and gets no normalization pass, so
# unlike CSV<->XLSX it has no structural guarantee of agreeing with the others. Rather than leave
# it to drift, the same write updates it: find the job's block by its `- **File:** <job file>`
# line (the one unambiguous identity line in the block) and insert or replace the summary line
# right after it.
# --------------------------------------------------------------------------- #
_MD_LINE_PREFIX = "- **Resume improvement:**"


def md_line(resume) -> str:
    b, i, d, why = resume
    return f"{_MD_LINE_PREFIX} base {b} → improved {i} (delta +{d}).{(' ' + why) if why else ''}"


def update_md(path: Path, job_file, resume) -> bool:
    if not resume or not job_file:
        return False
    lines = path.read_text(encoding="utf-8").split("\n")
    target = f"- **File:** {job_file}".strip()
    out, hit, drop_stale = [], False, False
    for line in lines:
        # Only a stale line belonging to THIS job is replaced — it sits immediately under the
        # File line we just matched. Every other job's line is left alone.
        if drop_stale:
            drop_stale = False
            if line.strip().startswith(_MD_LINE_PREFIX):
                continue
        out.append(line)
        if line.strip() == target:
            out.append(md_line(resume))
            hit, drop_stale = True, True
    if hit:
        path.write_text("\n".join(out), encoding="utf-8")
    return hit


def main(argv):
    ap = argparse.ArgumentParser(description="Write a tailor/cover-letter result back into the rankings.")
    ap.add_argument("--batch", required=True, help="Batch root, e.g. __READY_TO_REVIEW__PRIVATE_GITIGNORED/07-16-26")
    ap.add_argument("--job-file", default=None, help="Job .txt filename (fallback match key)")
    ap.add_argument("--url", default=None, help="Canonical job URL (primary match key)")
    ap.add_argument("--base", default=None, help="Resume base used (normalized to house style)")
    ap.add_argument("--cover-letter", action="store_true", help="Mark this row's Cover Letter? as Y")
    ap.add_argument("--base-score", type=int, default=None,
                    help="Base Resume Score (0-100, multiple of 5)")
    ap.add_argument("--improved-score", type=int, default=None,
                    help="Improved Resume Score (0-100, multiple of 5, >= base)")
    ap.add_argument("--why", default=None,
                    help="Why It Improves — one or two sentences naming the changes behind the delta")
    a = ap.parse_args(argv[1:])

    # The resume-comparison block is all-or-nothing, and its delta is DERIVED here rather than
    # accepted from the caller: a stored delta that disagrees with its operands is the one defect
    # a reader of the spreadsheet cannot detect by eye.
    resume = None
    given = [x is not None for x in (a.base_score, a.improved_score)]
    if any(given) or a.why is not None:
        if not all(given):
            raise SystemExit(
                "The resume-comparison block is written as a unit: pass BOTH --base-score and "
                "--improved-score (and --why). Leave all three off when the base could not be "
                "read — a blank block is the correct way to say 'not scored'.")
        delta = a.improved_score - a.base_score
        errs = norm_contracts.validate_resume_comparison(
            a.base_score, a.improved_score, delta, a.why)
        if errs:
            # Abort BEFORE any file is touched: a partially-written block would look like a
            # completed comparison whose improvement happened to be nothing.
            raise SystemExit("Refusing to write an invalid resume-comparison block:\n  - "
                             + "\n  - ".join(errs))
        resume = (a.base_score, a.improved_score, delta, a.why)

    if not a.base and not a.cover_letter and resume is None:
        raise SystemExit("Nothing to write: pass --base, --cover-letter, and/or the "
                         "--base-score/--improved-score/--why block.")
    if not a.job_file and not a.url:
        raise SystemExit("Need at least one match key: --job-file and/or --url.")

    rankings = Path(a.batch) / "1 - Rankings"
    if not rankings.is_dir():
        # LOUD, not a quiet note: this is the silent-no-op path. A tailor run whose batch
        # folder could not be resolved (e.g. it fell through to "manual/") lands here, and
        # the workflow used to report success while nothing was written anywhere.
        print(f"WARNING: no '1 - Rankings/' folder in {a.batch} — the Tailored?/Cover Letter "
              f"columns were NOT updated. The batch may be misrouted (check that the job file "
              f"lives under the batch's '3 - Source Material/'), or this job was tailored "
              f"outside any batch.")
        return 0

    base = terse_base(a.base) if a.base else None
    touched = []
    for p in sorted(rankings.glob("*-rankings.csv")):
        if update_csv(p, a.job_file, a.url, base, a.cover_letter, resume):
            touched.append(p.name)
    # Any .xlsx in the folder, not just `*-rankings.xlsx`: a real batch's workbook is often
    # renamed by hand (e.g. "UNIFIED-55-JOB-TRACKER-FINAL.xlsx"), and the narrower glob silently
    # skipped it — leaving the CSV updated and the spreadsheet the user actually reads untouched.
    for p in sorted(rankings.glob("*.xlsx")):
        if p.name.startswith("~$"):
            continue  # Excel lock file
        if update_xlsx(p, a.job_file, a.url, base, a.cover_letter, resume):
            touched.append(p.name)
    for p in sorted(rankings.glob("*-rankings.md")):
        if update_md(p, a.job_file, resume):
            touched.append(p.name)

    key = a.url or a.job_file
    if touched:
        bits = []
        if base:
            bits.append(f'base="{base}"')
        if a.cover_letter:
            bits.append("cover_letter=Y")
        if resume:
            bits.append(f"base_score={resume[0]} improved={resume[1]} delta=+{resume[2]}")
        print(f"Updated {', '.join(touched)} for {key}: {' '.join(bits)}")
    else:
        # Loud, not silent — a miss here is exactly how the column silently stayed empty before.
        print(f"WARNING: no rankings row matched {key} — '{rankings}' left unchanged. "
              f"The row may live in a different batch, or the job file/URL may not match.")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
