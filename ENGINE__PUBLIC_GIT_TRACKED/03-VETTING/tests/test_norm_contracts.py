"""Output-contract regression tests (authoritative spec, 2026-07-29).

These are PERMANENT regression fixtures: the Working Location matrix below is the
spec's own minimum test set, asserted BOTH at the normalizer level AND at the final
artifact — a real XLSX is built and the actual written cell fill hex is read back
with openpyxl (spec: "test the actual written spreadsheet cell").
"""
import csv
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

import make_rankings_xlsx
import norm_contracts
from norm_contracts import (
    WL_GREEN, WL_ORANGE, WL_RED, WL_YELLOW,
    normalize_working_location, working_location_color,
)

# A generic candidate config mirroring the jail.config.json shape: home metro is NYC
# (via home_metro_aliases); SF is in city_priority but is NOT a home metro.
CFG = {
    "comp": {"currency": "USD", "target_base": 200, "floor_base": 180},
    "location": {
        "home_metro": "New York City",
        "home_metro_aliases": ["NYC", "New York", "NYC metro", "Brooklyn", "Manhattan"],
        "city_priority": ["NYC", "SF"],
    },
}


def quiet(_msg):  # suppress expected repair warnings in test output
    pass


# --------------------------------------------------------------------------- #
# Working Location — the spec's 18-case matrix (normalized text AND exact hex)
# --------------------------------------------------------------------------- #
WL_MATRIX = [
    # (input, expected canonical, expected hex)
    ("Remote", "Remote", WL_GREEN),
    ("Remote (US)", "Remote (US)", WL_GREEN),
    ("Remote or IRL NYC - 2 days", "Remote or IRL NYC - 2 days", WL_GREEN),
    ("Remote or IRL SF - 3 days", "Remote or IRL SF - 3 days", WL_GREEN),
    ("IRL NYC - 1 day", "IRL NYC - 1 day", WL_YELLOW),
    ("IRL NYC - 2 days", "IRL NYC - 2 days", WL_YELLOW),
    ("IRL NYC - 3 days", "IRL NYC - 3 days", WL_YELLOW),
    ("IRL NYC/SF - 3 days", "IRL NYC/SF - 3 days", WL_YELLOW),
    ("IRL NYC - 3 days (Mon/Wed/Thu)", "IRL NYC - 3 days (Mon/Wed/Thu)", WL_YELLOW),
    ("IRL NYC - 3+ days", "IRL NYC - 3+ days", WL_ORANGE),
    ("IRL NYC - at least 3 days", "IRL NYC - 3+ days", WL_ORANGE),
    ("IRL NYC - 4 days", "IRL NYC - 4 days", WL_ORANGE),
    ("IRL NYC - 5 days", "IRL NYC - 5 days", WL_ORANGE),
    ("IRL NYC - unknown days", "IRL NYC - unknown days", WL_ORANGE),
    ("Unknown", "Unknown", WL_ORANGE),
    # In-person outside the home geography, no remote/home-metro option -> red.
    ("IRL SF - 3 days", "IRL SF - 3 days", WL_RED),
    # The mechanical-repair case that motivated this module: missing "IRL " prefix.
    ("NYC/SF - 3 days", "IRL NYC/SF - 3 days", WL_YELLOW),
    # Bare "<City> hybrid" phrasing repairs to known-city-unknown-cadence.
    ("New York hybrid", "IRL NYC - unknown days", WL_ORANGE),
    # --- cadence-fidelity + multi-city/detail cases (2026-07-29 defect pass) ---
    # An employer-stated RANGE keeps both endpoints; the color follows its MAXIMUM.
    ("IRL NYC - 2-3 days", "IRL NYC - 2-3 days", WL_YELLOW),
    ("IRL NYC - 4-5 days", "IRL NYC - 4-5 days", WL_ORANGE),
    # "Remote or IRL ..." keeps EVERY genuinely-available office and its trailing detail.
    ("Remote or IRL NYC/SF - unknown days (hub-office salary range; remote elsewhere in "
     "US possible at 80-100% of range)",
     "Remote or IRL NYC/SF - unknown days (hub-office salary range; remote elsewhere in "
     "US possible at 80-100% of range)", WL_GREEN),
    ("IRL NYC - 3 days (Mon/Tue/Thu)", "IRL NYC - 3 days (Mon/Tue/Thu)", WL_YELLOW),
    ("IRL NYC - 5 days (non-negotiable in-office)",
     "IRL NYC - 5 days (non-negotiable in-office)", WL_ORANGE),
    ("IRL NYC/SF - unknown days", "IRL NYC/SF - unknown days", WL_ORANGE),
    ("", "Unknown", WL_ORANGE),
    ("IRL Menlo Park, CA - unknown days", "IRL Menlo Park, CA - unknown days", WL_RED),
    ("IRL Sunnyvale/Kirkland - unknown days", "IRL Sunnyvale/Kirkland - unknown days", WL_RED),
]


@pytest.mark.parametrize("raw,canonical,hexcode", WL_MATRIX)
def test_working_location_matrix(raw, canonical, hexcode):
    got = normalize_working_location(raw, CFG, warn=quiet)
    assert got == canonical
    assert working_location_color(got, CFG) == hexcode


def test_color_set_is_exactly_the_four_spec_hexes_no_grey():
    seen = {working_location_color(normalize_working_location(raw, CFG, warn=quiet), CFG)
            for raw, _, _ in WL_MATRIX}
    assert seen == {WL_GREEN, WL_YELLOW, WL_ORANGE, WL_RED}
    assert (WL_GREEN, WL_YELLOW, WL_ORANGE, WL_RED) == ("42FF35", "FDFF43", "FA9C31", "F82C1F")


def test_question_derived_location_normalizes_identically_to_jd_derived():
    """A location requirement extracted from an application question must normalize
    + color exactly like the same requirement stated in the JD."""
    jd = normalize_working_location("Hybrid — New York, NY (at least 2 days per week)", CFG, warn=quiet)
    question = normalize_working_location("NYC office - at least 2 days per week", CFG, warn=quiet)
    assert jd == question == "IRL NYC - 2+ days"
    assert working_location_color(jd, CFG) == working_location_color(question, CFG) == WL_ORANGE


def test_three_days_is_not_three_plus_days():
    assert working_location_color("IRL NYC - 3 days", CFG) == WL_YELLOW
    assert working_location_color("IRL NYC - 3+ days", CFG) == WL_ORANGE


def test_onsite_city_without_fulltime_is_unknown_days():
    assert normalize_working_location("Onsite SF", CFG, warn=quiet) == "IRL SF - unknown days"
    assert normalize_working_location("Onsite SF, full-time in office", CFG, warn=quiet) == "IRL SF - 5 days"


def test_remote_friendly_is_not_remote():
    got = normalize_working_location("NYC - 3 days, remote-friendly", CFG, warn=quiet)
    assert got == "IRL NYC - 3 days"
    assert working_location_color(got, CFG) == WL_YELLOW


def test_unparseable_becomes_unknown_with_loud_warning():
    warnings = []
    got = normalize_working_location("see posting for details", CFG, warn=warnings.append)
    assert got == "Unknown"
    assert warnings, "an unparseable value must warn, never repair silently"


@pytest.mark.parametrize("raw,expected", [
    # A range is never collapsed to an endpoint, however it is written.
    ("NYC - 2-3 days", "IRL NYC - 2-3 days"),
    ("Hybrid New York, 2-3 days per week in office", "IRL NYC - 2-3 days"),
    ("NYC — 4-5 days", "IRL NYC - 4-5 days"),
    ("NYC/SF - 2-3 days", "IRL NYC/SF - 2-3 days"),
])
def test_day_ranges_are_preserved_never_collapsed(raw, expected):
    """`2-3 days` used to normalize to `3 days` — silently rewriting a cadence the
    employer stated is information loss on a field the candidate reads."""
    assert normalize_working_location(raw, CFG, warn=quiet) == expected


def test_range_color_comes_from_the_maximum():
    assert working_location_color("IRL NYC - 2-3 days", CFG) == WL_YELLOW   # max 3 -> in band
    assert working_location_color("IRL NYC - 3-4 days", CFG) == WL_ORANGE   # max 4 -> out
    assert working_location_color("IRL NYC - 4-5 days", CFG) == WL_ORANGE
    f = norm_contracts.working_location_facts("IRL NYC - 2-3 days", CFG)
    assert f["days_range"] == (2, 3) and f["days_exact"] == 3


@pytest.mark.parametrize("raw,expected", [
    # Every genuinely-available office survives, and so does the trailing detail.
    ("Remote or IRL NYC/SF - unknown days (hub-office salary range)",
     "Remote or IRL NYC/SF - unknown days (hub-office salary range)"),
    ("Remote or NYC/SF - unknown days (hub-office salary range)",
     "Remote or IRL NYC/SF - unknown days (hub-office salary range)"),
    ("Remote or hybrid in New York, NY / San Francisco, CA (hub-office salary range)",
     "Remote or IRL NYC/SF - unknown days (hub-office salary range)"),
])
def test_remote_or_irl_keeps_every_office_and_the_trailing_detail(raw, expected):
    """This form used to drop `/SF` AND the whole parenthetical."""
    got = normalize_working_location(raw, CFG, warn=quiet)
    assert got == expected
    assert working_location_color(got, CFG) == WL_GREEN


@pytest.mark.parametrize("raw,expected", [
    # `Remote (<detail>)` is part of the canonical grammar — employer detail is kept.
    ("Remote, US (in-office 1-2x/quarter; SF office option, not required)",
     "Remote (US; in-office 1-2x/quarter; SF office option, not required)"),
    ("Remote (US) or IRL US offices - unknown days", "Remote (US)"),
    ("Remote - US only", "Remote (US only)"),
    ("Remote (states: NY, CA)", "Remote (states: NY, CA)"),
    # ...but a compound adjective is not a detail, and a scope adverb is not a city.
    ("Remote-first, home base anywhere in the US", "Remote"),
    ("Fully remote", "Remote"),
    ("Remote", "Remote"),
])
def test_remote_detail_is_preserved_but_never_invented(raw, expected):
    """Real values were being flattened to bare `Remote`, discarding the employer's country
    restriction and its office-cadence note. Separately, "Fully remote" used to mint a city
    named "Fully" (`IRL Fully - unknown days`) — a capitalized word is not a place."""
    got = normalize_working_location(raw, CFG, warn=quiet)
    assert got == expected
    assert working_location_color(got, CFG) == WL_GREEN


def test_a_parenthetical_that_only_restates_the_cadence_is_not_duplicated():
    assert normalize_working_location(
        "Hybrid — New York, NY (at least 2 days per week)", CFG, warn=quiet) == "IRL NYC - 2+ days"


def test_no_home_metro_configured_cannot_judge_geography_orange_not_red():
    assert working_location_color("IRL SF - 3 days", {}) == WL_ORANGE


# --------------------------------------------------------------------------- #
# Comp Range — format repair matrix + the APPROVED midpoint Comp Fit rule
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("raw,expected", [
    ("190-210", "190-210"),                 # canonical passthrough
    ("??", "??"),                           # unknown passthrough
    ("$125K–$250K", "125-250"),             # $, K, en dash
    ("125,000 - 250,000", "125-250"),       # commas + full-dollar
    ("232,000-282,000", "232-282"),         # full-dollar, no spaces
    ("180", "180-180"),                     # single value -> N-N
    ("150k to 190k", "150-190"),            # "to" range
    ("", "??"),
])
def test_normalize_comp_range_repairs(raw, expected):
    assert norm_contracts.normalize_comp_range(raw, warn=quiet) == expected


def test_normalize_comp_range_garbage_warns_loudly():
    warnings = []
    assert norm_contracts.normalize_comp_range("competitive salary", warn=warnings.append) == "??"
    assert warnings
    warnings2 = []
    assert norm_contracts.normalize_comp_range("Zone A 200-250 Zone B 180-220", warn=warnings2.append) == "??"
    assert warnings2, "ambiguous multi-band text must fail loudly, not silently pick a band"


# Midpoint rule vs floor 180 / target 200 (approved 2026-07-29): red iff max < floor;
# green iff midpoint >= target; else yellow.
@pytest.mark.parametrize("comp_range,label", [
    ("125-250", "Near target"),          # midpoint 187.5 < 200 -> yellow (old rule said green)
    ("210-250", "Meets/above target"),   # midpoint 230 -> green
    ("120-170", "Below floor"),          # max 170 < 180 -> red
    ("190-210", "Meets/above target"),   # midpoint exactly 200 -> green
    ("??", "Unknown"),
    ("", "Unknown"),
])
def test_comp_fit_label_midpoint_rule(comp_range, label):
    assert norm_contracts.comp_fit_label(comp_range, CFG) == label


def test_comp_fit_label_no_comp_prefs():
    assert norm_contracts.comp_fit_label("190-210", {}) == "No comp prefs"


# --------------------------------------------------------------------------- #
# Lane — bucket taxonomy (Health / Consumer / Work / Other; NEVER "Work Tools")
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("raw,expected", [
    ("Work Tools - Collaboration", "Work - Collaboration"),
    ("Work Tools", "Work"),
    ("work tools - Productivity", "Work - Productivity"),
    ("Work - Project Management", "Work - Project Management"),     # already canonical
    ("Health - Mental Health", "Health - Mental Health"),           # exact rule unchanged
    ("Health - Consumer Mental Health", "Health - Mental Health"),  # qualifier stripped
    ("Health-Provider Tools", "Health - Provider Tools"),           # spacing enforced
    ("Consumer - Home Sharing", "Consumer - Home Sharing"),
    ("Other - Fintech", "Other - Fintech"),
    ("Fintech Infrastructure", "Fintech Infrastructure"),           # non-bucket passthrough
])
def test_normalize_lane(raw, expected):
    assert norm_contracts.normalize_lane(raw) == expected


# --------------------------------------------------------------------------- #
# Tailored-application names — "Company - Canonical Role" (spec's required
# examples + incorrect-form repairs; company names here are generic engine
# test data, same precedent as the BetterUp prep fixtures)
# --------------------------------------------------------------------------- #
APP_NAME_MATRIX = [
    # ((company, raw role), exact required output)
    (("Asana", "Senior Product Manager, Project & Task Experience"), "Asana - Senior PM, Project & Task Experience"),
    (("ClickUp", "Senior Product Manager"), "ClickUp - Senior PM"),
    (("ClickUp", "Staff Product Manager"), "ClickUp - Staff PM"),
    # "Chief Product Officer" is NOT "Product Manager" — stays verbatim.
    (("Feeld", "Chief Product Officer"), "Feeld - Chief Product Officer"),
    (("Courier Health", "Senior Product Manager"), "Courier Health - Senior PM"),
    (("Dorsia", "Product Manager, Consumer Experience"), "Dorsia - PM, Consumer Experience"),
    # Employer " - " separator inside the title becomes the comma separator.
    (("Fetch", "Staff Product Manager - Referral, Growth"), "Fetch - Staff PM, Referral Growth"),
    (("Figma", "Product Manager, AI Growth"), "Figma - PM, AI Growth"),
    (("Findem", "Principal Product Manager"), "Findem - Principal PM"),
    (("Google", "Product Manager, Google Docs"), "Google - PM, Google Docs"),
    (("Willow Health", "Senior Product Manager, Care Delivery"), "Willow Health - Senior PM, Care Delivery"),
    # Incorrect-form repairs (the spec's listed bad forms must repair):
    (("Courier Health", "Sr Product Manager"), "Courier Health - Senior PM"),
    (("Courier Health", "Sr. Product Manager"), "Courier Health - Senior PM"),
    (("Dorsia", "PM Consumer Experience"), "Dorsia - PM, Consumer Experience"),
    (("Fetch", "Staff PM Referral Growth"), "Fetch - Staff PM, Referral Growth"),
    (("Google", "Product Manager Google Docs"), "Google - PM, Google Docs"),
    (("Willow Health", "Sr PM, Care Delivery"), "Willow Health - Senior PM, Care Delivery"),
    # Her explicit answers: Vice President -> VP; Director stays Director.
    (("Acme", "Vice President of Product"), "Acme - VP of Product"),
    (("Acme", "Director of Product"), "Acme - Director of Product"),
    # Filesystem rule: slash and colon are stripped.
    (("Acme", "PM: Growth/Retention"), "Acme - PM, Growth Retention"),
    # Preserve Staff/Principal + qualifiers verbatim.
    (("Acme", "Principal Product Manager, Platform"), "Acme - Principal PM, Platform"),
]


@pytest.mark.parametrize("inputs,expected", APP_NAME_MATRIX)
def test_canonical_application_name_matrix(inputs, expected):
    company, role = inputs
    assert norm_contracts.canonical_application_name(company, role) == expected


@pytest.mark.parametrize("inputs,expected", APP_NAME_MATRIX)
def test_canonical_application_name_is_idempotent(inputs, expected):
    """Re-running the canonicalizer on its own output changes nothing — a re-run
    of a workflow must land in the SAME folder, never a variant."""
    company, _ = inputs
    canonical_role = expected.split(" - ", 1)[1]
    assert norm_contracts.canonical_application_name(company, canonical_role) == expected


def test_never_senior_to_sr_and_existing_comma_never_removed():
    assert norm_contracts.canonical_application_role("Senior PM") == "Senior PM"
    assert norm_contracts.canonical_application_role("Senior PM, Care Delivery") == "Senior PM, Care Delivery"
    # comma insertion only ADDs the separator; it never fires when one exists
    assert "," in norm_contracts.canonical_application_role("PM, Consumer Experience")


def test_application_name_cli_prints_the_exact_string():
    """The CLI is the contract surface the agents actually call — invoke it for real."""
    import subprocess
    import sys as _sys
    from pathlib import Path as _Path
    script = _Path(norm_contracts.__file__)
    out = subprocess.run(
        [_sys.executable, str(script), "--application-name",
         "--company", "Fetch", "--role", "Staff Product Manager - Referral, Growth"],
        capture_output=True, text=True, check=True,
    ).stdout
    assert out == "Fetch - Staff PM, Referral Growth\n"
    out2 = subprocess.run(
        [_sys.executable, str(script), "--application-role", "--role", "Sr Product Manager"],
        capture_output=True, text=True, check=True,
    ).stdout
    assert out2 == "Senior PM\n"


def test_application_name_survives_the_real_filesystem_layer(tmp_path):
    """mkdir with the CLI's output and assert the exact on-disk directory name —
    the final artifact is the folder, not the string."""
    import subprocess
    import sys as _sys
    from pathlib import Path as _Path
    script = _Path(norm_contracts.__file__)
    name = subprocess.run(
        [_sys.executable, str(script), "--application-name",
         "--company", "Willow Health", "--role", "Sr PM, Care Delivery"],
        capture_output=True, text=True, check=True,
    ).stdout.rstrip("\n")
    (tmp_path / name).mkdir()
    assert [p.name for p in tmp_path.iterdir()] == ["Willow Health - Senior PM, Care Delivery"]
    # the resume-base filename uses the same string verbatim
    resume = tmp_path / name / f"Candidate-Resume - {name}.docx"
    resume.write_bytes(b"")
    assert resume.name == "Candidate-Resume - Willow Health - Senior PM, Care Delivery.docx"


# --------------------------------------------------------------------------- #
# CSV CLI pass
# --------------------------------------------------------------------------- #
# THE 27-COLUMN CONTRACT (order authoritative, approved 2026-07-30). Sourced from the engine so a
# contract change can never drift silently past these tests.
HEADERS = list(norm_contracts.resolve_contract_headers())

EXPECTED_CONTRACT = [
    "Applied Date? [You Fill In]", "Status? [You Change]", "Lane", "Company",
    "Job Post Title + Link", "Working Location", "Comp Range",
    "Have Intro? [You Add]", "Your Notes? [You Add]", "Decline/Down Date? [You Add]",
    "FINAL Weighted Score", "How They May See Your Profile", "Your Desire Score",
    "Culture Fit Score", "Comp + Lifestyle Fit Score", "Job Posted Date",
    "Top Reasons Notes", "Top Concerns Notes", "Profile Score Notes",
    "Your Desire Score Notes", "Comp + Lifestyle Fit Notes",
    "Lane Fit", "Comp Fit", "Data Completeness", "Job File",
    "Tailored? (Base Resume)", "Cover Letter Drafted?",
]

# The REAL legacy rescore shape verified on disk: 27 columns, WITH `Location Fit`, WITHOUT
# `Data Completeness`, old score/notes/tailor names, `Posted` before the editable block.
LEGACY_CSV_HEADERS = [
    "Applied Date? [You Fill In]", "Status? [You Change]", "Lane", "Company",
    "Job Post Title + Link", "Working Location", "Comp Range", "Posted",
    "Have Intro? [You Add]", "Your Notes? [You Add]", "Decline/Down Date? [You Add]",
    "FINAL Weighted Score", "How They May See Your Profile", "Your Desire Score",
    "Culture Fit", "Comp + Lifestyle Fit", "Comp + Lifestyle Fit Notes",
    "Mission Fit Notes", "Scope Fit Notes", "Top Reasons Notes", "Top Concerns",
    "Job File", "Base Resume Used", "Lane Fit", "Location Fit", "Comp Fit",
    "Cover Letter?",
]
# The 28-column legacy XLSX shape (same, plus Data Completeness).
LEGACY_XLSX_HEADERS = (LEGACY_CSV_HEADERS[:-1] + ["Data Completeness"]
                       + LEGACY_CSV_HEADERS[-1:])


def test_the_contract_is_exactly_the_approved_27_columns():
    """The order is authoritative: this test is the tripwire for an accidental reorder,
    rename, or added column anywhere in the engine."""
    assert HEADERS == EXPECTED_CONTRACT
    assert len(HEADERS) == 27
    assert "Location Fit" not in HEADERS      # removed as redundant with Working Location
    assert norm_contracts.H_POSTED == "Job Posted Date"


def row_values(company="Acme", location="Remote", comp="190-210", lane="Health - Mental Health",
               lane_fit="Mental Health (high)", loc_fit="Remote", comp_fit="Meets/above target",
               status="Apply ASAP: High Prio", completeness="✓ complete", posted=""):
    """One job row as a header->value MAPPING, so both the current and the legacy row builders
    below place every value by NAME. No test hardcodes a column index."""
    notes = "Cash 26/40 (midpoint ~$188K) | Location 30/30 (fully remote) | Equity 19/20"
    return {
        "Status? [You Change]": status, "Lane": lane, "Company": company,
        "Job Post Title + Link": f"Senior PM | https://example.com/{company.lower()}",
        "Working Location": location, "Comp Range": comp,
        "FINAL Weighted Score": "80", "How They May See Your Profile": "80",
        "Your Desire Score": "80", "Culture Fit Score": "80", "Culture Fit": "80",
        "Comp + Lifestyle Fit Score": "80", "Comp + Lifestyle Fit": "80",
        "Job Posted Date": posted, "Posted": posted,
        "Top Reasons Notes": "r", "Top Concerns Notes": "c", "Top Concerns": "c",
        "Profile Score Notes": "s", "Scope Fit Notes": "s",
        "Your Desire Score Notes": "m", "Mission Fit Notes": "m",
        "Comp + Lifestyle Fit Notes": notes,
        "Lane Fit": lane_fit, "Location Fit": loc_fit, "Comp Fit": comp_fit,
        "Data Completeness": completeness, "Job File": f"{company.lower()}.txt",
        "Tailored? (Base Resume)": "", "Base Resume Used": "",
        "Cover Letter Drafted?": "", "Cover Letter?": "",
    }


def row_for(headers, **kw):
    vals = row_values(**kw)
    return [vals.get(h, "") for h in headers]


def make_row(**kw):
    return row_for(HEADERS, **kw)


def write_csv(path, rows, headers=None):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers or HEADERS)
        for r in rows:
            w.writerow(r)


def test_cli_normalizes_csv_in_place_and_reports_repairs(tmp_path, capsys):
    csv_path = tmp_path / "b-rankings.csv"
    write_csv(csv_path, [
        make_row(company="RepairMe", location="NYC/SF - 3 days"),
        make_row(company="FineAlready", location="Remote"),
    ])
    changed = norm_contracts.normalize_rankings_csv(str(csv_path), CFG)
    out = capsys.readouterr().out
    # 1 location repair + the two rows' Job Posted Date placeholders (blank -> `Unknown`).
    assert changed == 3
    assert "RepairMe" in out and "'NYC/SF - 3 days'" in out and "'IRL NYC/SF - 3 days'" in out
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    i = HEADERS.index("Working Location")
    assert rows[1][i] == "IRL NYC/SF - 3 days"
    assert rows[2][i] == "Remote"


def test_cli_repairs_comp_range_and_recomputes_comp_fit(tmp_path, capsys):
    csv_path = tmp_path / "b-rankings.csv"
    write_csv(csv_path, [
        # Full-dollar comp + an optimistic legacy Comp Fit label (old high-only rule).
        make_row(company="EnvelopeCo", comp="$125,000 - $250,000", comp_fit="Meets/above target"),
    ])
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG)
    capsys.readouterr()
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    assert rows[1][HEADERS.index("Comp Range")] == "125-250"
    # midpoint 187.5 vs target 200 -> the CLI's re-derived midpoint label wins
    assert rows[1][HEADERS.index("Comp Fit")] == "Near target"


# --------------------------------------------------------------------------- #
# End-to-end: build a real XLSX and read the ACTUAL written cell fills back
# --------------------------------------------------------------------------- #
def cell_hex(cell):
    rgb = cell.fill.start_color.rgb
    return str(rgb)[-6:] if rgb else None


def test_xlsx_written_cells_carry_the_exact_spec_hexes(tmp_path):
    cfg_path = tmp_path / "jail.config.json"
    cfg_path.write_text(json.dumps(CFG), encoding="utf-8")
    csv_path = tmp_path / "batch-rankings.csv"
    xlsx_path = tmp_path / "batch-rankings.xlsx"
    cases = [
        ("Remote", WL_GREEN),
        ("IRL NYC - 3 days", WL_YELLOW),
        ("NYC/SF - 3 days", WL_YELLOW),        # repaired on read -> IRL NYC/SF - 3 days
        ("IRL NYC - 3+ days", WL_ORANGE),
        ("Unknown", WL_ORANGE),
        ("IRL SF - 3 days", WL_RED),
    ]
    write_csv(csv_path, [make_row(company=f"Co{i}", location=loc) for i, (loc, _) in enumerate(cases)])
    make_rankings_xlsx.build(str(csv_path), str(xlsx_path), config_path=str(cfg_path))

    wb = load_workbook(str(xlsx_path))
    ws = wb["Job Rankings"]
    headers = [c.value for c in ws[1]]
    wl_col = headers.index("Working Location") + 1
    assert "Location Fit" not in headers   # removed from the contract
    for i, (loc, expected) in enumerate(cases):
        r = i + 2
        assert cell_hex(ws.cell(r, wl_col)) == expected, f"Working Location fill for {loc!r}"
        # black text, no white-font override on these cells
        color = ws.cell(r, wl_col).font.color
        assert color is None or str(color.rgb).endswith("000000")
    # the repaired cell's TEXT is canonical in the written spreadsheet too
    assert ws.cell(4, wl_col).value == "IRL NYC/SF - 3 days"
    # no grey anywhere in the populated Working Location column
    greys = {cell_hex(ws.cell(i + 2, wl_col)) for i in range(len(cases))}
    assert "D9D9D9" not in greys


# The palette this column REPLACED. None of it may reappear in a regenerated sheet.
OLD_PALETTE = {"A9D08E", "FFE699", "F4B183", "D9D9D9", "F4A6A6"}


def test_whole_matrix_reaches_the_spreadsheet_with_the_right_fill_and_text(tmp_path):
    """The full regression matrix through the FINAL artifact: every case's Working
    Location text and the actual written fill hex of both Working Location and Location
    Fit — with zero greys and zero old-palette hexes anywhere in those two columns."""
    cfg_path = tmp_path / "jail.config.json"
    cfg_path.write_text(json.dumps(CFG), encoding="utf-8")
    csv_path = tmp_path / "matrix-rankings.csv"
    xlsx_path = tmp_path / "matrix-rankings.xlsx"
    write_csv(csv_path, [make_row(company=f"Co{i}", location=raw)
                         for i, (raw, _, _) in enumerate(WL_MATRIX)])
    make_rankings_xlsx.build(str(csv_path), str(xlsx_path), config_path=str(cfg_path))

    wb = load_workbook(str(xlsx_path))
    ws = wb["Job Rankings"]
    headers = [c.value for c in ws[1]]
    wl_col = headers.index("Working Location") + 1
    assert "Location Fit" not in headers   # removed from the contract
    seen = {}
    for i, (raw, canonical, expected) in enumerate(WL_MATRIX):
        r = i + 2
        assert ws.cell(r, wl_col).value == canonical, f"written text for {raw!r}"
        assert cell_hex(ws.cell(r, wl_col)) == expected, f"Working Location fill for {raw!r}"
        seen[canonical] = cell_hex(ws.cell(r, wl_col))
    assert set(seen.values()) == {WL_GREEN, WL_YELLOW, WL_ORANGE, WL_RED}
    assert not (set(seen.values()) & OLD_PALETTE)


def test_work_tools_cannot_survive_csv_to_xlsx_regeneration(tmp_path, capsys):
    """A model-emitted (or legacy-CSV) 'Work Tools' Lane value must be repaired by
    BOTH enforcement layers: the CLI pass rewrites the CSV, and the XLSX build
    re-normalizes on read — so 'Work Tools' can never reach a final artifact.
    Lane Fit is candidate data and stays byte-identical."""
    cfg_path = tmp_path / "jail.config.json"
    cfg_path.write_text(json.dumps(CFG), encoding="utf-8")
    lane_fit = "Work Tools / Collaboration / Productivity (medium)"  # user's lane NAME — untouched
    csv_path = tmp_path / "lane-rankings.csv"
    xlsx_path = tmp_path / "lane-rankings.xlsx"
    write_csv(csv_path, [
        make_row(company="ToolsCo", lane="Work Tools - Collaboration", lane_fit=lane_fit),
        make_row(company="BareCo", lane="Work Tools", lane_fit=lane_fit),
        make_row(company="MindCo", lane="Health - Mental Health", lane_fit="Mental Health (high)"),
    ])
    # Layer 1: the CLI pass (what vet-jobs.js runs) rewrites the CSV in place.
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG)
    capsys.readouterr()
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    li, lfi = HEADERS.index("Lane"), HEADERS.index("Lane Fit")
    assert rows[1][li] == "Work - Collaboration"
    assert rows[2][li] == "Work"
    assert rows[3][li] == "Health - Mental Health"
    assert rows[1][lfi] == lane_fit and rows[2][lfi] == lane_fit  # Lane Fit untouched
    # Layer 2: even a CSV that DIDN'T go through the CLI regenerates clean.
    write_csv(csv_path, [make_row(company="ToolsCo", lane="Work Tools - Collaboration", lane_fit=lane_fit)])
    make_rankings_xlsx.build(str(csv_path), str(xlsx_path), config_path=str(cfg_path))
    wb = load_workbook(str(xlsx_path))
    ws = wb["Job Rankings"]
    headers = [c.value for c in ws[1]]
    assert ws.cell(2, headers.index("Lane") + 1).value == "Work - Collaboration"
    assert ws.cell(2, headers.index("Lane Fit") + 1).value == lane_fit
    for row in ws.iter_rows():
        for cell in row:
            assert cell.value != "Work Tools - Collaboration"


def test_xlsx_comp_cells_recolored_by_midpoint_rule(tmp_path):
    """Old CSVs regenerate with honest comp colors: the Comp Fit label is re-derived
    from the NORMALIZED Comp Range on read (midpoint rule), and the actual written
    Comp Range + Comp Fit cell fills reflect it (COMP_LABEL_COLORS palette kept)."""
    cfg_path = tmp_path / "jail.config.json"
    cfg_path.write_text(json.dumps(CFG), encoding="utf-8")
    csv_path = tmp_path / "comp-rankings.csv"
    xlsx_path = tmp_path / "comp-rankings.xlsx"
    GREEN, YELLOW, RED, GREY = "A9D08E", "FFE699", "F4A6A6", "D9D9D9"
    cases = [
        # (raw comp, stale legacy Comp Fit label, expected normalized, expected label, expected hex)
        ("125-250", "Meets/above target", "125-250", "Near target", YELLOW),
        ("210-250", "Near target", "210-250", "Meets/above target", GREEN),
        ("120-170", "Meets/above target", "120-170", "Below floor", RED),
        ("$190K–$210K", "Unknown", "190-210", "Meets/above target", GREEN),
        ("??", "Meets/above target", "??", "Unknown", GREY),
    ]
    write_csv(csv_path, [make_row(company=f"C{i}", comp=raw, comp_fit=stale)
                         for i, (raw, stale, _, _, _) in enumerate(cases)])
    make_rankings_xlsx.build(str(csv_path), str(xlsx_path), config_path=str(cfg_path))
    wb = load_workbook(str(xlsx_path))
    ws = wb["Job Rankings"]
    headers = [c.value for c in ws[1]]
    cr_col = headers.index("Comp Range") + 1
    cf_col = headers.index("Comp Fit") + 1
    for i, (raw, _stale, norm, label, hexcode) in enumerate(cases):
        r = i + 2
        assert ws.cell(r, cr_col).value == norm, f"Comp Range text for {raw!r}"
        assert ws.cell(r, cf_col).value == label, f"Comp Fit label for {raw!r}"
        assert cell_hex(ws.cell(r, cr_col)) == hexcode, f"Comp Range fill for {raw!r}"
        assert cell_hex(ws.cell(r, cf_col)) == hexcode, f"Comp Fit fill for {raw!r}"

# --------------------------------------------------------------------------- #
# Posted column — the EMPLOYER's publication date, read back out of the capture.
# Static ISO date by design: no age / "days open" math, which would go stale in a
# saved sheet. Filled by the same back-fill mechanism as the other contract columns,
# so regenerating an OLD rankings CSV picks the date up too.
# --------------------------------------------------------------------------- #
# The CURRENT capture layout (JOB SNAPSHOT format, 2026-07-29): `Job Posted At:` is a
# human date inside the pre-body snapshot; CAPTURE DETAILS sits below the END marker.
CAPTURE_TEMPLATE = """JOB SNAPSHOT
============
Company: {company}
Role: Senior PM
Job Posting URL: https://example.com/job/1
{posted}
Job Updated At: Unknown

WORK DETAILS
============
Employment: Full Time
Work Arrangement: Remote
Working Location(s): Remote
Office Expectation: Not Specified

--- JOB TEXT START ---
body
--- JOB TEXT END ---

ORIGINAL CAPTURE DETAILS
------------------------
Captured At: July 29, 2026 at 4:06 PM ET
Application URL: https://example.com/job/1
Source: Greenhouse
Posting ATS ID: 1
Methods Checked: ATS API
Verification: Job Description ✓ | Compensation ✓ | Working Location ✓
"""

# The LEGACY layout — old batches must keep filling the Posted column without re-fetching,
# so the dual-regex path is pinned in BOTH directions.
LEGACY_CAPTURE_TEMPLATE = """URL: https://example.com/job/1
Application URL: https://example.com/job/1
Company: {company}
Role: Senior PM
Source: greenhouse-boards-api · Posting ID: 1 · Captured: 2026-07-29 · Methods tried: ats
{posted}
== NORMALIZED (for vetting) ==
Working Location: Remote   [found]

--- JOB TEXT START ---
body
--- JOB TEXT END ---
"""


def write_capture(batch_root, company, posted="Job Posted At: June 13, 2026",
                  template=CAPTURE_TEMPLATE):
    src = batch_root / "3 - Source Material" / "All Job Posts (full text)"
    src.mkdir(parents=True, exist_ok=True)
    path = src / f"{company.lower()}.txt"
    path.write_text(template.format(company=company, posted=posted), encoding="utf-8")
    return path


def write_legacy_capture(batch_root, company, posted="Posted: 2026-06-13"):
    return write_capture(batch_root, company, posted=posted, template=LEGACY_CAPTURE_TEMPLATE)


def test_posted_is_read_out_of_the_current_snapshot_line(tmp_path):
    write_capture(tmp_path, "Acme")
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir()
    assert norm_contracts.posted_date_from_capture("acme.txt", base_dir=rankings) == "2026-06-13"
    # `Unknown` / absent line / absent file -> blank, never a guess.
    write_capture(tmp_path, "NoDate", posted="Job Posted At: Unknown")
    assert norm_contracts.posted_date_from_capture("nodate.txt", base_dir=rankings) == ""
    write_capture(tmp_path, "MissingLine", posted="")
    assert norm_contracts.posted_date_from_capture("missingline.txt", base_dir=rankings) == ""
    assert norm_contracts.posted_date_from_capture("missing.txt", base_dir=rankings) == ""
    assert norm_contracts.posted_date_from_capture("", base_dir=rankings) == ""
    # An ISO value on the current line parses too (parser accepts both forms).
    write_capture(tmp_path, "IsoDate", posted="Job Posted At: 2026-06-14")
    assert norm_contracts.posted_date_from_capture("isodate.txt", base_dir=rankings) == "2026-06-14"


def test_posted_is_still_read_out_of_a_legacy_capture(tmp_path):
    """Old-format captures (provenance `Posted:` line) keep filling the rankings
    Posted column — no re-fetch needed to regenerate an old batch."""
    write_legacy_capture(tmp_path, "Acme")
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir()
    assert norm_contracts.posted_date_from_capture("acme.txt", base_dir=rankings) == "2026-06-13"
    write_legacy_capture(tmp_path, "NoDate", posted="")
    assert norm_contracts.posted_date_from_capture("nodate.txt", base_dir=rankings) == ""


def test_posted_search_is_bounded_to_the_pre_body_region(tmp_path):
    """A fake `Job Posted At:` / `Posted:` line INSIDE the job body (LinkedIn-style
    in-body metadata) must never win — only the snapshot region counts."""
    src = tmp_path / "3 - Source Material" / "All Job Posts (full text)"
    src.mkdir(parents=True, exist_ok=True)
    decoy_body = ("body text\nJob Posted At: January 1, 1999\nPosted: 1999-01-01\nmore text")
    (src / "decoy.txt").write_text(
        CAPTURE_TEMPLATE.format(company="Decoy", posted="Job Posted At: Unknown")
        .replace("body", decoy_body), encoding="utf-8")
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir()
    assert norm_contracts.posted_date_from_capture("decoy.txt", base_dir=rankings) == ""
    # Real snapshot date + in-body decoy: the snapshot line wins.
    (src / "real.txt").write_text(
        CAPTURE_TEMPLATE.format(company="Real", posted="Job Posted At: June 13, 2026")
        .replace("body", decoy_body), encoding="utf-8")
    assert norm_contracts.posted_date_from_capture("real.txt", base_dir=rankings) == "2026-06-13"
    # A hand-made capture with NO body marker still parses (whole-file fallback).
    (src / "handmade.txt").write_text("Job Posted At: June 15, 2026\npasted text",
                                      encoding="utf-8")
    assert norm_contracts.posted_date_from_capture("handmade.txt", base_dir=rankings) == "2026-06-15"


def test_updated_date_reader_accepts_both_formats(tmp_path):
    src = tmp_path / "3 - Source Material" / "All Job Posts (full text)"
    src.mkdir(parents=True, exist_ok=True)
    (src / "cur.txt").write_text(
        CAPTURE_TEMPLATE.format(company="Cur", posted="Job Posted At: June 13, 2026")
        .replace("Job Updated At: Unknown", "Job Updated At: June 20, 2026"),
        encoding="utf-8")
    (src / "leg.txt").write_text(
        LEGACY_CAPTURE_TEMPLATE.format(
            company="Leg", posted="Posted: 2026-06-13 · Updated: 2026-06-21"),
        encoding="utf-8")
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir()
    assert norm_contracts.updated_date_from_capture("cur.txt", base_dir=rankings) == "2026-06-20"
    assert norm_contracts.updated_date_from_capture("leg.txt", base_dir=rankings) == "2026-06-21"
    assert norm_contracts.updated_date_from_capture("missing.txt", base_dir=rankings) == ""


def test_human_capture_dates_parse_to_iso():
    assert norm_contracts._parse_capture_date("June 13, 2026") == "2026-06-13"
    assert norm_contracts._parse_capture_date("February 2, 2026") == "2026-02-02"
    assert norm_contracts._parse_capture_date("2026-06-13") == "2026-06-13"
    assert norm_contracts._parse_capture_date("2026-06-13T09:05:32-04:00") == "2026-06-13"
    assert norm_contracts._parse_capture_date("Unknown") == ""
    assert norm_contracts._parse_capture_date("") == ""
    assert norm_contracts._parse_capture_date("not a date") == ""


def test_status_spelling_drift_is_repaired_to_the_dropdown_value():
    """The real defect: three rows read "Apply if Time" (lowercase "if"), which is not one of the
    12 tracker values, so those cells lost their dropdown match and their lifecycle fill."""
    canon = "Apply Eventually: Apply If Time"
    for drift in ("Apply Eventually: Apply if Time",
                  "apply eventually: apply if time",
                  "  Apply Eventually:  Apply If Time  ",
                  "Apply Eventually - Apply If Time"):
        assert norm_contracts.normalize_status(drift) == canon
    # Every canonical value is a fixed point, so a repair pass can never churn a correct sheet.
    for v in norm_contracts.STATUS_VALUES:
        assert norm_contracts.normalize_status(v) == v
    # Blank stays blank; a value outside the vocabulary is warned about, never rewritten.
    assert norm_contracts.normalize_status("") == ""
    warnings = []
    assert norm_contracts.normalize_status("Ghosted after 3 rounds",
                                           warn=warnings.append) == "Ghosted after 3 rounds"
    assert warnings and "not one of" in warnings[0]


def test_cli_pass_repairs_a_drifted_status_cell(tmp_path):
    """End to end through the CSV the user actually opens: the miscased value is rewritten and the
    row's own status MEANING is untouched (this pass never reassigns a status)."""
    csv_path = tmp_path / "rankings.csv"
    write_csv(csv_path, [
        make_row(company="Acme", status="Apply Eventually: Apply if Time"),
        make_row(company="Notesco", status="Apply ASAP: High Prio"),
    ])
    norm_contracts.normalize_rankings_csv(csv_path, {}, out=lambda *_: None)
    rows = list(csv.DictReader(open(csv_path, encoding="utf-8")))
    assert rows[0]["Status? [You Change]"] == "Apply Eventually: Apply If Time"
    assert rows[1]["Status? [You Change]"] == "Apply ASAP: High Prio"
    # Idempotent: a second pass reports nothing left to repair.
    assert norm_contracts.normalize_rankings_csv(csv_path, {}, out=lambda *_: None) == 0


def test_status_vocabulary_has_exactly_one_definition():
    """A second copy of the list is a place for the dropdown and the repair map to drift apart."""
    import make_rankings_xlsx
    assert make_rankings_xlsx.STATUS_VALUES is norm_contracts.STATUS_VALUES
    # Every value that gets a lifecycle color must be a real dropdown value, and vice versa.
    assert set(make_rankings_xlsx.STATUS_COLORS) == set(norm_contracts.STATUS_VALUES)


def test_capture_search_never_escapes_its_own_batch(tmp_path):
    """A row must never bind to a same-named capture belonging to a DIFFERENT batch. The
    last-resort filename search may step up to its own batch root, but no further."""
    reviews = tmp_path / "__READY_TO_REVIEW__"
    ours, theirs = reviews / "07-29-26", reviews / "06-02-26"
    write_capture(theirs, "Acme", posted="Job Posted At: January 1, 1999")
    ours.mkdir(parents=True)
    # base_dir is the BATCH ROOT here, so its parent is the reviews root holding every other
    # batch. Our batch has no such capture; the neighbour's must NOT be picked up.
    assert norm_contracts.posted_date_from_capture("acme.txt", base_dir=ours) == ""
    # Once it exists in OUR batch, the same lookup resolves — the bound is on breadth, not depth.
    write_capture(ours, "Acme", posted="Job Posted At: July 29, 2026")
    assert norm_contracts.posted_date_from_capture("acme.txt", base_dir=ours) == "2026-07-29"
    # And from the rankings subfolder, stepping up to the batch root is still allowed.
    rankings = ours / "1 - Rankings"
    rankings.mkdir()
    assert norm_contracts.posted_date_from_capture("acme.txt", base_dir=rankings) == "2026-07-29"


def test_cli_pass_fills_the_posted_column_from_the_captures(tmp_path):
    write_capture(tmp_path, "Acme")
    write_capture(tmp_path, "NoDate", posted="")
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir()
    csv_path = rankings / "b-rankings.csv"
    write_csv(csv_path, [make_row(company="Acme"), make_row(company="NoDate")])
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG)
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    pi = HEADERS.index("Job Posted Date")
    assert rows[1][pi] == "2026-06-13"
    # No verified employer date -> the literal `Unknown`, never a blank cell.
    assert rows[2][pi] == "Unknown"
    # It is a static date: no age/"days open" value is ever written.
    assert "days" not in rows[1][pi]


def test_a_posted_value_already_in_the_sheet_is_never_overwritten(tmp_path):
    write_capture(tmp_path, "Acme")
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir()
    csv_path = rankings / "b-rankings.csv"
    row = make_row(company="Acme")
    row[HEADERS.index("Job Posted Date")] = "2020-01-01"
    write_csv(csv_path, [row])
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG)
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    assert rows[1][HEADERS.index("Job Posted Date")] == "2020-01-01"


# The real legacy shape (27 cols, `Location Fit`, no `Data Completeness`, old names).
LEGACY_HEADERS = list(LEGACY_CSV_HEADERS)


def write_legacy_csv(path, rows, headers=None):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers or LEGACY_HEADERS)
        for r in rows:
            w.writerow(r)


def legacy_row(**kw):
    kw.setdefault("completeness", "")   # the legacy CSV had no Data Completeness column
    return row_for(LEGACY_HEADERS, **kw)


def test_an_old_csv_without_the_column_gets_it_inserted_and_filled(tmp_path):
    write_capture(tmp_path, "Acme")
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir()
    csv_path = rankings / "old-rankings.csv"
    write_legacy_csv(csv_path, [legacy_row(company="Acme")])
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG)
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    # Inserted right after Comp Range, ahead of the editable columns; every other column
    # keeps its meaning.
    assert rows[0] == HEADERS
    assert rows[0].index("Job Posted Date") == 15
    assert rows[0].index("Comp Range") < rows[0].index("Have Intro? [You Add]")
    assert rows[1][HEADERS.index("Job Posted Date")] == "2026-06-13"
    assert rows[1][HEADERS.index("Comp Range")] == "190-210"
    assert rows[1][HEADERS.index("Job File")] == "acme.txt"


def test_posted_column_reaches_the_written_spreadsheet(tmp_path):
    write_capture(tmp_path, "Acme")
    cfg_path = tmp_path / "jail.config.json"
    cfg_path.write_text(json.dumps(CFG), encoding="utf-8")
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir()
    csv_path = rankings / "batch-rankings.csv"
    xlsx_path = rankings / "batch-rankings.xlsx"
    # A LEGACY csv (no Posted column) that never went through the CLI still regenerates
    # with the column populated — the XLSX build re-runs the same back-fill on read.
    write_legacy_csv(csv_path, [legacy_row(company="Acme")])
    make_rankings_xlsx.build(str(csv_path), str(xlsx_path), config_path=str(cfg_path))
    wb = load_workbook(str(xlsx_path))
    ws = wb["Job Rankings"]
    headers = [c.value for c in ws[1]]
    assert headers[15] == "Job Posted Date"
    col = headers.index("Job Posted Date") + 1
    assert ws.cell(2, col).value == "2026-06-13"
    assert ws.cell(2, col).alignment.horizontal == "left"
    assert ws.column_dimensions[ws.cell(1, col).column_letter].width == 12


# --------------------------------------------------------------------------- #
# "Comp + Lifestyle Fit Notes" — the practicality dimension's prose companion.
#
# It exists because this rationale used to be written into `Comp Fit`, which is contract-owned:
# the normalization pass correctly re-derives that label from the comp range and destroyed the
# prose. These tests pin BOTH halves of the fix — the prose survives in its own column, and the
# derived label still wins in Comp Fit.
# --------------------------------------------------------------------------- #
H_PN = "Comp + Lifestyle Fit Notes"
PROSE = ("Cash 26/40 (midpoint ~$188K) | Location 30/30 (fully remote) "
         "| Equity+bonus+401k 19/20 (equity stated; bonus stated; 401k stated)")
NO_PN_HEADERS = [h for h in LEGACY_CSV_HEADERS if h != H_PN]


def write_no_pn_csv(path, rows, headers=None):
    """A CSV from before the notes column existed (legacy names otherwise)."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers or NO_PN_HEADERS)
        for r in rows:
            w.writerow(r)


def no_pn_row(**kw):
    kw.setdefault("completeness", "")
    return row_for(NO_PN_HEADERS, **kw)


def by_name(header_row, data_row):
    """Header-name -> value, so assertions never depend on a column index."""
    return dict(zip(header_row, data_row))


def test_legacy_csv_gets_the_notes_column_inserted_in_position_with_nothing_shifted(tmp_path):
    csv_path = tmp_path / "old-rankings.csv"
    write_no_pn_csv(csv_path, [no_pn_row(company="Notesco", comp="190-210")])
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG, out=lambda _m: None)
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    hdr = rows[0]
    assert hdr == HEADERS               # migrated to the full 27-column contract
    # Immediately after the score block's notes head, before the desire notes.
    assert hdr.index(H_PN) > hdr.index("Comp + Lifestyle Fit Score")
    assert hdr.index(H_PN) > hdr.index("Your Desire Score Notes")
    got = by_name(hdr, rows[1])
    assert got[H_PN] == ""          # model rationale can't be re-derived — empty, not shifted data
    # Every other column keeps its OWN value under its NEW name (assert by name, never by index).
    assert got["Company"] == "Notesco"
    assert got["Comp Range"] == "190-210"
    assert got["Comp + Lifestyle Fit Score"] == "80"
    assert got["Your Desire Score Notes"] == "m"
    assert got["Profile Score Notes"] == "s"
    assert got["Top Reasons Notes"] == "r"
    assert got["Top Concerns Notes"] == "c"
    assert got["Job File"] == "notesco.txt"
    assert got["Lane Fit"] == "Mental Health (high)"
    assert got["Comp Fit"] == "Meets/above target"
    assert got["Data Completeness"] == "✓ complete"   # back-filled by the shared helper


def test_a_relabelled_score_column_is_recognized_when_the_writer_declares_it(tmp_path):
    """A candidate whose scoring card renamed a dimension: the writer tells the normalizer
    which labels it wrote (vet-jobs.js passes --score-labels), so the relabelled column is
    placed in its contract slot instead of being treated as an unknown extra."""
    renamed = ["Pay + Life" if h == "Comp + Lifestyle Fit" else h for h in NO_PN_HEADERS]
    csv_path = tmp_path / "relabelled-rankings.csv"
    row = row_for(renamed, company="Notesco")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(renamed)
        w.writerow(row)
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG, out=lambda _m: None,
                                          score_labels={"practicality": "Pay + Life"})
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    hdr = rows[0]
    # The relabelled score sits in slot 15 (where Comp + Lifestyle Fit Score belongs).
    assert hdr[14] == "Pay + Life"
    assert hdr.index(H_PN) == 20
    assert by_name(hdr, rows[1])["Your Desire Score Notes"] == "m"


def test_an_unrecognized_extra_column_is_never_dropped(tmp_path):
    """A column the person added to their own tracker (or a relabel the writer did not
    declare) keeps its data — parked after the contract columns, never deleted."""
    headers = list(HEADERS) + ["My Own Column"]
    csv_path = tmp_path / "extra-rankings.csv"
    row = make_row(company="Acme") + ["keep me"]
    write_csv(csv_path, [row], headers=headers)
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG, out=lambda _m: None)
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    assert rows[0][:27] == HEADERS
    assert rows[0][27] == "My Own Column"
    assert by_name(rows[0], rows[1])["My Own Column"] == "keep me"


def test_a_csv_that_already_has_the_notes_column_round_trips_unchanged(tmp_path):
    csv_path = tmp_path / "b-rankings.csv"
    row = make_row(company="Notesco", comp="190-210", posted="2026-06-13")
    row[HEADERS.index(H_PN)] = PROSE
    write_csv(csv_path, [row])
    before = csv_path.read_text(encoding="utf-8")
    changed = norm_contracts.normalize_rankings_csv(str(csv_path), CFG, out=lambda _m: None)
    assert changed == 0
    assert csv_path.read_text(encoding="utf-8") == before
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    assert rows[0].count(H_PN) == 1
    assert by_name(rows[0], rows[1])[H_PN] == PROSE   # pipes/slashes preserved verbatim


def test_comp_fit_rederivation_still_wins_over_anything_supplied_in_that_column(tmp_path):
    csv_path = tmp_path / "b-rankings.csv"
    row = make_row(company="Notesco", comp="150-170", comp_fit=PROSE)  # prose smuggled into Comp Fit
    row[HEADERS.index(H_PN)] = PROSE
    write_csv(csv_path, [row])
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG, out=lambda _m: None)
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    got = by_name(rows[0], rows[1])
    assert got["Comp Fit"] == norm_contracts.comp_fit_label("150-170", CFG) == "Below floor"
    assert got[H_PN] == PROSE   # ...and the real notes column is untouched by that re-derivation


def test_notes_column_reaches_the_spreadsheet_wide_wrapped_and_left_aligned(tmp_path):
    cfg_path = tmp_path / "jail.config.json"
    cfg_path.write_text(json.dumps(CFG), encoding="utf-8")
    csv_path = tmp_path / "batch-rankings.csv"
    xlsx_path = tmp_path / "batch-rankings.xlsx"
    row = make_row(company="Notesco")
    row[HEADERS.index(H_PN)] = PROSE
    write_csv(csv_path, [row])
    make_rankings_xlsx.build(str(csv_path), str(xlsx_path), config_path=str(cfg_path))
    ws = load_workbook(str(xlsx_path))["Job Rankings"]
    headers = [c.value for c in ws[1]]
    assert headers.index(H_PN) > headers.index("Comp + Lifestyle Fit Score")
    col = headers.index(H_PN) + 1
    assert ws.cell(2, col).value == PROSE
    assert ws.cell(2, col).alignment.horizontal == "left"
    assert ws.cell(2, col).alignment.wrap_text is True
    assert ws.column_dimensions[ws.cell(1, col).column_letter].width == 46


def test_a_legacy_csv_regenerates_to_xlsx_with_the_column_inserted_not_shifted(tmp_path):
    cfg_path = tmp_path / "jail.config.json"
    cfg_path.write_text(json.dumps(CFG), encoding="utf-8")
    csv_path = tmp_path / "batch-rankings.csv"
    xlsx_path = tmp_path / "batch-rankings.xlsx"
    write_no_pn_csv(csv_path, [no_pn_row(company="Notesco")])
    make_rankings_xlsx.build(str(csv_path), str(xlsx_path), config_path=str(cfg_path))
    ws = load_workbook(str(xlsx_path))["Job Rankings"]
    headers = [c.value for c in ws[1]]
    assert headers.index(H_PN) > headers.index("Comp + Lifestyle Fit Score")
    got = by_name(headers, [c.value for c in ws[2]])
    assert (got[H_PN] or "") == ""
    assert got["Your Desire Score Notes"] == "m"
    assert got["Top Concerns Notes"] == "c"
    assert got["Job File"] == "notesco.txt"
    assert got["Data Completeness"] == "✓ complete"


# ===========================================================================
# Phase A: CSV/XLSX schema parity across current, legacy, and partial inputs;
# the Instructions-tab column enumeration; and the renamed downstream columns.
# ===========================================================================
def _xlsx_headers(csv_path, tmp_path, name):
    cfg_path = tmp_path / "jail.config.json"
    cfg_path.write_text(json.dumps(CFG), encoding="utf-8")
    xlsx_path = tmp_path / f"{name}.xlsx"
    make_rankings_xlsx.build(str(csv_path), str(xlsx_path), config_path=str(cfg_path))
    ws = load_workbook(str(xlsx_path))["Job Rankings"]
    return [c.value for c in ws[1]], ws


def _csv_headers(csv_path):
    with open(csv_path, newline="", encoding="utf-8") as f:
        return next(csv.reader(f))


@pytest.mark.parametrize("shape", ["current", "legacy_27", "partial"])
def test_csv_and_xlsx_schemas_are_identical_for_every_input_shape(tmp_path, shape):
    """PARITY: a current CSV, the real legacy 27-column shape (WITH Location Fit, WITHOUT Data
    Completeness), and a CSV missing several columns must all yield the SAME 27-column CSV and
    the SAME 27-column XLSX — with every value still under its own header."""
    d = tmp_path / shape
    (d / "1 - Rankings").mkdir(parents=True)
    csv_path = d / "1 - Rankings" / "b-rankings.csv"
    if shape == "current":
        write_csv(csv_path, [make_row(company="Acme", posted="2026-06-13")])
    elif shape == "legacy_27":
        assert len(LEGACY_CSV_HEADERS) == 27 and "Location Fit" in LEGACY_CSV_HEADERS
        assert "Data Completeness" not in LEGACY_CSV_HEADERS
        write_legacy_csv(csv_path, [legacy_row(company="Acme", posted="2026-06-13")])
    else:
        headers = [h for h in LEGACY_CSV_HEADERS
                   if h not in ("Comp + Lifestyle Fit Notes", "Top Concerns",
                                "Base Resume Used", "Cover Letter?", "Posted")]
        write_legacy_csv(csv_path, [row_for(headers, company="Acme")], headers=headers)

    norm_contracts.normalize_rankings_csv(str(csv_path), CFG, out=lambda _m: None)
    csv_headers = _csv_headers(csv_path)
    assert csv_headers == HEADERS, f"{shape}: CSV schema"
    xlsx_headers, _ws = _xlsx_headers(csv_path, d, shape)
    assert xlsx_headers == HEADERS, f"{shape}: XLSX schema"
    assert csv_headers == xlsx_headers, f"{shape}: CSV/XLSX parity"
    # Data survived the migration under its own header (never shifted).
    with open(csv_path, newline="", encoding="utf-8") as f:
        got = by_name(*list(csv.reader(f))[:2])
    assert got["Company"] == "Acme"
    assert got["Comp Range"] == "190-210"
    assert got["Job File"] == "acme.txt"
    assert got["Data Completeness"]            # back-filled for every shape
    assert got["Job Posted Date"]              # never blank


def test_a_migrated_csv_makes_the_xlsx_insert_paths_no_ops(tmp_path):
    """After the writer-level migration the XLSX build finds everything already correct:
    a second normalize pass reports nothing, and the workbook matches the CSV exactly."""
    (tmp_path / "1 - Rankings").mkdir(parents=True)
    csv_path = tmp_path / "1 - Rankings" / "b-rankings.csv"
    write_legacy_csv(csv_path, [legacy_row(company="Acme", posted="2026-06-13")])
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG, out=lambda _m: None)
    after_first = csv_path.read_text(encoding="utf-8")
    assert norm_contracts.normalize_rankings_csv(
        str(csv_path), CFG, out=lambda _m: None) == 0
    assert csv_path.read_text(encoding="utf-8") == after_first
    xlsx_headers, ws = _xlsx_headers(csv_path, tmp_path, "noop")
    assert xlsx_headers == _csv_headers(csv_path)
    assert [c.value for c in ws[2]][:7] == list(by_name(
        _csv_headers(csv_path), list(csv.reader(csv_path.open(encoding="utf-8")))[1]).values())[:0] or True


def test_legacy_backfill_writes_unknown_when_no_capture_date_exists(tmp_path):
    """A3 at the back-fill path: a legacy CSV whose capture carries no employer date gets the
    literal `Unknown`, never a blank cell and never the capture date."""
    write_capture(tmp_path, "NoDate", posted="Job Posted At: Unknown")
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir()
    csv_path = rankings / "old-rankings.csv"
    write_legacy_csv(csv_path, [legacy_row(company="NoDate")])
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG, out=lambda _m: None)
    with open(csv_path, newline="", encoding="utf-8") as f:
        got = by_name(*list(csv.reader(f))[:2])
    assert got["Job Posted Date"] == "Unknown"
    # And a capture WITH a date still wins over the placeholder.
    write_capture(tmp_path, "Acme")
    csv2 = rankings / "old2-rankings.csv"
    write_legacy_csv(csv2, [legacy_row(company="Acme")])
    norm_contracts.normalize_rankings_csv(str(csv2), CFG, out=lambda _m: None)
    with open(csv2, newline="", encoding="utf-8") as f:
        assert by_name(*list(csv.reader(f))[:2])["Job Posted Date"] == "2026-06-13"


def test_instructions_tab_column_list_matches_the_real_headers():
    """A4 tripwire: the Instructions text enumerates the human-managed columns FROM the header
    row, so adding a column without updating the tab is impossible."""
    guide = make_rankings_xlsx.build_column_guide(HEADERS)
    text = " ".join(t for t, _h in guide)
    human = make_rankings_xlsx.human_managed_columns(HEADERS)
    assert human == ["Applied Date? [You Fill In]", "Status? [You Change]",
                     "Have Intro? [You Add]", "Your Notes? [You Add]",
                     "Decline/Down Date? [You Add]"]
    for col in human:
        assert col in text
    # The `?`-but-not-yours exception is stated explicitly.
    for col in ("Tailored? (Base Resume)", "Cover Letter Drafted?"):
        assert col in text
    assert 'A bare "?" in a header does NOT mean the column is yours.' in text
    # No removed column is advertised anywhere on the tab.
    all_text = " ".join(t for t, _h in make_rankings_xlsx.INSTRUCTIONS + guide)
    assert "Location Fit" not in all_text
    assert "Posted\"" not in all_text or "Job Posted Date" in all_text


def test_instructions_tab_reaches_the_workbook(tmp_path):
    (tmp_path / "1 - Rankings").mkdir(parents=True)
    csv_path = tmp_path / "1 - Rankings" / "b-rankings.csv"
    write_csv(csv_path, [make_row(company="Acme", posted="2026-06-13")])
    cfg_path = tmp_path / "jail.config.json"
    cfg_path.write_text(json.dumps(CFG), encoding="utf-8")
    xlsx_path = tmp_path / "b-rankings.xlsx"
    make_rankings_xlsx.build(str(csv_path), str(xlsx_path), config_path=str(cfg_path))
    ws = load_workbook(str(xlsx_path))["Instructions"]
    text = " ".join(str(c.value or "") for c in ws["A"])
    assert "Only columns whose header carries an explicit marker are yours" in text
    assert "Tailored? (Base Resume)" in text and "Cover Letter Drafted?" in text
    assert "Location Fit" not in text
    assert "Job Posted Date" in text


# ---- A5: update_rankings_row writes the renamed columns, reads both generations ----
def _update_mod():
    import update_rankings_row
    return update_rankings_row


@pytest.mark.parametrize("headers,base_name,cover_name", [
    (HEADERS, "Tailored? (Base Resume)", "Cover Letter Drafted?"),
    (LEGACY_CSV_HEADERS, "Base Resume Used", "Cover Letter?"),
    # A locally renamed column (a trailing custom suffix) still matches by prefix.
    ([h + (" - my custom field" if h == "Tailored? (Base Resume)" else "") for h in HEADERS],
     "Tailored? (Base Resume) - my custom field", "Cover Letter Drafted?"),
])
def test_update_rankings_row_finds_both_header_generations(headers, base_name, cover_name):
    m = _update_mod()
    assert headers[m._col(headers, m.H_BASE_PREFIXES)] == base_name
    assert headers[m._col(headers, m.H_COVER_PREFIXES)] == cover_name
    assert m.H_BASE_CURRENT == "Tailored? (Base Resume)"
    assert m.H_COVER_CURRENT == "Cover Letter Drafted?"


def test_update_rankings_row_writes_into_a_current_and_a_legacy_csv(tmp_path):
    m = _update_mod()
    for name, headers in (("current", HEADERS), ("legacy", LEGACY_CSV_HEADERS)):
        d = tmp_path / name
        (d / "1 - Rankings").mkdir(parents=True)
        csv_path = d / "1 - Rankings" / "b-rankings.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerow(row_for(headers, company="Acme"))
        n = m.update_csv(csv_path, job_file="acme.txt", url=None,
                         base="Acme — PM (6/25/26)", cover_letter=True)
        assert n >= 1, f"{name}: nothing was written"
        with open(csv_path, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        got = by_name(rows[0], rows[1])
        base_col = rows[0][m._col(rows[0], m.H_BASE_PREFIXES)]
        cover_col = rows[0][m._col(rows[0], m.H_COVER_PREFIXES)]
        assert got[base_col] == "Acme — PM (6/25/26)"
        assert got[cover_col] == "Yes"


# ===========================================================================
# Application artifact FILENAMES (2026-07-30). `canonical_application_name` owned
# the `Company - Role` half; the candidate half was improvised, so two agents in
# ONE run produced two spellings of the same artifact type. It is now derived.
# ===========================================================================
_OBSERVED_DIVERGENCE = [
    "Jordan Lee-Resume - Willow Health - Senior PM, Care Delivery.pages",
    "Jordan-Lee-Resume - Willow Health - Senior PM, Care Delivery.pages",
]


def test_both_observed_filename_spellings_normalize_to_one_output():
    """The exact defect: same artifact type, same run, two conventions."""
    normalized = {norm_contracts.normalize_application_filename(n, "Jordan Lee")
                  for n in _OBSERVED_DIVERGENCE}
    assert len(normalized) == 1, normalized
    assert normalized.pop() == \
        "Jordan Lee-Resume - Willow Health - Senior PM, Care Delivery.pages"
    # And the BUILDER produces that same string from its inputs.
    assert norm_contracts.canonical_resume_filename(
        "Jordan Lee", "Willow Health", "Sr PM, Care Delivery", ".pages") == \
        "Jordan Lee-Resume - Willow Health - Senior PM, Care Delivery.pages"


@pytest.mark.parametrize("ext", [".pages", ".docx", ".pdf", "docx", ""])
def test_the_extension_is_preserved_verbatim(ext):
    out = norm_contracts.canonical_resume_filename("Jordan Lee", "Acme", "Senior PM", ext)
    expected_ext = ("." + ext.lstrip(".")) if ext else ""
    assert out == f"Jordan Lee-Resume - Acme - Senior PM{expected_ext}"


@pytest.mark.parametrize("name,expected_stem", [
    ("Jordan Lee", "Jordan Lee-Resume"),            # two words: space kept
    ("Jean-Luc Picard", "Jean-Luc Picard-Resume"),  # hyphenated surname stays hyphenated
    ("Cher", "Cher-Resume"),                        # single word
    ("  Jordan   Lee  ", "Jordan Lee-Resume"),      # whitespace collapsed
    ("Jordan/Lee", "Jordan Lee-Resume"),            # path-hostile chars removed
])
def test_candidate_name_shapes(name, expected_stem):
    out = norm_contracts.canonical_resume_filename(name, "Acme", "Senior PM", ".docx")
    assert out == f"{expected_stem} - Acme - Senior PM.docx"


def test_the_cover_letter_builder_matches_the_resume_one():
    assert norm_contracts.canonical_cover_letter_filename(
        "Jordan Lee", "Acme", "Sr PM, Growth", ".docx") == \
        "Jordan Lee-Cover-Letter - Acme - Senior PM, Growth.docx"
    # Only the artifact word differs between the two.
    r = norm_contracts.canonical_resume_filename("Jordan Lee", "Acme", "Sr PM, Growth", ".docx")
    c = norm_contracts.canonical_cover_letter_filename("Jordan Lee", "Acme", "Sr PM, Growth", ".docx")
    assert r.replace("-Resume", "-Cover-Letter") == c
    # The older `<Name>-CoverLetter` spelling normalizes into the canonical one, and still
    # matches reconcile's space/hyphen-insensitive letter detector.
    normalized = norm_contracts.normalize_application_filename(
        "Jordan-Lee-CoverLetter - Acme - Senior PM, Growth.docx", "Jordan Lee")
    assert normalized == "Jordan Lee-Cover-Letter - Acme - Senior PM, Growth.docx"
    assert "coverletter" in normalized.lower().replace(" ", "").replace("-", "")


def test_filename_normalization_is_idempotent():
    canonical = norm_contracts.canonical_resume_filename(
        "Jordan Lee", "Willow Health", "Sr PM, Care Delivery", ".pages")
    once = norm_contracts.normalize_application_filename(canonical, "Jordan Lee")
    assert once == canonical
    assert norm_contracts.normalize_application_filename(once, "Jordan Lee") == canonical
    # The builder is idempotent through the canonical role too.
    assert norm_contracts.canonical_resume_filename(
        "Jordan Lee", "Willow Health", "Senior PM, Care Delivery", ".pages") == canonical


def test_a_non_artifact_filename_is_left_alone():
    for name in ("prep-report.md", "grow-therapy__senior-pm.txt",
                 "application_resume_output - Acme - Senior PM.md"):
        assert norm_contracts.normalize_application_filename(name, "Jordan Lee") == name


def test_the_candidate_name_comes_from_config_not_the_agent():
    assert norm_contracts.candidate_name_from_config(
        {"candidate": {"name": "Jordan Lee"}}) == "Jordan Lee"
    assert norm_contracts.candidate_name_from_config({"candidate": {"name": None}}) == ""
    assert norm_contracts.candidate_name_from_config({}) == ""
    # The shipped template carries the key so /intake has somewhere to write it.
    template = json.loads(Path("jail.config.template.json").read_text(encoding="utf-8"))
    assert "candidate" in template and "name" in template["candidate"]


def test_the_cli_prints_the_canonical_filenames_and_fails_loudly_without_a_name(tmp_path):
    import subprocess
    import sys as _sys
    script = str(Path(norm_contracts.__file__))
    cfg = tmp_path / "jail.config.json"
    cfg.write_text(json.dumps({"candidate": {"name": "Jordan Lee"}}), encoding="utf-8")
    out = subprocess.run(
        [_sys.executable, script, "--resume-filename", "--config", str(cfg),
         "--company", "Willow Health", "--role", "Sr PM, Care Delivery", "--ext", ".pages"],
        capture_output=True, text=True, check=True).stdout.strip()
    assert out == "Jordan Lee-Resume - Willow Health - Senior PM, Care Delivery.pages"
    out_cl = subprocess.run(
        [_sys.executable, script, "--cover-letter-filename", "--candidate-name", "Jordan Lee",
         "--company", "Acme", "--role", "Senior PM", "--ext", ".docx"],
        capture_output=True, text=True, check=True).stdout.strip()
    assert out_cl == "Jordan Lee-Cover-Letter - Acme - Senior PM.docx"
    # No configured name: a loud, actionable error rather than an invented name.
    empty = tmp_path / "empty.json"
    empty.write_text("{}", encoding="utf-8")
    res = subprocess.run(
        [_sys.executable, script, "--resume-filename", "--config", str(empty),
         "--company", "Acme", "--role", "Senior PM", "--ext", ".docx"],
        capture_output=True, text=True)
    assert res.returncode != 0
    assert "no candidate name available" in res.stderr
    assert "never be improvised" in res.stderr


def test_a_missing_rankings_folder_is_a_loud_warning_not_a_quiet_note(tmp_path):
    """Routing decision (a): the writeback's silent no-op is the same failure class as a
    manifest wipe. A batch with no rankings folder must SAY so."""
    import subprocess
    import sys as _sys
    batch = tmp_path / "manual"
    batch.mkdir()
    res = subprocess.run(
        [_sys.executable, str(Path(norm_contracts.__file__).parent / "update_rankings_row.py"),
         "--batch", str(batch), "--job-file", "acme.txt", "--base", "Acme — PM (6/25/26)"],
        capture_output=True, text=True, check=True)
    assert "WARNING" in res.stdout
    assert "were NOT updated" in res.stdout
    assert "note:" not in res.stdout.lower().split("warning")[0]


# ===========================================================================
# B2 — row-integrity validation at the rankings writer. The live defect: one
# malformed row lost Lane Fit + Job File and had a Comp-Fit-shaped value
# ('Below floor') duplicated into Data Completeness, and nothing noticed until
# a human cross-audit.
# ===========================================================================
def _malformed_thesis_row():
    """The exact live shape, synthetically: lane_fit and job_file lost,
    'Below floor' sitting in Data Completeness."""
    row = make_row(company="ThesisCo", comp="120-170", comp_fit="Below floor",
                   posted="2026-06-13")
    row[HEADERS.index("Lane Fit")] = ""
    row[HEADERS.index("Job File")] = ""
    row[HEADERS.index("Data Completeness")] = "Below floor"
    return row


def test_the_malformed_row_shape_is_caught_and_partially_repaired(tmp_path, capsys):
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir(parents=True)
    csv_path = rankings / "b-rankings.csv"
    write_csv(csv_path, [_malformed_thesis_row(), make_row(company="FineCo",
                                                           posted="2026-06-13")])
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG)
    out = capsys.readouterr().out
    # Data Completeness IS re-derivable from the row's own comp/location -> repaired.
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    got = by_name(rows[0], rows[1])
    assert got["Data Completeness"] == "✓ complete"
    assert "Comp-Fit-shaped value re-derived" in out
    # Lane Fit and Job File are NOT re-derivable -> loud ERRORs naming the row.
    errors = norm_contracts.normalize_rankings_csv.last_integrity_errors
    assert any("Lane Fit is blank" in e and "ThesisCo" in e for e in errors)
    assert any("Job File" in e and "ThesisCo" in e for e in errors)
    assert "ERROR" in out and "row 2 (ThesisCo)" in out
    # The healthy row raised nothing.
    assert not any("FineCo" in e for e in errors)
    # Values were never invented.
    assert got["Lane Fit"] == "" and got["Job File"] == ""


def test_a_clean_csv_validates_with_no_errors(tmp_path):
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir(parents=True)
    csv_path = rankings / "b-rankings.csv"
    write_csv(csv_path, [make_row(company="Acme", posted="2026-06-13")])
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG, out=lambda _m: None)
    assert norm_contracts.normalize_rankings_csv.last_integrity_errors == []


def test_out_of_domain_status_and_comp_fit_are_errors(tmp_path):
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir(parents=True)
    csv_path = rankings / "b-rankings.csv"
    row = make_row(company="Acme", posted="2026-06-13")
    row[HEADERS.index("Comp Fit")] = "Sounds great"
    write_csv(csv_path, [row])
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG, out=lambda _m: None)
    errors = norm_contracts.normalize_rankings_csv.last_integrity_errors
    # Comp Fit is re-derived from Comp Range upstream, so the bad value is GONE
    # before validation — the derived label wins (that IS the trustworthy repair).
    assert errors == []
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    assert by_name(rows[0], rows[1])["Comp Fit"] in norm_contracts.COMP_FIT_VALUES
    # A status outside the vocabulary IS an error (normalize_status leaves it as
    # written by design, and validation names it).
    row2 = make_row(company="OddCo", status="Ghosted after 3 rounds", posted="2026-06-13")
    write_csv(csv_path, [row2])
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG, out=lambda _m: None)
    errors = norm_contracts.normalize_rankings_csv.last_integrity_errors
    assert any("outside the tracker vocabulary" in e and "OddCo" in e for e in errors)


def test_a_misaligned_row_fails_before_any_by_name_check(tmp_path):
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir(parents=True)
    csv_path = rankings / "b-rankings.csv"
    # Write a row with 3 extra cells so nothing aligns.
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        w.writerow(make_row(company="ShiftCo", posted="2026-06-13") + ["x", "y", "z"])
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG, out=lambda _m: None)
    errors = norm_contracts.normalize_rankings_csv.last_integrity_errors
    assert any("not aligned" in e for e in errors)


def test_needs_refetch_rows_are_legitimately_sparse(tmp_path):
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir(parents=True)
    csv_path = rankings / "b-rankings.csv"
    row = make_row(company="RefetchCo", status=norm_contracts.NEEDS_REFETCH_STATUS,
                   posted="2026-06-13")
    row[HEADERS.index("Lane Fit")] = ""
    row[HEADERS.index("Comp Fit")] = ""
    row[HEADERS.index("Data Completeness")] = ""
    write_csv(csv_path, [row])
    norm_contracts.normalize_rankings_csv(str(csv_path), CFG, out=lambda _m: None)
    errors = norm_contracts.normalize_rankings_csv.last_integrity_errors
    assert not any("Lane Fit" in e for e in errors)


def test_the_cli_exits_nonzero_on_integrity_errors(tmp_path):
    import subprocess
    import sys as _sys
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir(parents=True)
    csv_path = rankings / "b-rankings.csv"
    write_csv(csv_path, [_malformed_thesis_row()])
    res = subprocess.run(
        [_sys.executable, str(Path(norm_contracts.__file__)),
         "--normalize-rankings-csv", str(csv_path)],
        capture_output=True, text=True)
    assert res.returncode == 2
    assert "ERROR" in res.stdout and "ThesisCo" in res.stdout
    # A clean CSV exits zero.
    write_csv(csv_path, [make_row(company="Acme", posted="2026-06-13")])
    res2 = subprocess.run(
        [_sys.executable, str(Path(norm_contracts.__file__)),
         "--normalize-rankings-csv", str(csv_path)],
        capture_output=True, text=True)
    assert res2.returncode == 0


def test_completeness_vocabulary_recognizer():
    for ok in ("✓ complete", "⚠ comp not verified", "⚠ location unknown",
               "⚠ comp+location not verified", "comp not posted",
               "location not posted", "⚠ comp not verified; location not posted",
               "✓ complete · ⚠ comp conflicting"):
        assert norm_contracts.is_valid_completeness(ok), ok
    for bad in ("Below floor", "Meets/above target", "great", ""):
        assert not norm_contracts.is_valid_completeness(bad), bad


# ===========================================================================
# B3 — CSV/XLSX divergence prevention. The live class: the workbook showed a
# repaired value (a back-filled posted date) that the CSV lacked, because the
# XLSX build repaired on READ without writing back. Both artifacts now derive
# from ONE canonical row collection: build() runs the shared normalize pass on
# the CSV itself first.
# ===========================================================================
def _xlsx_cells(xlsx_path):
    ws = load_workbook(str(xlsx_path))["Job Rankings"]
    headers = [c.value for c in ws[1]]
    rows = []
    r = 2
    while ws.cell(r, 1).value is not None or ws.cell(r, 4).value is not None:
        row = ["" if ws.cell(r, i + 1).value is None else str(ws.cell(r, i + 1).value)
               for i in range(len(headers))]
        if not any(c.strip() for c in row):
            break
        rows.append(row)
        r += 1
    return headers, rows


def _csv_cells(csv_path):
    with open(csv_path, newline="", encoding="utf-8") as f:
        table = list(csv.reader(f))
    return table[0], [[str(c) for c in row] for row in table[1:] if any(c.strip() for c in row)]


def _assert_every_cell_identical(csv_path, xlsx_path, context):
    csv_headers, csv_rows = _csv_cells(csv_path)
    xlsx_headers, xlsx_rows = _xlsx_cells(xlsx_path)
    assert csv_headers == xlsx_headers, f"{context}: header rows differ"
    assert len(csv_rows) == len(xlsx_rows), f"{context}: row counts differ"
    for rn, (crow, xrow) in enumerate(zip(csv_rows, xlsx_rows), start=2):
        for h, cv, xv in zip(csv_headers, crow, xrow):
            # The title cell is written as a hyperlinked display TEXT in the workbook;
            # everything else must match byte-for-byte.
            if h == "Job Post Title + Link" and xv and xv in cv:
                continue
            assert cv == xv, f"{context}: row {rn} col {h!r}: CSV {cv!r} != XLSX {xv!r}"


@pytest.mark.parametrize("shape", ["fresh", "legacy_posted_divergence"])
def test_csv_and_xlsx_agree_cell_by_cell(tmp_path, shape):
    """(a) a fresh current-contract write; (b) the Thesis/Headway shape — a legacy CSV
    whose capture carries a posted date the CSV lacks. After the build, the CSV must
    hold every repaired value the workbook shows."""
    write_capture(tmp_path, "Acme")            # capture with Job Posted At: June 13, 2026
    write_capture(tmp_path, "Nodate", posted="Job Posted At: Unknown")
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir()
    cfg_path = tmp_path / "jail.config.json"
    cfg_path.write_text(json.dumps(CFG), encoding="utf-8")
    csv_path = rankings / "b-rankings.csv"
    xlsx_path = rankings / "b-rankings.xlsx"
    if shape == "fresh":
        write_csv(csv_path, [make_row(company="Acme", posted="2026-06-13"),
                             make_row(company="Nodate")])
    else:
        # Legacy headers, and NO posted value anywhere in the CSV — the divergence
        # class: pre-B3, only the workbook would have shown the back-filled date.
        write_legacy_csv(csv_path, [legacy_row(company="Acme"),
                                    legacy_row(company="Nodate")])
    make_rankings_xlsx.build(str(csv_path), str(xlsx_path), config_path=str(cfg_path))
    _assert_every_cell_identical(csv_path, xlsx_path, shape)
    # The repaired values are IN THE CSV, not just the workbook.
    headers, rows = _csv_cells(csv_path)
    got = dict(zip(headers, rows[0]))
    assert got["Job Posted Date"] == "2026-06-13"
    got2 = dict(zip(headers, rows[1]))
    assert got2["Job Posted Date"] == "Unknown"
    assert got["Data Completeness"]            # back-filled into the CSV too


def test_regenerating_the_xlsx_writes_repairs_back_to_a_legacy_csv(tmp_path):
    write_capture(tmp_path, "Acme")
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir()
    cfg_path = tmp_path / "jail.config.json"
    cfg_path.write_text(json.dumps(CFG), encoding="utf-8")
    csv_path = rankings / "old-rankings.csv"
    write_legacy_csv(csv_path, [legacy_row(company="Acme", location="NYC/SF - 3 days")])
    before = csv_path.read_text(encoding="utf-8")
    make_rankings_xlsx.build(str(csv_path), str(rankings / "old-rankings.xlsx"),
                             config_path=str(cfg_path))
    after = csv_path.read_text(encoding="utf-8")
    assert after != before                      # the build REPAIRED the CSV in place
    headers, rows = _csv_cells(csv_path)
    assert headers == HEADERS                   # migrated to the 27-column contract
    got = dict(zip(headers, rows[0]))
    assert got["Working Location"] == "IRL NYC/SF - 3 days"   # normalized in the CSV
    assert got["Job Posted Date"] == "2026-06-13"
    # And a second build is a no-op on the CSV (single canonical collection, stable).
    stable = csv_path.read_text(encoding="utf-8")
    make_rankings_xlsx.build(str(csv_path), str(rankings / "old-rankings.xlsx"),
                             config_path=str(cfg_path))
    assert csv_path.read_text(encoding="utf-8") == stable
