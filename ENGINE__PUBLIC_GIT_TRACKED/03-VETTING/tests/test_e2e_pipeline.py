"""B8 — the end-to-end fixture pipeline test.

Drives the REAL wiring, not reimplementations: stubbed fetch (network fully stubbed)
→ real extraction (the actual ATS fixture parsers) → canonical identity → registry →
`prep_common.process_urls` (capture writer + QA gate) → question filtering →
scoring-row synthesis through the SAME 32-column contract → the real
`norm_contracts.py --normalize-rankings-csv` CLI (subprocess) → the real
`make_rankings_xlsx.build`. Assertions land on the FINAL artifacts.

The critical property: this test FAILS if a live-path component is disconnected even
when its helper-level unit tests pass. The only stage stubbed above the helper level
is the genuinely-LLM scorer, whose CSV rows are synthesized through the same
HEADERS contract the JS writer uses (pinned against `resolve_contract_headers`).

Source shapes covered: Greenhouse (boards-API fixture), Ashby (posting-API fixture +
rendered apply-form questions), Lever (postings-API fixture through the real
`_fetch_lever` with a faked requests layer), employer-hosted Ashby (declared-name
enrichment), and a rendered generic page (JSON-LD JobPosting).
"""
import csv
import json
from datetime import date as dt_date, datetime
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

import make_rankings_xlsx
import norm_contracts

REPO = Path(__file__).resolve().parents[3]
PREP = REPO / "ENGINE__PUBLIC_GIT_TRACKED" / "02-PREP"
sys.path.insert(0, str(PREP))
sys.path.insert(0, str(PREP / "tests"))

import ats_fetchers as af  # noqa: E402
import prep_common as pc  # noqa: E402
import qa_captures as qa  # noqa: E402

FIXTURES = PREP / "tests" / "fixtures"
CFG = {
    "comp": {"currency": "USD", "target_base": 200, "floor_base": 180},
    "location": {
        "home_metro": "New York City",
        "home_metro_aliases": ["NYC", "New York", "NYC metro", "Brooklyn", "Manhattan"],
        "city_priority": ["NYC", "SF"],
    },
}


@pytest.fixture(autouse=True)
def _no_network_and_isolated_registry(tmp_path, monkeypatch):
    """The E2E must be as network-proof as the unit suite: the registry default is
    isolated and the employer-name lookup's network step is a recording no-op."""
    monkeypatch.setattr(pc, "DEFAULT_REGISTRY_PATH",
                        tmp_path / "_registry" / "capture-history-registry.json")
    monkeypatch.setattr(pc, "EMPLOYER_NAME_FETCHER",
                        lambda url, **kw: pytest.fail(
                            f"E2E made a live employer-name request for {url}"))


def _load(name):
    return json.loads((FIXTURES / name).read_text(encoding="utf-8"))


class _FakeResp:
    def __init__(self, status, payload):
        self.status_code = status
        self._payload = payload

    def json(self):
        return self._payload


def _result(meta, questions=None):
    return {"ok": True, "title": meta["title"], "company": meta["company"],
            "body": meta["text"], "method": "ats", "error": None,
            "meta": dict(meta, method="ats", structured_source=True),
            "questions": questions if questions is not None else meta.get("questions") or []}


def _build_fetchers(monkeypatch):
    """One stubbed fetch_one per source shape, each built from the REAL parser."""
    gh = af._greenhouse_job_to_result(
        _load("greenhouse_bloomerang_4705550005.json"), "bloomerang", "Bloomerang",
        "https://boards.greenhouse.io/bloomerang/jobs/4705550005")

    ashby = af._ashby_job_to_result(_load("ashby_betterup_principal_pm.json"), "betterup")
    kept = af.filter_questions(
        af.normalize_ashby_apply_fields(_load("ashby_betterup_apply_form.json")))
    ashby = dict(ashby, questions=kept)

    lever_payload = _load("lever_findem_posting.json")

    class LeverRequests:
        def get(self, url, **kw):
            return _FakeResp(200, lever_payload) if "api.lever.co" in url \
                else _FakeResp(404, {})

    monkeypatch.setattr(af, "requests", LeverRequests())
    lever = af._fetch_lever(
        "https://jobs.lever.co/findem/d1f48556-a8c9-46b7-b089-4317ec2dd280")
    assert lever and lever["source"] == "lever-postings-api"

    employer_hosted = dict(ashby)
    employer_hosted["company"] = "Helpscout"          # the board-token guess
    employer_hosted["title"] = "Lead/Principal Product Manager, Resolve"
    employer_hosted["source"] = "ashby-posting-api (custom-domain jid)"
    employer_hosted["questions"] = []

    rendered_meta = {
        "title": "Staff Product Manager", "company": "Acme Health",
        "text": ("About the role\nResponsibilities include shipping product.\n"
                 "Qualifications: 7+ years of product experience.\n"
                 "The base salary range for this role is $190,000 - $240,000 annually.\n"
                 + ("value delivered. " * 60)),
        "source": "playwright/html", "structured_source": True,
        "compensation": "$190,000 - $240,000", "working_location": "Remote, US",
        "employment_type": "FULL_TIME", "posted_date": "2026-07-01",
        "questions": [],
    }

    by_url = {
        "https://boards.greenhouse.io/bloomerang/jobs/4705550005": _result(gh),
        "https://jobs.ashbyhq.com/betterup/fa0a5d05-39f9-47d5-9fc9-0a0540ff9018":
            _result(ashby, kept),
        "https://jobs.lever.co/findem/d1f48556-a8c9-46b7-b089-4317ec2dd280":
            _result(lever),
        "https://www.helpscout.com/company/careers/?ashby_jid="
        "fa0a5d05-39f9-47d5-9fc9-0a0540ff9018": _result(employer_hosted),
        "https://acmehealth.example.com/jobs/staff-product-manager": {
            "ok": True, "title": rendered_meta["title"],
            "company": rendered_meta["company"], "body": rendered_meta["text"],
            "method": "playwright", "error": None,
            "meta": dict(rendered_meta, method="playwright"), "questions": []},
    }

    def fetch_one(url):
        return dict(by_url[url])

    return list(by_url), fetch_one


def test_the_whole_pipeline_end_to_end(tmp_path, monkeypatch):
    batch = tmp_path / "07-30-26"
    src = batch / "3 - Source Material" / "All Job Posts (full text)"
    src.mkdir(parents=True)
    reg_path = tmp_path / "registry.json"
    urls, fetch_one = _build_fetchers(monkeypatch)
    name_calls = []

    # ---- REAL process_urls: extraction results -> identity -> registry -> writer
    #      -> QA gate. One employer-name lookup is expected (Helpscout shape). ----
    manifest = pc.process_urls(
        urls, src, fetch_one, registry_path=reg_path,
        employer_name_fetcher=lambda url, **kw: (
            name_calls.append(url) or
            af.employer_declared_name(
                "<title>Lead/Principal Product Manager, Resolve – Careers at "
                "Help Scout</title>")))
    entries = {e["original_url"]: e for e in manifest["entries"]}
    assert all(e["status"] == pc.USABLE for e in entries.values()), {
        u: (e["status"], e["notes"]) for u, e in entries.items()}
    assert len(entries) == 5
    assert name_calls == [u for u in urls if "helpscout" in u]

    # Canonical filenames, incl. the employer-declared-name repair.
    files = {u: Path(e["output_path"]).name for u, e in entries.items()}
    assert files["https://boards.greenhouse.io/bloomerang/jobs/4705550005"] == \
        "bloomerang__sr-product-manager.txt"
    assert files[[u for u in urls if "helpscout" in u][0]] == \
        "help-scout__lead-principal-product-manager-resolve.txt"

    # Every capture passes the QA gate standalone too, and section order holds.
    for u, e in entries.items():
        text = (src / Path(e["output_path"]).name).read_text(encoding="utf-8")
        assert qa.validate_capture(text, filename=Path(e["output_path"]).name) == [], u
        assert text.index("JOB SNAPSHOT") < text.index("WORK DETAILS") \
            < text.index("COMPENSATION") \
            < text.index("APPLICATION QUESTIONS WORTH PREPARING") \
            < text.index("--- JOB TEXT START ---") \
            < text.index("ORIGINAL CAPTURE DETAILS")

    # Kept vs excluded questions in the FINAL artifact (Greenhouse fixture: three
    # essays kept, routine/identity/comp-expectation/EEO excluded).
    gh_text = (src / files["https://boards.greenhouse.io/bloomerang/jobs/4705550005"]
               ).read_text(encoding="utf-8")
    qsection = gh_text.split("APPLICATION QUESTIONS WORTH PREPARING")[1].split(
        "--- JOB TEXT START ---")[0]
    assert "customer outcome" in qsection.lower()
    assert "[Optional]" in qsection            # the Bloomerang essays are optional
    for excluded in ("First Name", "Requested Compensation", "authorized to work",
                     "LinkedIn", "Gender"):
        assert excluded not in qsection, excluded

    # ---- Re-fetch ONE job: registry-backed ORIGINAL/LATEST in the artifact. ----
    gh_url = "https://boards.greenhouse.io/bloomerang/jobs/4705550005"
    manifest2 = pc.process_urls([gh_url], src, fetch_one, force=True,
                                registry_path=reg_path)
    assert len(manifest2["entries"]) == 5          # partial run wiped nothing
    gh_text2 = (src / files[gh_url]).read_text(encoding="utf-8")
    assert "ORIGINAL CAPTURE DETAILS" in gh_text2
    assert "LATEST CAPTURE DETAILS" in gh_text2
    first_captured = next(l for l in gh_text.splitlines() if l.startswith("Captured At:"))
    assert first_captured in gh_text2               # original immutable in the artifact
    registry = pc.load_capture_registry(reg_path)
    assert len(registry["postings"]["greenhouse:bloomerang:4705550005"]["history"]) == 2

    # ---- Scorer boundary: synthesize rows through the SAME 32-column contract. ----
    headers = norm_contracts.resolve_contract_headers()
    assert len(headers) == 31                       # mirror-pin of the JS writer's HEADERS
    rankings = batch / "1 - Rankings"
    rankings.mkdir()
    csv_path = rankings / "e2e-rankings.csv"

    def row(company, title, wl, comp, lane, lane_fit, job_file):
        vals = {"Status? [You Change]": "Apply ASAP: High Prio", "Lane": lane,
                "Company": company,
                "Job Post Title + Link": f"{title} | https://example.com/{company.lower()}",
                "Working Location": wl, "Comp Range": comp,
                "FINAL Weighted Score": "82", "How They May See Your Profile": "80",
                "Your Desire Score": "84", "Culture Fit Score": "78",
                "Comp + Lifestyle Fit Score": "76", "Posting Last Update": "Unknown",
                "Top Reasons Notes": "r", "Top Concerns Notes": "c",
                "Profile Score Notes": "s", "Your Desire Score Notes": "m",
                "Comp + Lifestyle Fit Notes": "Cash 30/40 | Location 30/30 | Equity 16/20",
                "Lane Fit": lane_fit, "Comp Fit": "Unknown", "Data Completeness": "",
                "Job File": job_file, "Tailored? (Base Resume)": "",
                "Cover Letter Drafted?": ""}
        return [vals.get(h, "") for h in headers]

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerow(row("Betterup", "Principal Product Manager",
                       "IRL NYC/SF/Austin/DC - 2+ days", "236-296",
                       "Work - Coaching", "Outside lanes (medium)",
                       "betterup__principal-product-manager.txt"))
        w.writerow(row("Acme Health", "Staff Product Manager", "Remote", "190-240",
                       "Health - Provider Tools", "Provider Tools (high)",
                       "acme-health__staff-product-manager.txt"))

    # ---- The REAL normalize CLI pass (subprocess, exactly as vet-jobs runs it). ----
    cfg_path = tmp_path / "jail.config.json"
    cfg_path.write_text(json.dumps(CFG), encoding="utf-8")
    res = subprocess.run(
        [sys.executable, str(REPO / "ENGINE__PUBLIC_GIT_TRACKED" / "03-VETTING"
                             / "norm_contracts.py"),
         "--normalize-rankings-csv", str(csv_path), "--config", str(cfg_path)],
        capture_output=True, text=True)
    assert res.returncode == 0, res.stdout + res.stderr

    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    assert rows[0] == headers
    got_b = dict(zip(rows[0], rows[1]))
    got_a = dict(zip(rows[0], rows[2]))
    # Comp envelope: the BetterUp capture lists TWO zones; the scorer's single-band
    # 236-296 is overridden by the mechanical envelope of both.
    assert got_b["Comp Range"] == "213-296"
    assert got_b["Comp Fit"] == norm_contracts.comp_fit_label("213-296", CFG)
    # Posting Last Update back-filled from the captures (never blank).
    # Posting Last Update: neither fixture publishes an update timestamp (the Ashby
    # payload has no updatedAt, the rendered page no dateModified), so the merged
    # column falls back to the first-posted date — never JAIL's fetch date.
    assert got_b["Posting Last Update"] == "2026-03-25"
    assert got_a["Posting Last Update"] == "2026-07-01"
    # Data Completeness back-filled; Job File links to a real capture.
    assert norm_contracts.is_valid_completeness(got_b["Data Completeness"])
    assert (src / got_b["Job File"]).is_file()
    assert (src / got_a["Job File"]).is_file()

    # ---- The REAL workbook build, then cell-by-cell parity + WL hexes. ----
    xlsx_path = rankings / "e2e-rankings.xlsx"
    make_rankings_xlsx.build(str(csv_path), str(xlsx_path), config_path=str(cfg_path))
    ws = load_workbook(str(xlsx_path))["Job Rankings"]
    xlsx_headers = [c.value for c in ws[1]]
    assert xlsx_headers == headers
    with open(csv_path, newline="", encoding="utf-8") as f:
        final_rows = list(csv.reader(f))[1:]
    for rn, crow in enumerate(final_rows, start=2):
        for i, h in enumerate(headers):
            xv = ws.cell(rn, i + 1).value
            if isinstance(xv, (datetime, dt_date)):
                xv = xv.strftime("%Y-%m-%d")      # real date cells render back to ISO
            xv = "" if xv is None else str(xv)
            cv = str(crow[i])
            if h == "Job Post Title + Link" and xv and xv in cv:
                continue
            assert cv == xv, f"row {rn} col {h}: CSV {cv!r} != XLSX {xv!r}"
    wl_col = headers.index("Working Location") + 1

    def cell_hex(cell):
        rgb = cell.fill.start_color.rgb
        return str(rgb)[-6:] if rgb else None

    # "2+ days" is an OPEN-ENDED minimum -> orange per the frozen four-hex spec
    # (yellow requires an exact 1-3 day count); genuine remote -> green.
    assert cell_hex(ws.cell(2, wl_col)) == norm_contracts.WL_ORANGE
    assert cell_hex(ws.cell(3, wl_col)) == norm_contracts.WL_GREEN
