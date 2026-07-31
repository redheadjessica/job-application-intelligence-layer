"""The resume-comparison block: contract position, score invariants, artifact resolution, and
the in-place writeback's preservation guarantees.

The four columns exist to answer a question the ranking structurally cannot. `How They May See
Your Profile` is scored from the candidate's canonical profile and the posting and never opens a
resume, so nothing in the tracker said whether the resume actually being sent communicates that
fit — or whether tailoring it would change anything. These tests pin the parts of that answer
that are mechanical: the shape of the columns, the arithmetic, which file gets read, and the
promise that writing a score never disturbs anything the candidate typed.

What is deliberately NOT tested here: whether the model's judgment is good. Evidence-diversity
credit, keyword-stuffing penalties and new-content upper bounds are prompt-governed
(`00-job_application_agent.md`, Step 9.8), so the tests below assert those rules are PRESENT and
unambiguous in the spec — a drift guard, not a behavioral proof. The behavior itself is what the
four-job pilot exists to check.
"""
import csv
import subprocess
import sys
from pathlib import Path

import pytest

VETTING = Path(__file__).resolve().parent.parent
TAILOR = VETTING.parent / "04-TAILOR"
sys.path.insert(0, str(VETTING))
sys.path.insert(0, str(TAILOR))

import norm_contracts  # noqa: E402
import resume_artifacts  # noqa: E402
import update_rankings_row as urr  # noqa: E402

H_BASE = norm_contracts.H_BASE_SCORE
H_IMPROVED = norm_contracts.H_IMPROVED_SCORE
H_DELTA = norm_contracts.H_DELTA
H_WHY = norm_contracts.H_WHY_IMPROVES


# --------------------------------------------------------------------------- #
# Contract position
# --------------------------------------------------------------------------- #
def test_the_four_columns_are_appended_immediately_after_cover_letter_drafted():
    """Appended as a unit at the END, so no existing column's position moves and a user's
    left-to-right reading of the tracker is untouched."""
    headers = norm_contracts.resolve_contract_headers()
    assert headers[-5:] == [
        "Cover Letter Drafted?", H_BASE, H_IMPROVED, H_DELTA, H_WHY,
    ]
    assert len(headers) == 32


def test_the_numeric_members_exclude_the_prose_column():
    """`Why It Improves` must never be int-parsed or range-checked with the scores."""
    assert norm_contracts.RESUME_SCORE_COLUMNS == (H_BASE, H_IMPROVED, H_DELTA)
    assert H_WHY in norm_contracts.RESUME_COMPARISON_COLUMNS
    assert H_WHY not in norm_contracts.RESUME_SCORE_COLUMNS


def test_a_legacy_csv_migrates_and_gains_blank_columns_without_losing_user_data():
    """The migration joins by header NAME, so adding four columns must not shift a single
    existing value — including the ones only the candidate can supply."""
    legacy = ["Applied Date? [You Fill In]", "Status? [You Change]", "Lane", "Company",
              "Job Post Title + Link", "Working Location", "Comp Range", "Posted",
              "Have Intro? [You Add]", "Your Notes? [You Add]", "Decline/Down Date? [You Add]",
              "FINAL Weighted Score", "How They May See Your Profile", "Your Desire Score",
              "Culture Fit", "Comp + Lifestyle Fit", "Comp + Lifestyle Fit Notes",
              "Mission Fit Notes", "Scope Fit Notes", "Top Reasons Notes", "Top Concerns",
              "Job File", "Base Resume Used", "Lane Fit", "Location Fit", "Comp Fit",
              "Cover Letter?"]
    row = ["2026-07-01", "Interviewing: Onsite", "Health - Mental Health", "Acme",
           "Senior PM | https://x.test/j", "Remote", "180-200", "June 1, 2026",
           "Yes - via Dana", "MY PRIVATE NOTE", "2026-07-20",
           "81", "88", "85", "75", "68", "cash ok", "mission notes", "scope notes",
           "reasons", "concerns", "acme.txt", "Talkspace (7/1/26)", "p1 (high)",
           "Remote", "Meets/above target", "Yes"]
    rows = [legacy, row]
    norm_contracts.migrate_rankings_headers(rows)
    headers = rows[0]
    got = dict(zip(headers, rows[1]))

    for col in (H_BASE, H_IMPROVED, H_DELTA, H_WHY):
        assert col in headers, f"{col} was not inserted by the migration"
        assert got[col] == "", "a migrated row must start with an UNSCORED block, not a zero"
    # Every user-owned and pipeline-owned value survives the reshape intact.
    assert got["Applied Date? [You Fill In]"] == "2026-07-01"
    assert got["Status? [You Change]"] == "Interviewing: Onsite"
    assert got["Your Notes? [You Add]"] == "MY PRIVATE NOTE"
    assert got["Have Intro? [You Add]"] == "Yes - via Dana"
    assert got["Decline/Down Date? [You Add]"] == "2026-07-20"
    assert got["Tailored? (Base Resume)"] == "Talkspace (7/1/26)"
    assert got["Cover Letter Drafted?"] == "Yes"
    assert got["How They May See Your Profile"] == "88"


# --------------------------------------------------------------------------- #
# Score invariants
# --------------------------------------------------------------------------- #
def test_a_wholly_blank_block_is_valid_because_unscored_is_a_real_state():
    assert norm_contracts.validate_resume_comparison("", "", "", "") == []


@pytest.mark.parametrize("base,improved,delta", [(70, 85, 15), (0, 0, 0), (95, 100, 5), (40, 40, 0)])
def test_valid_blocks_pass(base, improved, delta):
    assert norm_contracts.validate_resume_comparison(base, improved, delta, "because reasons") == []


@pytest.mark.parametrize("base,improved", [(72, 85), (70, 83), (71, 84)])
def test_scores_must_be_multiples_of_five(base, improved):
    """Two model-produced scores each drift a few points run to run, so 82-vs-84 is precision
    the numbers cannot support."""
    errs = norm_contracts.validate_resume_comparison(base, improved, improved - base, "why")
    assert any("multiple of 5" in e for e in errs)


@pytest.mark.parametrize("base,improved", [(-5, 10), (10, 105), (105, 110)])
def test_scores_must_stay_within_0_100(base, improved):
    errs = norm_contracts.validate_resume_comparison(base, improved, improved - base, "why")
    assert any("outside 0-100" in e for e in errs)


def test_improved_may_never_fall_below_base():
    """The improved version is the BEST available option and keeping the base unchanged is
    always available, so a negative delta is definitionally impossible."""
    errs = norm_contracts.validate_resume_comparison(80, 70, -10, "why")
    assert any("is below" in e for e in errs)
    assert any("negative" in e for e in errs)


def test_delta_must_equal_improved_minus_base():
    errs = norm_contracts.validate_resume_comparison(70, 85, 20, "why")
    assert any("!=" in e for e in errs)


def test_a_half_written_block_is_rejected():
    """Two of three cells filled would read in the spreadsheet as a completed comparison whose
    improvement happened to be nothing."""
    errs = norm_contracts.validate_resume_comparison(70, "", "", "why")
    assert any("half-written" in e for e in errs)


def test_a_scored_row_must_carry_its_reason():
    errs = norm_contracts.validate_resume_comparison(70, 85, 15, "   ")
    assert any(H_WHY in e for e in errs)


def test_non_numeric_cells_are_reported_not_crashed():
    errs = norm_contracts.validate_resume_comparison("high", 85, 15, "why")
    assert errs and any("not a whole number" in e for e in errs)


def test_row_validation_surfaces_a_bad_block_in_a_real_table():
    headers = norm_contracts.resolve_contract_headers()
    row = [""] * len(headers)
    idx = {h: i for i, h in enumerate(headers)}
    row[idx["Company"]] = "Acme"
    row[idx["Job Post Title + Link"]] = "Senior PM | https://x.test/j"
    row[idx["Job File"]] = "acme.txt"
    row[idx[norm_contracts.H_POSTED]] = "Unknown"
    row[idx[norm_contracts.H_UPDATED]] = "Unknown"
    row[idx[H_BASE]], row[idx[H_IMPROVED]], row[idx[H_DELTA]] = "70", "60", "-10"
    row[idx[H_WHY]] = "worse somehow"
    errors = norm_contracts.validate_rankings_rows([headers, row], out=lambda *_: None)
    assert any("is below" in e for e in errors)


# --------------------------------------------------------------------------- #
# Which file actually gets scored
# --------------------------------------------------------------------------- #
def test_a_pages_base_resolves_to_its_pdf_sibling_not_the_index(tmp_path):
    """The whole point: a .pages file cannot be read, and the resume index is prose ABOUT the
    resume. Scoring must land on the document."""
    pages = tmp_path / "Jessica-Barnett-Resume - Acme - Senior PM.pages"
    pages.write_text("pages bundle")
    pdf = tmp_path / "Jessica-Barnett-Resume - Acme - Senior PM.pdf"
    pdf.write_text("%PDF-1.4")
    art = resume_artifacts.resolve_base_artifact(pages)
    assert art.status == resume_artifacts.STATUS_OK
    assert art.scoreable
    assert art.readable_path == str(pdf)
    assert art.reader == "pdf skill"


def test_a_missing_pdf_yields_no_readable_artifact_and_a_clear_reason(tmp_path):
    pages = tmp_path / "Resume - Acme.pages"
    pages.write_text("pages bundle")
    art = resume_artifacts.resolve_base_artifact(pages)
    assert art.status == resume_artifacts.STATUS_NO_PDF
    assert not art.scoreable
    assert art.readable_path is None
    assert "NOT scored" in art.note and "Export a PDF" in art.note


def test_two_ambiguous_pdfs_are_refused_rather_than_guessed(tmp_path):
    pages = tmp_path / "Resume - Acme.pages"
    pages.write_text("pages bundle")
    (tmp_path / "one.pdf").write_text("%PDF")
    (tmp_path / "two.pdf").write_text("%PDF")
    assert resume_artifacts.resolve_base_artifact(pages).status == resume_artifacts.STATUS_NO_PDF


def test_a_pdf_older_than_its_pages_is_flagged_stale_but_still_scoreable(tmp_path):
    """A stale PDF is missing the candidate's most recent manual Pages edits, so the base score
    it produces is a lower bound — usable, but the caveat has to travel with it."""
    import os
    pages = tmp_path / "Resume - Acme.pages"
    pages.write_text("pages bundle")
    pdf = tmp_path / "Resume - Acme.pdf"
    pdf.write_text("%PDF")
    os.utime(pdf, (1_700_000_000, 1_700_000_000))
    os.utime(pages, (1_700_600_000, 1_700_600_000))   # ~7 days newer
    art = resume_artifacts.resolve_base_artifact(pages)
    assert art.status == resume_artifacts.STATUS_STALE
    assert art.scoreable, "stale is a warning, not a refusal"
    assert "STALE" in art.note and "lower bound" in art.note


def test_a_fresh_export_is_not_flagged_stale(tmp_path):
    import os
    pages = tmp_path / "Resume - Acme.pages"
    pages.write_text("pages")
    pdf = tmp_path / "Resume - Acme.pdf"
    pdf.write_text("%PDF")
    os.utime(pdf, (1_700_000_000, 1_700_000_000))
    os.utime(pages, (1_700_000_030, 1_700_000_030))   # exported 30s apart
    assert resume_artifacts.resolve_base_artifact(pages).status == resume_artifacts.STATUS_OK


def test_directly_readable_formats_pass_through_with_the_right_reader(tmp_path):
    for name, reader in (("r.md", "Read tool"), ("r.txt", "Read tool"),
                         ("r.pdf", "pdf skill"), ("r.docx", "docx skill")):
        p = tmp_path / name
        p.write_text("x")
        art = resume_artifacts.resolve_base_artifact(p)
        assert art.status == resume_artifacts.STATUS_OK and art.reader == reader


def test_a_nonexistent_base_is_not_found(tmp_path):
    art = resume_artifacts.resolve_base_artifact(tmp_path / "nope.pages")
    assert art.status == resume_artifacts.STATUS_NOT_FOUND
    assert not art.scoreable


# --------------------------------------------------------------------------- #
# The in-place writeback
# --------------------------------------------------------------------------- #
def _batch(tmp_path, extra_headers=True):
    rankings = tmp_path / "1 - Rankings"
    rankings.mkdir(parents=True)
    headers = list(norm_contracts.resolve_contract_headers())
    if not extra_headers:      # simulate a pre-block file
        headers = [h for h in headers if h not in norm_contracts.RESUME_COMPARISON_COLUMNS]
    idx = {h: i for i, h in enumerate(headers)}
    row = [""] * len(headers)
    row[idx["Company"]] = "Acme"
    row[idx["Job Post Title + Link"]] = "Senior PM | https://x.test/j"
    row[idx["Job File"]] = "acme.txt"
    row[idx["Status? [You Change]"]] = "Interviewing: Onsite"
    row[idx["Your Notes? [You Add]"]] = "MY PRIVATE NOTE"
    row[idx["Applied Date? [You Fill In]"]] = "2026-07-01"
    row[idx["Tailored? (Base Resume)"]] = "Talkspace (7/1/26)"
    row[idx["Cover Letter Drafted?"]] = "Yes"
    row[idx["How They May See Your Profile"]] = "88"
    other = list(row)
    other[idx["Company"]] = "Other Co"
    other[idx["Job File"]] = "other.txt"
    other[idx["Job Post Title + Link"]] = "Lead PM | https://y.test/j"
    csv_path = rankings / "b-rankings.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows([headers, row, other])
    return tmp_path, csv_path


def _read(csv_path):
    rows = list(csv.reader(csv_path.open(newline="", encoding="utf-8")))
    return rows[0], [dict(zip(rows[0], r)) for r in rows[1:]]


def _run(batch, *args):
    return subprocess.run(
        [sys.executable, str(VETTING / "update_rankings_row.py"), "--batch", str(batch), *args],
        capture_output=True, text=True)


def test_writing_a_block_preserves_every_user_controlled_field(tmp_path):
    batch, csv_path = _batch(tmp_path)
    r = _run(batch, "--job-file", "acme.txt", "--base-score", "70",
             "--improved-score", "85", "--why", "Adds the AI evidence the base buries.")
    assert r.returncode == 0, r.stderr
    _, recs = _read(csv_path)
    acme = recs[0]
    assert (acme[H_BASE], acme[H_IMPROVED], acme[H_DELTA]) == ("70", "85", "15")
    assert acme[H_WHY] == "Adds the AI evidence the base buries."
    # Nothing the candidate owns, and nothing already scored, moved.
    assert acme["Your Notes? [You Add]"] == "MY PRIVATE NOTE"
    assert acme["Status? [You Change]"] == "Interviewing: Onsite"
    assert acme["Applied Date? [You Fill In]"] == "2026-07-01"
    assert acme["Tailored? (Base Resume)"] == "Talkspace (7/1/26)"
    assert acme["Cover Letter Drafted?"] == "Yes"
    assert acme["How They May See Your Profile"] == "88"


def test_only_the_matched_row_is_touched(tmp_path):
    batch, csv_path = _batch(tmp_path)
    _run(batch, "--job-file", "acme.txt", "--base-score", "70", "--improved-score", "85",
         "--why", "reason")
    _, recs = _read(csv_path)
    assert recs[1]["Company"] == "Other Co"
    assert recs[1][H_BASE] == "" and recs[1][H_DELTA] == ""


def test_the_delta_is_derived_and_cannot_be_supplied(tmp_path):
    """A stored delta disagreeing with its operands is the one defect a reader cannot catch by
    eye, so the CLI computes it rather than accepting it."""
    batch, _ = _batch(tmp_path)
    r = _run(batch, "--job-file", "acme.txt", "--base-score", "70", "--improved-score", "85",
             "--why", "reason", "--delta", "40")
    assert r.returncode != 0
    assert "--delta" in (r.stderr + r.stdout)


@pytest.mark.parametrize("args,expect", [
    (["--base-score", "72", "--improved-score", "85", "--why", "r"], "multiple of 5"),
    (["--base-score", "80", "--improved-score", "70", "--why", "r"], "is below"),
    (["--base-score", "70", "--improved-score", "85"], "Refusing"),
    (["--base-score", "70", "--why", "r"], "written as a unit"),
])
def test_an_invalid_block_aborts_before_touching_the_file(tmp_path, args, expect):
    batch, csv_path = _batch(tmp_path)
    before = csv_path.read_text(encoding="utf-8")
    r = _run(batch, "--job-file", "acme.txt", *args)
    assert r.returncode != 0
    assert expect in (r.stderr + r.stdout)
    assert csv_path.read_text(encoding="utf-8") == before, "file was modified despite the abort"


def test_a_file_predating_the_block_gains_the_columns_by_append(tmp_path):
    batch, csv_path = _batch(tmp_path, extra_headers=False)
    r = _run(batch, "--job-file", "acme.txt", "--base-score", "60", "--improved-score", "60",
             "--why", "The base already says everything this job needs.")
    assert r.returncode == 0, r.stderr
    headers, recs = _read(csv_path)
    assert headers[-4:] == [H_BASE, H_IMPROVED, H_DELTA, H_WHY]
    assert recs[0][H_DELTA] == "0", "a zero delta is a real answer, not a failure"
    assert recs[1]["Company"] == "Other Co" and recs[1][H_BASE] == ""


def test_the_markdown_gains_the_same_numbers_so_the_artifacts_agree(tmp_path):
    """The rankings .md is generated independently and gets no normalization pass, so it is the
    artifact that silently goes stale. The same write updates it."""
    batch, csv_path = _batch(tmp_path)
    md = csv_path.parent / "b-rankings.md"
    md.write_text("# Job Rankings\n\n## 81 — Acme: Senior PM\n\n- **File:** acme.txt\n\n"
                  "## 70 — Other Co: Lead PM\n\n- **File:** other.txt\n", encoding="utf-8")
    _run(batch, "--job-file", "acme.txt", "--base-score", "70", "--improved-score", "85",
         "--why", "Adds AI evidence.")
    text = md.read_text(encoding="utf-8")
    assert "- **Resume improvement:** base 70 → improved 85 (delta +15). Adds AI evidence." in text
    # The other job's block is untouched, and its File line still terminates its section.
    assert text.count("Resume improvement") == 1
    assert "- **File:** other.txt" in text


def test_rewriting_a_job_replaces_its_markdown_line_instead_of_stacking(tmp_path):
    batch, csv_path = _batch(tmp_path)
    md = csv_path.parent / "b-rankings.md"
    md.write_text("# Job Rankings\n\n## 81 — Acme: Senior PM\n\n- **File:** acme.txt\n\n"
                  "## 70 — Other Co: Lead PM\n\n- **File:** other.txt\n", encoding="utf-8")
    for improved in ("85", "90"):
        _run(batch, "--job-file", "acme.txt", "--base-score", "70",
             "--improved-score", improved, "--why", "Adds AI evidence.")
    text = md.read_text(encoding="utf-8")
    assert text.count("Resume improvement") == 1
    assert "improved 90 (delta +20)" in text


def test_a_hand_renamed_workbook_is_still_updated(tmp_path):
    """Real batches carry hand-renamed workbooks ('UNIFIED-55-JOB-TRACKER-FINAL.xlsx'). The old
    `*-rankings.xlsx` glob skipped those, updating the CSV while leaving the spreadsheet the
    candidate actually reads untouched."""
    openpyxl = pytest.importorskip("openpyxl")
    batch, csv_path = _batch(tmp_path)
    headers, _ = _read(csv_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    rows = list(csv.reader(csv_path.open(newline="", encoding="utf-8")))[1:]
    for r in rows:
        ws.append(r)
    xlsx = csv_path.parent / "MY-TRACKER-FINAL.xlsx"
    wb.save(xlsx)

    r = _run(batch, "--job-file", "acme.txt", "--base-score", "70", "--improved-score", "85",
             "--why", "Adds AI evidence.")
    assert r.returncode == 0, r.stderr
    assert "MY-TRACKER-FINAL.xlsx" in r.stdout

    ws2 = openpyxl.load_workbook(xlsx).active
    hdr = [c.value for c in ws2[1]]
    got = dict(zip(hdr, [c.value for c in ws2[2]]))
    assert got[H_BASE] == 70 and got[H_IMPROVED] == 85 and got[H_DELTA] == 15
    assert isinstance(got[H_DELTA], int), "scores must land as numbers so the color ramp fires"
    assert got["Your Notes? [You Add]"] == "MY PRIVATE NOTE"


def test_columns_added_to_an_existing_workbook_are_styled_like_the_sheet(tmp_path):
    """The tracker is hand-finished and must not be regenerated to gain a column, so the
    in-place path has to make the new columns look native rather than leaving bare white
    cells against a fully styled sheet."""
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    batch, csv_path = _batch(tmp_path, extra_headers=False)
    headers, _ = _read(csv_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in list(csv.reader(csv_path.open(newline="", encoding="utf-8")))[1:]:
        ws.append(r)
    thin = Side(style="thin", color="D9D9D9")
    for c in ws[1]:
        c.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"
    xlsx = csv_path.parent / "HAND-FINISHED.xlsx"
    wb.save(xlsx)

    assert _run(batch, "--job-file", "acme.txt", "--base-score", "70",
                "--improved-score", "85", "--why", "Adds AI evidence.").returncode == 0

    ws2 = openpyxl.load_workbook(xlsx).active
    hdr = [c.value for c in ws2[1]]
    for name in (H_BASE, H_IMPROVED, H_DELTA, H_WHY):
        col = hdr.index(name) + 1
        header_cell = ws2.cell(1, col)
        assert header_cell.fill.start_color.rgb.endswith("305496"), f"{name} header unstyled"
        assert header_cell.font.bold
        assert ws2.cell(2, col).alignment.wrap_text, f"{name} body cells unstyled"
    # The filter now reaches the new columns, so one-click sorting still covers the sheet.
    assert ws2.auto_filter.ref.split(":")[1].rstrip("0123456789") == \
        openpyxl.utils.get_column_letter(ws2.max_column)


# --------------------------------------------------------------------------- #
# Spec drift guards — the judgment rules live in the prompt, so pin that they are stated
# --------------------------------------------------------------------------- #
SPEC = (TAILOR / "00-job_application_agent.md").read_text(encoding="utf-8")


def test_the_spec_forbids_scoring_the_resume_index_instead_of_the_resume():
    assert "resume_artifacts.py" in SPEC
    assert "never score the resume index" in SPEC.lower() or \
           "wrong** input for *scoring*" in SPEC or "grades the description" in SPEC


def test_the_spec_requires_distinct_evidence_to_reinforce_and_repetition_not_to():
    assert "distinct evidence vs. repeated assertion" in SPEC.lower()
    assert "keyword stuffing" in SPEC.lower()


def test_the_spec_requires_disclosure_when_new_content_is_assumed():
    assert "assumes the proposed" in SPEC
    assert "contingent on work they have not done yet" in SPEC


def test_the_spec_pins_the_five_point_scale_and_the_never_below_base_invariant():
    assert "steps of five" in SPEC.lower()
    assert "improved is never below base" in SPEC.lower()


def test_the_spec_states_the_profile_score_is_not_a_ceiling():
    assert "not a ceiling" in SPEC.lower()


def test_the_spec_keeps_the_ledger_out_of_the_spreadsheet_cell():
    assert "Never paste the ledger into that cell." in SPEC
